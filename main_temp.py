from fpdf import FPDF,HTMLMixin
from os.path import join
from os import getcwd
from PIL import ImageColor
import sys
import os, glob
from jsonparser import *
import requests
import datetime as dt
from datetime import datetime
import openpyxl
from fpdf import FPDF, HTMLMixin
import matplotlib.pyplot as plt
import pandas as pd
import seaborn as sns
import numpy as np
import matplotlib.ticker as tick
from matplotlib import font_manager as fm
import matplotlib.font_manager as font_manager
import math
import matplotlib.pyplot as plt
from matplotlib.pyplot import gca

import PyPDF2


# API Setup

print(sys.argv)    
raw_data = sys.argv[1]
data_type = raw_data.split('.')


if data_type[-1] == 'xlsx' :
    # json_data = xlsx_to_dict(sys.argv[1])
    json_data = xlsx_to_dict1(sys.argv[1])
    json_data = json.loads(json_data)
    
elif data_type[-1] == 'json':
    json_data = json_to_dict(sys.argv[1])
    json_data = json.loads(json_data)
else:
    print('Wrong File')
    
url_client_credentials = 'https://account.onefin.app/connect/token'
headers_client_credentials = {'content-type': 'application/x-www-form-urlencoded', 'Accept-Charset': 'UTF-8'}
payload_client_credentials = { 'client_id':'CustomerAppClientCredentials',
            'client_secret':'e3b8e12e-c11c-22ab-c085-33b433d40522',
            'scope':'customer-login',
            'grant_type':'client_credentials'  }

try:          
    response_client_credentials = requests.post(url=url_client_credentials, headers=headers_client_credentials, data=payload_client_credentials).json()
except:
    print('token failed')

access_token = response_client_credentials['access_token']  
df = pd.DataFrame.from_dict(json_data['Name'])

MOBILE_NUMBER = str(int(df['Mobile no'][0]))
url_token = 'https://account.onefin.app/api/user/login-user'
headers_token = {'content-type': 'application/json',
                 'Accept': 'application/json',
                 'Authorization':f'Bearer {access_token}',
                 'Accept-Charset': 'UTF-8'}
raw_data = '{"user_name":"%s"}' % MOBILE_NUMBER
response_token = requests.post(url=url_token, headers=headers_token, data=raw_data).json()
user_id = response_token['data']['user_id']


url_ms_data = f'https://ms.onefin.app/UserMoneySignReport/PDF?user_code={user_id}'
response_ms_data = requests.get(url=url_ms_data).json()
user_data = response_ms_data
# for i,j in user_data.items():
#     print(i,j)
# user_name =user_data['name'].title()


# Fetching Users Data from the API 

#Money Sign Image 
user_data['image']
image_url = f'https://ms.onefin.app{user_data["image"]}'

#Money Sign text

#Money Sign description
ms_dictription = user_data['description']

#Our Sugesstion description
our_sugesstion =user_data['analBehav']

#//*----------setting of Pdf Pages---*//
# Unit conversionss
def px2MM(val):
  # Sauce: https://www.figma.com/community/plugin/841435609952260079/Unit-Converter
#   return val * (25.4 / 72)
  return val * 0.264583333338

def mm2PX(val):
  # Sauce: https://www.figma.com/community/plugin/841435609952260079/Unit-Converter
  return val * 3.7795275591

def hex2RGB(val):
  return list(ImageColor.getcolor(val, "RGB"))

def px2pts(val):
    return val*0.75
  
# class PDF(FPDF, HTMLMixin):
#     def html(self, html):
#         # Write HTML code
#         self.write_html(html)
    
  

# remove pkl files
for f in glob.glob("*.pkl"):
  os.remove(f)
  
# reportpath=os.getcwd()+'/public/money-sign-reports/'
cwd = script_dir = os.path.abspath( os.path.dirname(__file__) )
logo = join(cwd,'assets','images','logo','1FBlack.png')
logo2 = join(cwd,'assets','images','logo','1FBlackPB.png')

pdf  = FPDF('L','mm',(px2MM(1080), px2MM(1920)))

def money_sign_pdf(pdf,json_data):
    
    # pdf = FPDF()
    # pdf = FPDF('L','mm',(px2MM(1080), px2MM(1920)))
    pdf.set_auto_page_break(False)
    pdf.add_font('LeagueSpartan-SemiBold', '', join(cwd, 'assets', 'fonts', 'League_Spartan','static', 'LeagueSpartan-SemiBold.ttf'))
    pdf.add_font('LeagueSpartan-Bold', '', join(cwd, 'assets', 'fonts', 'League_Spartan','static', 'LeagueSpartan-Bold.ttf'))
    pdf.add_font('LeagueSpartan-Regular', '', join(cwd, 'assets', 'fonts', 'League_Spartan','static', 'LeagueSpartan-Regular.ttf'))
    pdf.add_font('LeagueSpartan-Medium', '', join(cwd, 'assets', 'fonts', 'League_Spartan', 'static', 'LeagueSpartan-Medium.ttf'))
    pdf.add_font('LeagueSpartan-Light', '', join(cwd, 'assets', 'fonts', 'League_Spartan', 'static', 'LeagueSpartan-Light.ttf'))
    pdf.add_font('Prata', '', join(cwd, 'assets', 'fonts', 'Prata','Prata-Regular.ttf'))

    
    # try:
    #   # df = pd.read_excel(path,sheet_name='Name')   # Getting (Name) Sheet data from the given Excel 
    #     df = pd.DataFrame.from_dict(json_data["Money Sign Description"])
    # except:
    #     return None
    
    # c_MoneyS = json_data["Money Sign Description"]["Money_sign_image"]
    # c_MoneyS = df['Money_sign_image'][0]
    c_MoneyS = user_data['moneySign'].split(' ')
    c_MoneyS = c_MoneyS[-1].strip()
    print(c_MoneyS)
    
    money_signData={
        'Eagle':{
            'Front_P':{
                'Ms_image':'Eagle.svg',
                'Vt_line':'#7C5FF2',
                'Date_c':'#C6B9FF'
            },
            'content':['#C6B9FF', '#A792FF','#7C5FF2','#5641AA'],
            'Money_Sign':['#E6E0FF','#7C5FF2','Far-Sighted Eagle'],
            #//*-behav_bias = image,color,x-axis,y-axis,width,height
            'behav_bias':['Eagle_bb.svg','#7C5FF2',837,567,1083,519,'#A792FF'],
            'gen_profile':['#5641AA','#A792FF','#7C5FF2'],
            'fin_profile':['#E6E0FF']
        },
        'Horse':{
            'Front_P':{
                'Ms_image':'Horse.svg',
                'Vt_line':'#4DC3A7',
                'Date_c':'#ACE4D7'
            },
            'content':['#ACE4D7','#82DBC6','#4DC3A7','#229479'],
            'Money_Sign':['#DEF7F1','#4DC3A7','Persistent Horse'],
            'behav_bias':['Horse_bb.svg','#82DBC6',1162,322,688,688,'#82DBC6'],
            'gen_profile':['#229479','#82DBC6','#4DC3A7'],
            'fin_profile':['#DEF7F1']
        },
        'Tiger':{
            'Front_P':{
                'Ms_image':'Tiger.svg',
                'Vt_line':'#FFCA41',
                'Date_c':'#FFE6A8'
            },
            'content':['#FFE6A8','#FFD976','#FFCA41','#D2A530'],
            'Money_Sign':['#FFF3DB','#FFCA41','Tactical Tiger'],
            'behav_bias':['Tiger_bb.svg','#FFCA41',1170,330,680,680,'#FFD976'],
            'gen_profile':['#D2A530','#FFD976','#FFCA41'],
            'fin_profile':['#FFF3DB']
        },
        'Lion':{
            'Front_P':{
                'Ms_image':'Lion.png',
                'Vt_line':'#FFCA41',
                'Date_c':'#FFE6A8'
            },
            'content':['#FFE6A8','#FFD976','#FFCA41','#D2A530'],
            'Money_Sign':['#FFF3DB','#FFCA41','Opportunistic Lion'],
            'behav_bias':['Lion_bb.svg','#FFCA41',1177,337,673,673,'#FFD976'],
            'gen_profile':['#D2A530','#FFD976','#FFCA41'],
            'fin_profile':['#DEF7F1']
        },
        'Elephant':{
            'Front_P':{
                'Ms_image':'Elephant.svg',
                'Vt_line':'#4DC3A7',
                'Date_c':'#ACE4D7'
            },
            'content':['#ACE4D7','#82DBC6','#4DC3A7','#229479'],
            'Money_Sign':['#DEF7F1','#4DC3A7','Virtuous Elephant'],
            'behav_bias':['Elephant_bb.svg','#4DC3A7',1177,377,673,673,'#82DBC6'],
            'gen_profile':['#229479','#82DBC6','#4DC3A7'],
            'fin_profile':['#DEF7F1']
        },
        'Turtle':{
            'Front_P':{
                'Ms_image':'Turtle.svg',
                'Vt_line':'#649DE5',
                'Date_c':'#ADD0FB'
            },
            'content':['#ADD0FB','#90BEF8','#649DE5','#3D7DD0'],
            'Money_Sign':['#DEEDFF','#649DE5','Vigilant Turtle'],
            'behav_bias':['Turtle_bb.svg','#649DE5',1150,310,700,700,'#90BEF8'],
            'gen_profile':['#3D7DD0','#90BEF8','#649DE5'],
            'fin_profile':['#DEEDFF']
        },
        'Whale':{
            'Front_P':{
                'Ms_image':'Whale.svg',
                'Vt_line':'#649DE5',
                'Date_c':'#ADD0FB'
            },
            'content':['#ADD0FB','#90BEF8','#649DE5','#3D7DD0'],
            'Money_Sign':['#DEEDFF','#649DE5','Enlightened Whale'],
            'behav_bias':['Whale_bb.svg','#649DE5',1177,337,673,673,'#90BEF8'],
            'gen_profile':['#3D7DD0','#90BEF8','#649DE5'],
            'fin_profile':['#DEEDFF']
        },
        'Shark':{
            'Front_P':{
                'Ms_image':'Shark.svg',
                'Vt_line':'#7C5FF2',
                'Date_c':'#C6B9FF'
            },
            'content':['#C6B9FF', '#A792FF','#7C5FF2','#5641AA'],
            'Money_Sign':['#E6E0FF','#7C5FF2','Stealthy Shark'],
            'behav_bias':['Shark_bb.svg','#7C5FF2',1170,330,680,680,'#A792FF'],
            'gen_profile':['#5641AA','#A792FF','#7C5FF2'],
            'fin_profile':['#E6E0FF']
        }
       
    }
    
    #//*----Pasing pdf_setting,Json Data, MoneySign Name,Money sign wise all Images,Backgrounds to function
    Banner(pdf,json_data,c_MoneyS,money_signData,user_data)
    content(pdf,json_data,c_MoneyS,money_signData,user_data)
    fin_profile(pdf, json_data,c_MoneyS,money_signData,user_data)
    fbs(pdf,json_data,c_MoneyS,money_signData,user_data)
    money_sign(pdf,json_data,c_MoneyS,money_signData,user_data)
    behave_bias(pdf,json_data,c_MoneyS,money_signData,user_data)
    gen_profile(pdf,json_data,c_MoneyS,money_signData)
    your_1_view_detail(pdf,json_data,c_MoneyS,money_signData,user_data)
    assets_chart(pdf,json_data,c_MoneyS,money_signData,user_data)
    liabilities_chart(pdf,json_data,c_MoneyS,money_signData,user_data)
    emergency_planning(pdf,json_data,c_MoneyS,money_signData)
    exp_lib_mang(pdf,json_data,c_MoneyS,money_signData)
    asset_allocation(pdf,json_data,c_MoneyS,money_signData)
    net_worth(pdf,json_data,c_MoneyS,money_signData)
    net_worth_projection(pdf,json_data,c_MoneyS,money_signData)
    bureao_report(pdf,json_data,c_MoneyS,money_signData)
    libility_management_1(pdf,json_data,c_MoneyS,money_signData)
    libility_management_2(pdf,json_data,c_MoneyS,money_signData)


    assumptions(pdf, json_data,c_MoneyS,money_signData)
    fin_wellness_plan(pdf, json_data,c_MoneyS,money_signData)
    cashflow_plan(pdf, json_data,c_MoneyS,money_signData)
    term_insurance(pdf,json_data,c_MoneyS,money_signData)
    health_insurance(pdf,json_data,c_MoneyS,money_signData)
    mutual_fund(pdf,json_data,c_MoneyS,money_signData)
    disclaimer(pdf,json_data,c_MoneyS,money_signData)
    lastpage(pdf,json_data,c_MoneyS,money_signData)

    
    # # directory = os.path.join(cwd,'FWR pdf',f'{names}')
    # directory = os.path.join(cwd,'FWR pdf')

    # pdf.output(directory,'F')
    
    # with open(directory, 'wb') as f:
    # # pdf.output(f'FWR pdf\\{names}')
    
    dir_name = "FWR pdf"
    if not os.path.exists(dir_name):
        os.mkdir(dir_name)
    
    names = json_data['Name'][0]['Name']+'.pdf'

    st1 = join(cwd,'FWR pdf')
    dirs = join(st1,names)
    pdf.output(dirs)
    
    # dir_name = "FWR pdf"
    # if not os.path.exists(dir_name):
    #     os.mkdir(dir_name)

    # # set the filename and filepath
    # filename = names
    # filepath = os.path.join(dir_name, filename)

    # # create a PDF file
    # pdf = PyPDF2.PdfWriter()

    # with open(filepath, 'wb') as f:
    #     pdf.write(f)
    
    
 
    

    # print(f"PDF file saved to {filepath}")






    
    
#//*------Banner-----*//
def Banner(pdf,json_data,c_MoneyS,money_signData,user_data):
    try:
      # df = pd.read_excel(path,sheet_name='Name')   # Getting (Name) Sheet data from the given Excel 
        # df = pd.DataFrame.from_dict(json_data['Name'])
        user_name = user_data['name']
    except:
        return None
    
    # pdf = FPDF('L','mm','A4')
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(0, 0, px2MM(1120), px2MM(1080), 'F')
    
    MoneyS_color = money_signData[c_MoneyS]['Front_P']['Vt_line']
    Date_c = money_signData[c_MoneyS]['Front_P']['Date_c']
    Ms_Image = money_signData[c_MoneyS]['Front_P']['Ms_image']
    
    #/**--For Money sigh right banner

    pdf.set_fill_color(*hex2RGB(MoneyS_color))
    pdf.rect(px2MM(1120), px2MM(0), px2MM(800), px2MM(1080), 'F')
 
    pdf.image(join(cwd,'assets', 'images','money_sign_png',Ms_Image),px2MM(1120), px2MM(0), px2MM(800), px2MM(1080))
 
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(0, 0, px2MM(1120), px2MM(1080), 'F')
    
    # pdf.image(join(cwd,'assets', 'images','money_sign_png',c_MoneyS+'.svg'),px2MM(1120), px2MM(12), px2MM(736), px2MM(1100))
    # pdf.rect(0, 0, px2MM(1120), px2MM(1080), 'F')
    
    #//*---1F logo--*/
    pdf.image(logo,px2MM(120), px2MM(80), px2MM(98), px2MM(113))
    
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(120))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.set_xy(px2MM(120),px2MM(333))
    pdf.multi_cell(px2MM(796), px2MM(168),'Financial\nWellness Plan')
    # pdf.multi_cell(181.6,20,'Financial Wellness\nPlan',border=0)
    
    # Test of User name and Date
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(80))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.set_xy(px2MM(120),px2MM(804))
    pdf.cell(px2MM(486), px2MM(112),user_name.title())
    # pdf.cell(px2MM(486), px2MM(112),'Dummy Data')
    
    pdf.set_font('LeagueSpartan-Light', size=px2pts(60))
    pdf.set_text_color(*hex2RGB(Date_c))
    Day=dt.datetime.now().strftime("%d")

    month=dt.datetime.now().strftime("%b")
    year=dt.datetime.now().strftime("%Y")
    if 4 <= int(Day) <= 20 or 24 <= int(Day) <= 30:
        suffix = "th"
    else:
        suffix = ["st", "nd", "rd"][int(Day) % 10 - 1]
    # pdf.set_xy(px2MM(120),px2MM(916))
    # pdf.cell(px2MM(341), px2MM(84),str(Day)+suffix+' '+str(month)+','+str(year))
    pdf.set_fill_color(*hex2RGB('#ffffff'))
    pdf.set_xy(px2MM(120),px2MM(916))
    pdf.cell(px2MM(60), px2MM(84),str(Day),border=1,align='R')
    
    d_x = pdf.get_x()
    pdf.set_font('LeagueSpartan-Light', size=px2pts(36))
    pdf.set_xy(px2MM(mm2PX(d_x)-5),px2MM(898))
    pdf.cell(px2MM(32), px2MM(84),suffix)
    
    d_x2 = pdf.get_x()
    pdf.set_font('LeagueSpartan-Light', size=px2pts(60))
    pdf.set_xy(px2MM(mm2PX(d_x2)),px2MM(916))
    pdf.cell(px2MM(100), px2MM(84),' '+str(month)+', '+str(year))
    #//*---Th suffix---*//
    
    #//*---Left Bottom Vertical Line
    pdf.set_xy(px2MM(0),px2MM(804))
    pdf.set_fill_color(*hex2RGB(MoneyS_color))
    pdf.rect(px2MM(0), px2MM(804), px2MM(20), px2MM(196), 'F')
    
    
    
# //*----Contents----*//  

def content(pdf,json_data,c_MoneyS,money_signData,user_data):
    pdf.add_page()
    # pdf.rect()
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080), 'DF')
    
    #//*--Contents banner
    pdf.set_xy(px2MM(140),px2MM(78))
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.cell(px2MM(600), px2MM(84), 'CONTENTS')
    
    #//*----- for vertical dash
    basy_y = 244
    y_gap = 160
    # h_y = 128.83
    fill_color = money_signData[c_MoneyS]['content']
    
    #//*----For Content headings and para
    # cont_rect_basey = 254
    # cont_rect_gap_y =33.91
    
    cont_head_basey = 260
    cont_head_gap_y =160
    
    cont_para_basey = 316
    cont_para_gap_y =160
    
    cont_headings = ['Your Financial Profile','Your 1 view','Your Financial Analysis','Your Financial Wellness Plan']


    htm = '''Financial behavioural Score, MoneySign<sup>TM</sup>, Genration Profile,'''
    # cont_para = ['Financial behavioural Score, MoneySignᵀᴹ, Genration Profile','Snapshot, Detailed view','Financial Metrix, Network Projection','Key Takeaways, Equity MF Suggestion, Near-term Cash Flow Plan'] 
    cont_para = ['Financial Behaviour Score, MoneySign  , Generation Profile, Life stage','Snapshot, Detailed Snapshot','Financial Metrics, Net Worth Projection',"Key takeaways, Next 3 Months' Action Plan, Financial Products Featured List"] 
    
    for i in range(4):
        pdf.set_fill_color(*hex2RGB(fill_color[i]))
        pdf.rect(px2MM(280), px2MM(basy_y), px2MM(8), px2MM(120), 'F')
        basy_y += y_gap

        #//*---Contents for each vertical dash

        pdf.set_draw_color(*hex2RGB('#000000'))
        # pdf.rect(px2MM(325.62),px2MM(cont_rect_basey),px2MM(651),px2MM(128))
        pdf.set_xy(px2MM(325.62),px2MM(cont_head_basey))
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(40))
        pdf.set_text_color(*hex2RGB('#FFFFFF'))
        pdf.set_line_width(px2MM(2))
        pdf.cell(px2MM(600), px2MM(56),cont_headings[i])


        pdf.set_xy(px2MM(325.62),px2MM(cont_para_basey))
        pdf.set_font('LeagueSpartan-Medium', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#898B90'))
        pdf.cell(px2MM(1000), px2MM(32),cont_para[i])

        cont_head_basey+=cont_head_gap_y
        cont_para_basey+=cont_para_gap_y
    
    #//*--To print superscritp TM  
    pdf.set_xy(px2MM(706), px2MM((322)))
    pdf.set_font('LeagueSpartan-Medium', size=9)
    pdf.set_text_color(*hex2RGB('#898B90'))
    pdf.cell(px2MM(16), px2MM(8), 'TM')  
      

#//*----Financial Behaviour Score----*//  

def fbs(pdf,json_data,c_MoneyS,money_signData,user_data):
    try:
        # df = pd.DataFrame.from_dict(json_data['Money Sign Description'])
        # print(json_data['Score'][0]['C_Score'])
        score = json_data['Score'][0]['C_Score']
    except:
        return None
    # score = df['Finanical_score'][0]
    #//*---Page setup
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080), 'DF')
    
    #//*--Heading vertical line
    vl_color = money_signData[c_MoneyS]['content'][2]
    pdf.set_xy(px2MM(125),px2MM(78))
    pdf.set_fill_color(*hex2RGB(vl_color))
    pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84), 'F') 
    
    #//*---heading statement
    pdf.set_xy(px2MM(120),px2MM(80))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(659), px2MM(84),'Financial Behaviour Score') 
    
    #/*--Description--*/
    txt = '''Financial Behaviour Score is a numerical representation of your financial well-being - offering an in-depth assessment of how closely your financial choices align with your personality, demography, generation, life constraints, and the macro-economic environment.'''
    pdf.set_xy(px2MM(941),px2MM(325))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(40))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.multi_cell(px2MM(860), px2MM(56),txt,align='L') 
    
    #//*----Desclamer---*//
     
    txt = '''Disclaimer: Financial Behaviour Score is part of 1 Finance's patent-pending holistic financial planning framework that is aimed at generating a wellness plan for the members to help them achieve financial well-being.'''

    pdf.set_xy(px2MM(941),px2MM(701))  
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#65676D'))
    pdf.multi_cell(px2MM(860), px2MM(32),txt,align='L') 
    
   

    #//*---Scale width is 640(excluding the curve corner) so 1%= 6.4 
    
    if score>=0 and score<=20:
        vl_x = 138+(score*5.75)
    elif score==21:
        vl_x = 272
    elif score>21 and score<=40:
        vl_x = 272+((score-20)*5.75)
    elif score==41:
        vl_x = 406
    elif score>41 and score<=60:
        vl_x = 406+((score-40)*5.75)
    elif score==61:
        vl_x = 540
    elif score>61 and score<=80:
        vl_x = 540+((score-60)*5.75)
    elif score==81:
        vl_x = 674
    elif score>81 and score<=100:
        vl_x = 674+((score-80)*5.75)
    else:
        vl_x=138+(score*6.64)

        

    # rect_x = 13 +(score*6.64)
    # text_x = 63+ (score*6.64)
    if score>=0 and score<=21:
        rect_x = 120
        text_x = 165 
    elif score>=82 and score<=100:
        rect_x = 520
        text_x = 565
    else:
        rect_x = (score*6.64)-17
        text_x = 28+ (score*6.64)    
    # vl_x = 138+(score*6.6)
    
     #//*---Score---*//
    if score>=0 and score<=20:
        pdf.image(join(cwd,'assets','images','BehaviourMeter','meter_1_20.png'),px2MM(120), px2MM(627),px2MM(700), px2MM(134))
    elif score>20 and score<=40:
        pdf.image(join(cwd,'assets','images','BehaviourMeter','meter_20_40.png'),px2MM(120), px2MM(627),px2MM(700), px2MM(134))
    elif score>40 and score<=60:
        pdf.image(join(cwd,'assets','images','BehaviourMeter','meter_40_60.png'),px2MM(120), px2MM(627),px2MM(700), px2MM(134))
    elif score>60 and score<=80:
        pdf.image(join(cwd,'assets','images','BehaviourMeter','meter_60_80.png'),px2MM(120), px2MM(627),px2MM(700), px2MM(134))
    else :
        pdf.image(join(cwd,'assets','images','BehaviourMeter','meter_80_100.png'),px2MM(120), px2MM(627),px2MM(700), px2MM(134))
    
      
    #//*---Vertical Line of Score box
    pdf.set_fill_color(*hex2RGB('#FFFFFF'))
    # pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.rect(px2MM(vl_x), px2MM(532), px2MM(13), px2MM(95), 'F') 
    
    #//*---Score Box
    pdf.set_fill_color(*hex2RGB('#FFFFFF'))
    pdf.rect(px2MM(rect_x),px2MM(284), px2MM(300), px2MM(248), 'F')
    pdf.set_xy(px2MM(text_x),px2MM(324)) 
    pdf.set_font('Prata', size=px2pts(120))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(209), px2MM(168),str(int(score)),align='C')
    
    #//*---Scale---*/
    pdf.set_xy(px2MM(120),px2MM(782)) 
    pdf.set_font('LeagueSpartan-Medium', size=px2pts(39))
    pdf.set_text_color(*hex2RGB('#898B90'))
    pdf.cell(px2MM(600), px2MM(52),'0')
    
    pdf.set_xy(px2MM(761),px2MM(782)) 
    pdf.set_font('LeagueSpartan-Medium', size=px2pts(39))
    pdf.set_text_color(*hex2RGB('#898B90'))
    pdf.cell(px2MM(59), px2MM(52),'100')

#//*----MoneySign----*//  
  
def money_sign(pdf,json_data,c_MoneyS,money_signData,user_data):
    try:
        # df = pd.DataFrame.from_dict(json_data['Money Sign Description'])
        moneySing_desc = user_data['description']
    except:
        pass
    # moneySing_desc = df["Desc"][0]
    bg_color = money_signData[c_MoneyS]['Money_Sign'][0]
    vt_line_color = money_signData[c_MoneyS]['Money_Sign'][1]
    ms_name = money_signData[c_MoneyS]['Money_Sign'][2]
    texture = c_MoneyS+'_text.svg'
    #//*---Page setup
    pdf.add_page()
    
    pdf.set_draw_color(*hex2RGB(bg_color))
    pdf.set_fill_color(*hex2RGB(bg_color))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080), 'F') 
    
    #//*----Money Sign Background-****
    pdf.rect(px2MM(0), px2MM(0), px2MM(789.4), px2MM(1080))
    pdf.image(join(cwd,'assets', 'images','MoneySign',texture),px2MM(0), px2MM(0), px2MM(789.4), px2MM(1080))
  
    pdf.rect(px2MM(0), px2MM(0), px2MM(789.4), px2MM(1080))
    pdf.image(join(cwd,'assets', 'images','MoneySign',c_MoneyS+'_overlay.png'),px2MM(0), px2MM(0), px2MM(789.4), px2MM(1080))

    #//*--Purple vertical line
    # pdf.set_xy(px2MM(125),px2MM(78))
    pdf.set_fill_color(*hex2RGB(vt_line_color))
    pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84), 'F') 
    
    #//*---heading statement
    pdf.set_xy(px2MM(120),px2MM(80))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(600), px2MM(84),pdf.write_html('MoneySign<sup>TM</sup>')) 
    
    #//*---Money Sign Logog---*//
    # pdf.rect(0, 0, px2MM(1120), px2MM(1080), 'F')
    pdf.image(join(cwd,'assets', 'images','MoneySign',c_MoneyS+'.svg'),px2MM(120), px2MM(224), px2MM(700), px2MM(700))
    
    #//*---Money Sign Name
    pdf.set_xy(px2MM(290),px2MM(924))  
    pdf.set_font('Prata', size=px2pts(42))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(400), px2MM(84),ms_name,align='C') 
    
    #//*----Description---*//

    desc = moneySing_desc.replace('<br><br>','\n')
    desc = desc.replace('<br>','\n')
    pdf.set_draw_color(*hex2RGB('#E6E0FF'))
    pdf.set_font('LeagueSpartan-Regular',size=px2pts(30))
    pdf.set_text_color(*hex2RGB('#000000'))
    # pdf.multi_cell(px2MM(860), px2MM(42),moneySing_desc,align='L')
    # if len(desc) > 1:
    #     pdf.set_xy(px2MM(940), px2MM(233))
    # else:
    # #     pdf.set_xy(px2MM(940), px2MM(330))
    # pdf.set_xy(px2MM(940), px2MM(330))
        
    # pdf.multi_cell(px2MM(860), px2MM(42),desc,align='L')
    
    # if len(desc)>1:
    #     pdf.set_xy(px2MM(940), px2MM(mm2PX(pdf.get_y())))
    #     pdf.multi_cell(px2MM(860), px2MM(42),desc[-1],align='L')
    
    if c_MoneyS=='Eagle':
        txt = """You’re a committed individual who possesses a great sense of loyalty towards your loved ones. You also have the ability to cut through the clutter and focus your time and energy on distant prospects. This enables you to assess high-potential opportunities with the big picture in mind, and strike just when the time is right. When a rewarding opportunity presents itself, your confidence, bold decision-making, and visionary mindset give you an edge over the competition."""
        pdf.set_xy(px2MM(940), px2MM(338))
        
    elif c_MoneyS=='Horse':
        txt = """Being a creature of habit, you firmly believe in following a daily routine. You are often surrounded by a large social circle and prefer companionship over solitude. You take particular joy in being present in an environment that allows you to share and absorb ideas. You have the patience and perseverance to go the extra mile in order to achieve a goal. To safeguard your long-term well-being, you tend to follow time-tested, conventional wisdom, and dislike uncertainty. In the pursuit of stability, you steer clear of situations that may be unpredictable or volatile."""
        pdf.set_xy(px2MM(940), px2MM(317))
        
    elif c_MoneyS=='Tiger':
        txt="""An acute awareness of your surroundings, combined with the ability to act with agility and swiftness, makes you perfectly capable of achieving difficult goals. You assert and maintain control over the outcome of a situation by keeping a watchful eye and never letting your guard down. You possess the optimal blend of strategic skills and emotional intelligence — you tend to be bold and risk-taking, while also being careful; you’re clever and tactful, patient and persistent, and you can also be zealous when required. You’re equipped to be self-sufficient owing to your robust decision-making, observant nature and tendency to adapt easily to situations."""
        pdf.set_xy(px2MM(940), px2MM(296))
        
    elif c_MoneyS=='Lion':
        txt="""You have a striking demeanour and a curious mind, which means you’re always on the lookout for the next big opportunity. Your formula for success has always been to fill in the gaps in your skill set by drawing on the expertise of people in your social circle. This helps you make the most of the options presented to you, even when the window of opportunity is narrow. You’ll do what it takes to emerge ahead of your competitors. Being highly self-motivated allows you to keep exploring new pursuits no matter what the outcome may be. While taking time off to celebrate a significant milestone is important to you, you will not think twice before jumping right back into action if there’s something exciting on the horizon."""
        pdf.set_xy(px2MM(940), px2MM(275))
        
    elif c_MoneyS=='Elephant':
        txt="""A natural leader, you possess a great sense of practicality and sage wisdom, which makes you capable of deriving original solutions to complicated problems. People look up to you because of your ability to maintain composure during a crisis and your unwavering focus on the end goal. You offer a sense of warmth and belonging to those around you by ensuring that they’re well looked after and aren’t threatened by external factors. A sharp intellect, great communication skills, and remarkable self-control are the most important tools in your arsenal, accompanied by a spirited attitude."""
        pdf.set_xy(px2MM(940), px2MM(317))
        
    elif c_MoneyS=='Turtle':
        txt="""You’re an incredibly composed and grounded individual with a strong sense of self-control, which comes from being satisfied with your achievements. You firmly believe in working hard to achieve your goals and follow a very systematic and disciplined approach while solving problems across different facets of your life. You dislike uncertainty because it leaves you feeling anxious, and you’re apprehensive about deviating from tried-and-tested conventions as you believe it exposes you to unwanted risks. You’re extremely skilled at negotiating difficult situations by applying a watertight system of checks and balances that you’ve established over time. You’re careful and highly observant, which allows you to anticipate undesirable outcomes and prepare in advance. You believe in taking one step at a time, and that holds you in good stead."""
        pdf.set_xy(px2MM(940), px2MM(233))
        
    elif c_MoneyS=='Whale':
        txt="""When it comes to decision-making, you’re a realist. You don’t allow irrational sentiments to cloud your judgement. You can handle the pressures of working with tight deadlines and are unperturbed by situations that are beyond your control. While you are blessed with the gift of a sharp memory and a knack to tackle complex situations, you are usually modest about your intellectual prowess. You avoid making hasty decisions and prefer to give your undivided attention to things that matter. Your easy-going attitude towards life, coupled with a sense of self-awareness, allow you to maintain a calm disposition even when provoked. And you take care of yourself by blowing off some steam once in a while."""
        pdf.set_xy(px2MM(940), px2MM(275))
        
    elif c_MoneyS=='Shark':
        txt="""You prefer to maintain an air of mystery by being discreet about your intentions, and taking measured actions. You’re highly inquisitive and aren’t afraid to venture into the unknown if the reward could prove worth it. This, combined with your strategic thought process, helps you achieve solid results. Your high levels of ambition can trigger restlessness, but also motivate you to seek prospects beyond the ordinary. You exercise autonomy while making decisions, even as you closely monitor your competitor’s movements to scout for rewarding opportunities. When a lucrative proposition is on the table, your decisions are swift and deliberate, regardless of what the outcome may be."""
        pdf.set_xy(px2MM(940), px2MM(275))
        
    pdf.multi_cell(px2MM(860), px2MM(42),desc,align='L')    
        
        
        
        
    #//*----Desclaimer---*//
    dsc = f'''Disclaimer: MoneySign    is 1 Finance's patent-pending personality assessment framework that implements one of the most scientifically validated models in psychology and helps in hyper-personalising the financial suggestions.'''

    # pdf.set_xy(px2MM(940), px2MM(777))
    desc_y = mm2PX(pdf.get_y())+35
    pdf.set_xy(px2MM(940), px2MM(mm2PX(pdf.get_y())+32))
    pdf.set_font('LeagueSpartan-Regular',size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#4B4C51'))
    pdf.multi_cell(px2MM(860), px2MM(32),dsc,align='L')
    
    #//*--To print superscritp TM  
    pdf.set_xy(px2MM(1158), px2MM(desc_y))
    pdf.set_font('LeagueSpartan-Medium', size=9)
    pdf.set_text_color(*hex2RGB('#4B4C51'))
    pdf.cell(px2MM(16), px2MM(9), 'TM')  


#//*----Behavioural Bias----*//
    
def behave_bias(pdf,json_data,c_MoneyS,money_signData,user_data):
    try:
        # df = pd.DataFrame.from_dict(json_data['Behavioural Biases'])
        df = pd.DataFrame.from_dict(user_data['behav_Description'])
    except:
        return None
    
    page_data = money_signData[c_MoneyS]['behav_bias']
    m_image = page_data[0]
    m_color = page_data[1]
    rect_color = page_data[6]
    img_x = page_data[2]
    img_y = page_data[3]
    img_w = page_data[4]
    img_4 = page_data[5]
    ini = 0
    k = 2
    if len(df)>1:
        txt2 = """We have also identified some behavioural biases that you’re likely to display while making financial decisions, and should be conscious of:"""
    elif len(df)<2:
        txt2 = """We have also identified a behavioural bias that you’re likely to display while making financial decisions, and should be conscious of:"""

    for i in range(0,len(df),2):
        #//*---Page setup
        pdf.add_page()
        pdf.set_fill_color(*hex2RGB('#FCF8ED'))
        pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F') 
        
        #//*---Cloud Images
        pdf.set_draw_color(*hex2RGB('#000000'))
        pdf.set_xy(px2MM(750), px2MM(520)) 
        # pdf.image(join(cwd,'assets', 'images','BehaviourBias','bias.png'),px2MM(750), px2MM(520), px2MM(1187.63), px2MM(570.77))
        pdf.image(join(cwd,'assets', 'images','BehaviourBias',m_image),px2MM(img_x), px2MM(img_y), px2MM(img_w), px2MM(img_4))
        
        
        #//*--Purple vertical line
        # pdf.set_xy(px2MM(125),px2MM(78))
        pdf.set_fill_color(*hex2RGB(m_color))
        pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84), 'F') 
        
        
        #//*---heading 
        pdf.set_xy(px2MM(120),px2MM(80))  
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.cell(px2MM(600), px2MM(84),'Behavioural Biases') 
        
        #//*---heading statement
        # txt2 = '''We have also identified a behavioural biases that you're likely to display while making financial decisions, and should be conscious of:'''
        pdf.set_xy(px2MM(120),px2MM(240))  
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(40))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.multi_cell(px2MM(1300), px2MM(56),txt2) 
        
        
        #//*----Content-----*//
        h_bullet = 414
        h_heading = 396
        h_para = 472
        
        gap_bullet = 284
        gap_heading = 284
        gap_para = 284
        
        for j in range(ini,k):
            
            try:
                # if df["headings"][j]:
                if df["bias_type"][j]:
                    #//* bullet
                    pdf.set_xy(px2MM(120),px2MM(h_bullet))
                    pdf.set_fill_color(*hex2RGB(rect_color))  
                    pdf.rect(px2MM(120),px2MM(h_bullet),px2MM(20),px2MM(20),'F')
                
                #//*--heading
                pdf.set_xy(px2MM(165),px2MM(h_heading))
                pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(40))
                pdf.set_text_color(*hex2RGB('#000000'))
                # pdf.multi_cell(px2MM(1300), px2MM(56),df["headings"][j])
                pdf.multi_cell(px2MM(1255), px2MM(56),df["bias_type"][j])
                
                #//*---para
                pdf.set_xy(px2MM(165),px2MM(h_para))
                pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
                pdf.set_text_color(*hex2RGB('#1A1A1D'))
                pdf.multi_cell(px2MM(1255), px2MM(42),df["description"][j],align='L')  
                
                h_bullet += gap_bullet
                h_heading += gap_heading
                h_para += gap_para
            except:
                # print(sys.exc_info())
                pass
        
        ini +=2
        k +=2 
   
#//*----Genration Profile----*//    
def gen_profile(pdf,json_data,c_MoneyS,money_signData):
    try:
        df = pd.DataFrame.from_dict(json_data["Genration"])
    except:
        return None
    

    #//*---Page setup----*//
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F') 
    
    #//*--Purple vertical line
    # pdf.set_xy(px2MM(125),px2MM(78))
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84), 'F') 
    
    #//*---heading statement
    pdf.set_xy(px2MM(120),px2MM(80))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.cell(px2MM(600), px2MM(84),'Generation Profile')
    
    gen_color = money_signData[c_MoneyS]['gen_profile'][0]
    your_profile_color = money_signData[c_MoneyS]['gen_profile'][1]
    bullet_profile_color = money_signData[c_MoneyS]['gen_profile'][2]
    
    if df["Gen Profile"][0]=='Gen 1':
        pdf.set_fill_color(*hex2RGB(gen_color))
        pdf.rect(px2MM(120), px2MM(204), px2MM(527), px2MM(915), 'F')
        pdf.image(join(cwd,'assets','images','genration profile','shade.png'),px2MM(120), px2MM(204), px2MM(527), px2MM(915))
        
        pdf.set_fill_color(*hex2RGB('#1A1A1D'))
        pdf.rect(px2MM(697), px2MM(204), px2MM(527), px2MM(915), 'F')
        
        pdf.set_fill_color(*hex2RGB('#1A1A1D'))
        pdf.rect(px2MM(1273), px2MM(204), px2MM(527), px2MM(915), 'F')
        # sq_bullet = [bullet_profile_color,'#313236','#313236']
        sq_bullet_1 = bullet_profile_color
        sq_bullet_2 = '#313236'
        sq_bullet_3 = '#313236'
        

        # highlited_color=['#1e173b','#1A1A1D','#1A1A1D'] 
    elif df["Gen Profile"][0]=='Gen 2':
        pdf.set_fill_color(*hex2RGB('#1A1A1D'))
        pdf.rect(px2MM(120), px2MM(204), px2MM(527), px2MM(915), 'F')
        
        pdf.set_fill_color(*hex2RGB(gen_color))
        pdf.rect(px2MM(697), px2MM(204), px2MM(527), px2MM(915), 'F')
        pdf.image(join(cwd,'assets','images','genration profile','shade.png'),px2MM(697), px2MM(204), px2MM(527), px2MM(915))
        
        pdf.set_fill_color(*hex2RGB('#1A1A1D'))
        pdf.rect(px2MM(1273), px2MM(204), px2MM(527), px2MM(915), 'F')
        sq_bullet = ['#313236',bullet_profile_color,'#313236']
        sq_bullet_1 = '#313236'
        sq_bullet_2 = bullet_profile_color
        sq_bullet_3 = '#313236'
        # highlited_color=['#1A1A1D','#1e173b','#1A1A1D']
    else:
        pdf.set_fill_color(*hex2RGB('#1A1A1D'))
        pdf.rect(px2MM(120), px2MM(204), px2MM(527), px2MM(915), 'F')
        
        pdf.set_fill_color(*hex2RGB('#1A1A1D'))
        pdf.rect(px2MM(697), px2MM(204), px2MM(527), px2MM(915), 'F')
        
        pdf.set_fill_color(*hex2RGB(gen_color))
        pdf.rect(px2MM(1273), px2MM(204), px2MM(527), px2MM(915), 'F')
        pdf.image(join(cwd,'assets','images','genration profile','shade.png'),px2MM(1273), px2MM(195), px2MM(527), px2MM(915))
        sq_bullet = ['#313236','#313236',bullet_profile_color]
        sq_bullet_1 = '#313236'
        sq_bullet_2 = '#313236'
        sq_bullet_3 = bullet_profile_color
        # highlited_color=['#1A1A1D','#1A1A1D','#1e173b'] 
        
    
    
    
        
    #//*---For base Rectangle---*//
    #//*-------------Card 1-----------*//
    #//*----For Heading (Genrations)---*//
    pdf.set_xy(px2MM(277),px2MM(244))  
    pdf.set_font('LeagueSpartan-Medium', size=px2pts(40))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.cell(px2MM(215), px2MM(56),'Generation 1')

    
    #//*----Personality Traits---*/
    pdf.set_xy(px2MM(160),px2MM(330))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#898B90'))
    
    # pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.cell(px2MM(447), px2MM(35),'PERSONALITY TRAITS')
    pdf.image(join(cwd,'assets','images','genration profile','shade_line.png'),px2MM(160), px2MM(372), px2MM(447), px2MM(1))
    
        
    #//*--Point 1---*//
    
    pdf.set_fill_color(*hex2RGB(sq_bullet_1))
    pdf.rect(px2MM(160), px2MM(393), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(190),px2MM(380))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Primary bread-earner in family',align='L')
    
    pdf.set_fill_color(*hex2RGB(sq_bullet_1))
    pdf.rect(px2MM(160), px2MM(446), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(190),px2MM(419))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Work hard to provide for their loved ones despite limited education',align='L') 

    
    #//*----Financial Behaviour---*/
    pdf.set_xy(px2MM(160),px2MM(523))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#898B90'))
    pdf.cell(px2MM(447), px2MM(35),'FINANCIAL BEHAVIOUR')
    pdf.image(join(cwd,'assets','images','genration profile','shade_line.png'),px2MM(160), px2MM(565), px2MM(447), px2MM(1))

    
    #//*--Point 3 to 4---*//
    pdf.set_fill_color(*hex2RGB(sq_bullet_1))
    pdf.rect(px2MM(160), px2MM(586), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(190),px2MM(575))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Earning for basic sustenance',align='L')
    
    pdf.set_fill_color(*hex2RGB(sq_bullet_1))
    pdf.rect(px2MM(160), px2MM(639), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(190),px2MM(612))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Prioritize stability and security over taking risks with their finances',align='L') 
    
    #//*----ASPIRATIONA---*/
    pdf.set_xy(px2MM(160),px2MM(716))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#898B90'))
    pdf.cell(px2MM(447), px2MM(30),'ASPIRATIONS')
    pdf.image(join(cwd,'assets','images','genration profile','shade_line.png'),px2MM(160), px2MM(758), px2MM(447), px2MM(1))
    
    #//*--Point 5 to 3---*//
    pdf.set_fill_color(*hex2RGB(sq_bullet_1))
    pdf.rect(px2MM(160), px2MM(779), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(190),px2MM(768))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Providing social security to family',align='L')
    
    pdf.set_fill_color(*hex2RGB(sq_bullet_1))
    pdf.rect(px2MM(160), px2MM(816), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(190),px2MM(805))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Giving basic lifestyle to next generation',align='L') 

    #//*----Examples of Priorities---*/
    pdf.set_xy(px2MM(160),px2MM(877))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#898B90'))
    pdf.cell(px2MM(457), px2MM(35),'EXAMPLE OF PRIORITIES')
    pdf.image(join(cwd,'assets','images','genration profile','shade_line.png'),px2MM(160), px2MM(919), px2MM(447), px2MM(1))
    
    #//*--Point 5 to 3---*//
    pdf.set_fill_color(*hex2RGB(sq_bullet_1))
    pdf.rect(px2MM(160), px2MM(956), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(190),px2MM(929))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Upgrading existing living facility to one with basic comfort and necessities')
    
    pdf.set_fill_color(*hex2RGB(sq_bullet_1))
    pdf.rect(px2MM(160), px2MM(1009), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(190),px2MM(998))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(32),'Living a dignified life in society') 
    
    
     #//*-------------Card 2-----------*//
    #//*----For Heading (Genrations)---*//
    pdf.set_xy(px2MM(850),px2MM(244))  
    pdf.set_font('LeagueSpartan-Medium', size=px2pts(40))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.cell(px2MM(215), px2MM(56),'Generation 2')

    
    #//*----Personality Traits---*/
    pdf.set_xy(px2MM(737),px2MM(330))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#898B90'))
    
    # pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.cell(px2MM(447), px2MM(35),'PERSONALITY TRAITS')
    pdf.image(join(cwd,'assets','images','genration profile','shade_line.png'),px2MM(737), px2MM(372), px2MM(447), px2MM(1))
    
        
    #//*--Point 1---*//
    
    pdf.set_fill_color(*hex2RGB(sq_bullet_2))
    pdf.rect(px2MM(737), px2MM(393), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(767),px2MM(380))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Well-educated and skilled professional',align='L')
    
    pdf.set_fill_color(*hex2RGB(sq_bullet_2))
    pdf.rect(px2MM(737), px2MM(430), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(767),px2MM(419))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Focused on improving current lifestyle',align='L') 

    
    #//*----Financial Behaviour---*/
    pdf.set_xy(px2MM(737),px2MM(481))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#898B90'))
    pdf.cell(px2MM(447), px2MM(35),'FINANCIAL BEHAVIOUR')
    pdf.image(join(cwd,'assets','images','genration profile','shade_line.png'),px2MM(737), px2MM(523), px2MM(447), px2MM(1))

    
    #//*--Point 3 to 4---*//
    pdf.set_fill_color(*hex2RGB(sq_bullet_2))
    pdf.rect(px2MM(737), px2MM(544), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(767),px2MM(533))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(425), px2MM(30),'Save mindfully to build a reasonable corpus',align='L')
    
    pdf.set_fill_color(*hex2RGB(sq_bullet_2))
    pdf.rect(px2MM(737), px2MM(613), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(767),px2MM(570))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Prefer traditional investment options such as bank deposits, mutual funds, insurance plus investment plans etc.',align='L') 
    
    #//*----ASPIRATIONA---*/
    pdf.set_xy(px2MM(737),px2MM(696))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#898B90'))
    pdf.cell(px2MM(447), px2MM(30),'ASPIRATIONS')
    pdf.image(join(cwd,'assets','images','genration profile','shade_line.png'),px2MM(737), px2MM(738), px2MM(447), px2MM(1))
    
    #//*--Point 5 to 3---*//
    pdf.set_fill_color(*hex2RGB(sq_bullet_2))
    pdf.rect(px2MM(737), px2MM(775), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(767),px2MM(748))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Providing a good lifestyle and education for future generations',align='L')
    
    pdf.set_fill_color(*hex2RGB(sq_bullet_2))
    pdf.rect(px2MM(737), px2MM(844), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(767),px2MM(817))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Achieving financial freedom to have more control over time',align='L') 

    #//*----Examples of Priorities---*/
    pdf.set_xy(px2MM(737),px2MM(911)) 
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#898B90'))
    pdf.cell(px2MM(457), px2MM(35),'EXAMPLE OF PRIORITIES')
    pdf.image(join(cwd,'assets','images','genration profile','shade_line.png'),px2MM(737), px2MM(953), px2MM(447), px2MM(1))
    
    #//*--Point 5 to 3---*//
    pdf.set_fill_color(*hex2RGB(sq_bullet_2))
    pdf.rect(px2MM(737), px2MM(974), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(767),px2MM(963))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Creating secondary source of income')
    
    pdf.set_fill_color(*hex2RGB(sq_bullet_2))
    pdf.rect(px2MM(737), px2MM(1027), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(767),px2MM(1000))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(32),'Buying a quality car and a home with good amenities')
    
    #//*----------Card 3-------*//
    #//*----For Heading (Genrations)---*//
    pdf.set_xy(px2MM(1426),px2MM(244))  
    pdf.set_font('LeagueSpartan-Medium', size=px2pts(40))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.cell(px2MM(215), px2MM(56),'Generation 3')

    
    #//*----Personality Traits---*/
    pdf.set_xy(px2MM(1313),px2MM(330))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#898B90'))
    
    # pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.cell(px2MM(447), px2MM(35),'PERSONALITY TRAITS')
    pdf.image(join(cwd,'assets','images','genration profile','shade_line.png'),px2MM(1313), px2MM(372), px2MM(447), px2MM(1))
    
        
    #//*--Point 1---*//
    
    pdf.set_fill_color(*hex2RGB(sq_bullet_3))
    pdf.rect(px2MM(1313), px2MM(409), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(1343),px2MM(380))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Early adopter of new trends and global products',align='L')
    
    pdf.set_fill_color(*hex2RGB(sq_bullet_3))
    pdf.rect(px2MM(1313), px2MM(478), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(1343),px2MM(451))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Willing to take high risks in pursuit of potential rewards',align='L') 

    
    #//*----Financial Behaviour---*/
    pdf.set_xy(px2MM(1313),px2MM(545))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#898B90'))
    pdf.cell(px2MM(447), px2MM(35),'FINANCIAL BEHAVIOUR')
    pdf.image(join(cwd,'assets','images','genration profile','shade_line.png'),px2MM(1313), px2MM(587), px2MM(447), px2MM(1))

    
    #//*--Point 3 to 4---*//
    pdf.set_fill_color(*hex2RGB(sq_bullet_3))
    pdf.rect(px2MM(1313), px2MM(608), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(1343),px2MM(597))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(425), px2MM(30),'Focused on building wealth',align='L')
    
    pdf.set_fill_color(*hex2RGB(sq_bullet_3))
    pdf.rect(px2MM(1313), px2MM(645), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(1343),px2MM(634))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Like experimenting with risky asset classes',align='L') 
    
    #//*----ASPIRATIONA---*/
    pdf.set_xy(px2MM(1313),px2MM(696))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#898B90'))
    pdf.cell(px2MM(447), px2MM(30),'ASPIRATIONS')
    pdf.image(join(cwd,'assets','images','genration profile','shade_line.png'),px2MM(1313), px2MM(738), px2MM(447), px2MM(1))
    
    #//*--Point 5 to 3---*//
    pdf.set_fill_color(*hex2RGB(sq_bullet_3))
    pdf.rect(px2MM(1313), px2MM(759), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(1343),px2MM(748))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Interested in luxury purchases',align='L')
    
    pdf.set_fill_color(*hex2RGB(sq_bullet_3))
    pdf.rect(px2MM(1313), px2MM(812), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(1343),px2MM(785))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Driven to start new businesses and pursue hobbies as a profession',align='L') 

    #//*----Examples of Priorities---*/
    pdf.set_xy(px2MM(1313),px2MM(879)) 
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#898B90'))
    pdf.cell(px2MM(457), px2MM(35),'EXAMPLE OF PRIORITIES')
    pdf.image(join(cwd,'assets','images','genration profile','shade_line.png'),px2MM(1313), px2MM(921), px2MM(447), px2MM(1))
    
    #//*--Point 5 to 3---*//
    pdf.set_fill_color(*hex2RGB(sq_bullet_3))
    pdf.rect(px2MM(1313), px2MM(958), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(1343),px2MM(931))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(30),'Growing investment portfolio by investing in alternative assets')
    
    pdf.set_fill_color(*hex2RGB(sq_bullet_3))
    pdf.rect(px2MM(1313), px2MM(1027), px2MM(10), px2MM(10),'F')
    
    pdf.set_xy(px2MM(1343),px2MM(1000))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#E9EAEE'))
    pdf.set_draw_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(417), px2MM(32),'Staying informed and educated about financial trends and new products')
    
        
    #//**----For Your Profile box---*//
    if df["Gen Profile"][0]=='Gen 1':
        pdf.set_fill_color(*hex2RGB(your_profile_color))
        pdf.rect(px2MM(120), px2MM(204), px2MM(117), px2MM(35),'F')
        pdf.set_xy(px2MM(132),px2MM(209))  
        pdf.set_font('LeagueSpartan-Medium', size=px2pts(18))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.cell(px2MM(87), px2MM(25.2),'Your Profile') 
    
    elif df["Gen Profile"][0]=='Gen 2':
        pdf.set_fill_color(*hex2RGB(your_profile_color))
        pdf.rect(px2MM(697), px2MM(204), px2MM(117), px2MM(35),'F')
        pdf.set_xy(px2MM(710),px2MM(209))  
        pdf.set_font('LeagueSpartan-Medium', size=px2pts(18))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.cell(px2MM(87), px2MM(25),'Your Profile')
        
    elif df["Gen Profile"][0]=='Gen 3':
        pdf.set_fill_color(*hex2RGB(your_profile_color))
        pdf.rect(px2MM(1273), px2MM(204), px2MM(117), px2MM(35),'F')
        pdf.set_xy(px2MM(1288),px2MM(209))  
        pdf.set_font('LeagueSpartan-Medium', size=px2pts(18))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.cell(px2MM(87), px2MM(25.2),'Your Profile')  
    else:
        pass

#//*----Net Worth------*//

def net_worth(pdf,json_data,c_MoneyS,money_signData):
    try:
        df = pd.DataFrame.from_dict(json_data["Net Worth"])
        df2 = pd.DataFrame.from_dict(json_data['val_un_adv'])
    except:
        return None
    
   
    #//*---Page setup----*//
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F') 
    
    #//*--Purple vertical line
    # pdf.set_xy(px2MM(125),px2MM(78))
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84), 'F') 
    
    #//*---heading statement
    pdf.set_xy(px2MM(120),px2MM(80))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(600), px2MM(84),'Net worth')
    
    #//*---What is Net worth
    pdf.set_xy(px2MM(400),px2MM(244))  
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(40))
    # pdf.set_text_color(*hex2RGB('#1A1A1D'))
    text2 = '''Your net worth is simply the difference between what you own (like your house, retirement funds, etc) and what you owe (your liabilities such as mortgage, credit card debt and so forth).'''
    pdf.multi_cell(px2MM(1130), px2MM(56),text2,align='C',markdown=True)
    
    #//*---Rect---*//
    
    #//*--White rect dynamic x
    white_rect = ('Net worth','Assets','Liabilities')
    white_rect_x = 140
    white_rect_x_gap = 560
    white_rect_text_x = 300
    white_rect_text_x_gap = 560
    
    #//*--Color rect dynamic x
    color_rect = ('#4DC3A7','#7C5FF2','#FFCA41')
    tot_assets = '₹ ' +"{:.2f}".format(float(df['Total Assets'][0])) +' '+ str(df['Unit1'][0])
    tot_liab = '₹ ' +"{:.2f}".format(float(df['Total Liabilities'][0])) +' '+ str(df['Unit2'][0])
    tot_networth = '₹ ' +"{:.2f}".format(float(df['Networth'][0])) +' '+ df['Unit3'][0]
    
    #//*---Total networth = Total_Assets - Total_Liabilities
    color_rect_val = (tot_networth,tot_assets,tot_liab)
    color_rect_x = 235
    color_rect_x_gap = 560
    color_rect_text_x = 275
    color_rect_text_x_gap = 560
    
    for i in range(3):
        #//*---White rectangle with text
        pdf.set_fill_color(*hex2RGB('#FFFFFF'))
        pdf.set_draw_color(*hex2RGB('#F5F5F5'))
        pdf.rect(px2MM(white_rect_x),px2MM(492),px2MM(520),px2MM(173),'FD')
        
        pdf.set_xy(px2MM(white_rect_text_x),px2MM(532))  
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(40))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.cell(px2MM(200), px2MM(56),white_rect[i],border=0,align='C')
        white_rect_x+=white_rect_x_gap
        white_rect_text_x+=white_rect_text_x_gap
        
        #//*---Color Rect with text---*//
        pdf.set_fill_color(*hex2RGB(color_rect[i]))
        pdf.rect(px2MM(color_rect_x),px2MM(618),px2MM(330),px2MM(158),'F')
        
        pdf.set_xy(px2MM(color_rect_text_x),px2MM(658))  
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(56))
        pdf.set_text_color(*hex2RGB('#FFFFFF'))
        pdf.cell(px2MM(250), px2MM(78),color_rect_val[i],border=0,align='C')
        color_rect_x+=color_rect_x_gap
        color_rect_text_x+=color_rect_text_x_gap
      
    #//*---For circle operator symbol 
    
    white_circle1_x = 639                              
    common_gap = 564 
                              
    color_circle_x = 653                                
    
    opt_x = 667.33  
    opt_val = ('=','-')
    opt_height=(13.33,3.33) 
                           
    for i in range(2):
        #//*---white outer circle---*//
        pdf.set_fill_color(*hex2RGB('#FFFFFF'))
        pdf.circle(x=px2MM(white_circle1_x),y=px2MM(539),r=px2MM(80),style='F')
        
        #//*---Color Inner circle---*//
        pdf.set_fill_color(*hex2RGB('#FCF8ED'))
        pdf.circle(x=px2MM(color_circle_x),y=px2MM(553),r=px2MM(52),style='F')
        
        #//*---For operator
        pdf.set_xy(px2MM(opt_x),px2MM(572.33))  
        pdf.set_font('LeagueSpartan-Light', size=px2pts(70))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.cell(px2MM(23.33), px2MM(opt_height[i]),opt_val[i],border=0,align='C')
        
        white_circle1_x+=common_gap
        color_circle_x+=common_gap
        opt_x+=common_gap
        
    #//*----For Value under Adivisoary---*//
    
        pdf.set_xy(px2MM(687),px2MM(892))  
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(40))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.cell(px2MM(374), px2MM(56),'Value Under Advisory:',border=0,align='L')
        
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(40))
        pdf.set_xy(px2MM(1071),px2MM(892)) 
        # pdf.cell(px2MM(374), px2MM(56),tot_networth,border=0,align='L')
        # pdf.cell(px2MM(374), px2MM(56),str(df['Total Liabilities'][0]+df['Networth'][0]),border=0,align='L')
        val_ud_adv = '₹ '+"{:.2f}".format(float(df2['Value Under Advisory'].iloc[0]))+' '+df2['Value'].iloc[0]
        # pdf.cell(px2MM(374), px2MM(56), '₹ '+"{:.2f}".format(float(df['Total Liabilities'][0])+float(df['Total Assets'][0]))+' Cr',border=0,align='L')
        pdf.cell(px2MM(374), px2MM(56),val_ud_adv,border=0,align='L')
        
        pdf.set_xy(px2MM(728.5),px2MM(968)) 
        pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
        pdf.cell(px2MM(463), px2MM(32),'This includes total of your assets and liabilities.',border=0,align='C')
        
        
#//*----Expense and Liability Management------*//

def exp_lib_mang(pdf,json_data,c_MoneyS,money_signData):
    try:
        # df = pd.DataFrame.from_dict(json_data["Financial Metrics - Expenses"])
        df = pd.DataFrame.from_dict(json_data["Exp_lib_management"])
    except:
        return None
    
   
    #//*---Page setup----*//
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F') 
    
    pdf.image(join(cwd,'assets','images','backgrounds','doubleLine.png'),px2MM(1449),px2MM(0),px2MM(471),px2MM(1080))
    
    #//*--Purple vertical line
    # pdf.set_xy(px2MM(125),px2MM(78))
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84), 'F')
    
    #//*----Purple Rectange of Heading Expense and Liability Management
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(px2MM(1310), px2MM(81), px2MM(490), px2MM(82),'F')
    
    pdf.set_xy(px2MM(1330),px2MM(101)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
    pdf.set_text_color(*hex2RGB('#ffffff'))
    pdf.cell(px2MM(450), px2MM(42),'Expense and Liability Management')
     
    #//*---heading statement
    pdf.set_xy(px2MM(120),px2MM(80))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(589), px2MM(84),'Your Financial Analysis')
    

    all_statements = df['Comments']
    #//*----6 Boxes--*//
    main_box_x =main_box_x1 = 120
    heading_label_x = heading_label_x1 = 160
    x_common_gap = x1_common_gap = 577
    score_box_x = score_box_x1 = 160
    score_x = score_x1 = 170
    ideal_range_x = ideal_range_x1 = 403
    all_stat_x = all_stat_x1 = 160
    
    
    # ideal_min = (df['Ideal'][0],df['Ideal'][1],df['Ideal'][2],df['Ideal'][3],df['Ideal'][4],df['Ideal'][5])
    ideal_min = df["Ideal Range"]
    # ideal_max = (47.8,1.0,55.6,30.4,2.8,43.1)

    for i in range(3):
        
        #//*---vor horizontol Boxes Row 1
        
        pdf.set_fill_color(*hex2RGB('#ffffff'))
        pdf.rect(px2MM(main_box_x), px2MM(204), px2MM(527), px2MM(362),'F')
        
        #//*----Box Headings----*//
        pdf.set_xy(px2MM(heading_label_x),px2MM(244)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(28))
        pdf.set_text_color(*hex2RGB('#2E3034'))
        pdf.cell(px2MM(447), px2MM(39),df["Ratios"][i],align='L')
        
        #//*----Color Score Box---*//
        # if float(df['Actual'][i])>=ideal_mean[i] and float(df['Actual'][i])<=ideal_max[i]:
        #     pdf.set_fill_color(*hex2RGB('#71EBB8'))
        # else:
        #     pdf.set_fill_color(*hex2RGB('#FF937B'))
        
        if df['Color'][i]=='green':
            pdf.set_fill_color(*hex2RGB('#71EBB8'))
        else:
            pdf.set_fill_color(*hex2RGB('#FF937B'))
            
        pdf.rect(px2MM(score_box_x), px2MM(313), px2MM(90), px2MM(52),'F')
        
        pdf.set_xy(px2MM(score_x),px2MM(318)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.cell(px2MM(70), px2MM(42),str(int(df["Actual"][i]*100))+'%',align='C')
        
        #//*-----Ideal Ranges
        pdf.set_xy(px2MM(ideal_range_x),px2MM(323)) 
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#313236'))
        pdf.cell(px2MM(204), px2MM(32),'Ideal: '+str(ideal_min[i]),align='C')
        
        #//*----Statements----*//
        pdf.set_xy(px2MM(all_stat_x),px2MM(395)) 
        pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.multi_cell(px2MM(447), px2MM(32),all_statements[i],align='L')
        
        main_box_x+= x_common_gap
        score_box_x+=x_common_gap
        score_x+=x_common_gap
        ideal_range_x+=x_common_gap
        heading_label_x+=x_common_gap
        all_stat_x+=x_common_gap
        
        
    #//*----Lower 3 boxes----*//    
    main_box_x1 = 120
    heading_label_x1 = 160
    x1_common_gap = 577
    score_box_x1 = 160
    score_x1 = 170
    ideal_range_x1 = 403
    all_stat_x1 = 160
        
    for i in range(3,6):
        
        #//*---vor horizontol Boxes Row 2
        
        pdf.set_fill_color(*hex2RGB('#ffffff'))
        pdf.rect(px2MM(main_box_x1), px2MM(616), px2MM(527), px2MM(362),'F')
        
        #//*----Box Headings----*//
        pdf.set_xy(px2MM(heading_label_x1),px2MM(656)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(28))
        pdf.set_text_color(*hex2RGB('#2E3034'))
        pdf.cell(px2MM(447), px2MM(39),df["Ratios"][i],align='L')
        
        #//*----Color Score Box---*//
        # if float(df['Actual'][i])>=ideal_mean[i] and float(df['Actual'][i])<=ideal_max[i]:
        #     pdf.set_fill_color(*hex2RGB('#71EBB8'))
        # else:
        #     pdf.set_fill_color(*hex2RGB('#FF937B'))
        
        if df['Color'][i]=='green':
            pdf.set_fill_color(*hex2RGB('#71EBB8'))
        else:
            pdf.set_fill_color(*hex2RGB('#FF937B'))
            
        pdf.rect(px2MM(score_box_x1), px2MM(715), px2MM(90), px2MM(52),'F')
        
        pdf.set_xy(px2MM(score_x1),px2MM(720)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.cell(px2MM(70), px2MM(42),str(int(df["Actual"][i]*100))+'%',align='C')
        
        #//*-----Ideal Ranges
        pdf.set_xy(px2MM(ideal_range_x1),px2MM(735)) 
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#313236'))
        pdf.cell(px2MM(204), px2MM(32),'Ideal: '+str(ideal_min[i]),align='C')
        
        #//*----Statements----*//
        pdf.set_xy(px2MM(all_stat_x1),px2MM(807)) 
        pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.multi_cell(px2MM(447), px2MM(32),all_statements[i],align='L')
        
        main_box_x1+= x1_common_gap
        score_box_x1+=x1_common_gap
        score_x1+=x1_common_gap
        ideal_range_x1+=x1_common_gap
        heading_label_x1+=x1_common_gap
        all_stat_x1+=x1_common_gap
        
    pdf.set_xy(px2MM(250),px2MM(1019)) 
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(18))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    txt = '''Disclaimer: The red shade denotes a value that falls outside of the suggested range for a given metric, while a green shade indicates a value that falls within that suggested range.'''
    pdf.cell(px2MM(1420), px2MM(21.09),txt,align='C')
       
       
#//*----Asset Alocation------*//

def asset_allocation(pdf,json_data,c_MoneyS,money_signData):
    try:
        # df = pd.DataFrame.from_dict(json_data["Financial Metrics - Asset Alloc"])
        df = pd.DataFrame.from_dict(json_data["Asset Allocation_2"])
    except:
        return None
    
   
    #//*---Page setup----*//
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F') 
    
    pdf.image(join(cwd,'assets','images','backgrounds','doubleLine.png'),px2MM(1449),px2MM(0),px2MM(471),px2MM(1080))
    
    #//*--Purple vertical line
    # pdf.set_xy(px2MM(125),px2MM(78))
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84), 'F')
    
    #//*----Black Rectange of Heading Expense and Liability Management
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(px2MM(1554), px2MM(81), px2MM(246), px2MM(82),'F')
    
    pdf.set_xy(px2MM(1574),px2MM(101)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
    pdf.set_text_color(*hex2RGB('#ffffff'))
    pdf.cell(px2MM(206), px2MM(42),'Asset Allocation')
     
    #//*---heading statement
    pdf.set_xy(px2MM(120),px2MM(80))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(589), px2MM(84),'Your Financial Analysis')

    all_statements = df['Comments']
    #//*----6 Boxes--*//
    main_box_x =main_box_x1 = 120
    heading_label_x = heading_label_x1 = 160
    x_common_gap = x1_common_gap = 577
    score_box_x = score_box_x1 = 160
    score_x = score_x1 = 170
    ideal_range_x = ideal_range_x1 = 403
    all_stat_x = all_stat_x1 = 160
    
    
    # ideal_min = (df['Ideal'][0],df['Ideal'][1],df['Ideal'][2],df['Ideal'][3],df['Ideal'][4])
    ideal_min = df['Ideal Range']
    # ideal_max = (47.8,1.0,55.6,30.4,2.8,43.1)
    for i in range(3):
        
        #//*---vor horizontol Boxes Row 1
        
        pdf.set_fill_color(*hex2RGB('#ffffff'))
        pdf.rect(px2MM(main_box_x), px2MM(204), px2MM(527), px2MM(362),'F')
        
        #//*----Box Headings----*//
        pdf.set_xy(px2MM(heading_label_x),px2MM(244)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(28))
        pdf.set_text_color(*hex2RGB('#2E3034'))
        pdf.cell(px2MM(447), px2MM(39),df["Ratios"][i],align='L')
        
        #//*----Color Score Box---*//
        # if float(df['Actual'][i])>=ideal_mean[i] and float(df['Actual'][i])<=ideal_max[i]:
        #     pdf.set_fill_color(*hex2RGB('#71EBB8'))
        # else:
        #     pdf.set_fill_color(*hex2RGB('#FF937B'))
        
        if df['Color'][i]=='green':
            pdf.set_fill_color(*hex2RGB('#71EBB8'))
        else:
            pdf.set_fill_color(*hex2RGB('#FF937B'))
            
        pdf.rect(px2MM(score_box_x), px2MM(318), px2MM(90), px2MM(52),'F')
        
        pdf.set_xy(px2MM(score_x),px2MM(322)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.cell(px2MM(70), px2MM(42),str(int(df["Actual"][i]*100))+'%',align='C')
        
        #//*-----Ideal Ranges
        pdf.set_xy(px2MM(ideal_range_x),px2MM(323)) 
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#313236'))
        pdf.cell(px2MM(204), px2MM(32),'Ideal: '+str(ideal_min[i]),align='C')
        
        #//*----Statements----*//
        pdf.set_xy(px2MM(all_stat_x),px2MM(395)) 
        pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.multi_cell(px2MM(447), px2MM(32),all_statements[i],align='L')
        
        main_box_x+= x_common_gap
        score_box_x+=x_common_gap
        score_x+=x_common_gap
        ideal_range_x+=x_common_gap
        heading_label_x+=x_common_gap
        all_stat_x+=x_common_gap
        
        
    #//*----Lower 3 boxes----*//    
    main_box_x1 = 120
    heading_label_x1 = 160
    x1_common_gap = 577
    score_box_x1 = 160
    score_x1 = 170
    ideal_range_x1 = 403
    all_stat_x1 = 160
        
    for i in range(3,5):
        
        #//*---vor horizontol Boxes Row 1
        
        pdf.set_fill_color(*hex2RGB('#ffffff'))
        pdf.rect(px2MM(main_box_x1), px2MM(616), px2MM(527), px2MM(362),'F')
        
        #//*----Box Headings----*//
        pdf.set_xy(px2MM(heading_label_x1),px2MM(656)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(28))
        pdf.set_text_color(*hex2RGB('#2E3034'))
        pdf.cell(px2MM(447), px2MM(39),df["Ratios"][i],align='L')
        
        #//*----Color Score Box---*//
        # if float(df['Actual'][i])>=ideal_mean[i] and float(df['Actual'][i])<=ideal_max[i]:
        #     pdf.set_fill_color(*hex2RGB('#71EBB8'))
        # else:
        #     pdf.set_fill_color(*hex2RGB('#FF937B'))
        
        if df['Color'][i]=='green':
            pdf.set_fill_color(*hex2RGB('#71EBB8'))
        else:
            pdf.set_fill_color(*hex2RGB('#FF937B'))
            
        pdf.rect(px2MM(score_box_x1), px2MM(725), px2MM(90), px2MM(52),'F')
        
        pdf.set_xy(px2MM(score_x1),px2MM(730)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.cell(px2MM(70), px2MM(42),str(int(df["Actual"][i]*100))+'%',align='C')
        
        #//*-----Ideal Ranges
        pdf.set_xy(px2MM(ideal_range_x1),px2MM(735)) 
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#313236'))
        pdf.cell(px2MM(204), px2MM(32),'Ideal: '+str(ideal_min[i]),align='C')
        
        #//*----Statements----*//
        pdf.set_xy(px2MM(all_stat_x1),px2MM(807)) 
        pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.multi_cell(px2MM(447), px2MM(32),all_statements[i],align='L')
        
        main_box_x1+= x1_common_gap
        score_box_x1+=x1_common_gap
        score_x1+=x1_common_gap
        ideal_range_x1+=x1_common_gap
        heading_label_x1+=x1_common_gap
        all_stat_x1+=x1_common_gap
        
    pdf.set_xy(px2MM(250),px2MM(1019)) 
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(18))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    txt = '''Disclaimer: The red shade denotes a value that falls outside of the suggested range for a given metric, while a green shade indicates a value that falls within that suggested range.'''
    pdf.cell(px2MM(1420), px2MM(21.09),txt,align='C')
             
#//*----Emergency Planning------*//

def emergency_planning(pdf,json_data,c_MoneyS,money_signData):
    try:
        # df = pd.DataFrame.from_dict(json_data["Financial Metrics - Emergency"])
        df = pd.DataFrame.from_dict(json_data["Emergence Planning"])
    except:
        return None
    
   
    #//*---Page setup----*//
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F') 
    
    pdf.image(join(cwd,'assets','images','backgrounds','doubleLine.png'),px2MM(1449),px2MM(0),px2MM(471),px2MM(1080))
    
    #//*--Purple vertical line
    # pdf.set_xy(px2MM(125),px2MM(78))
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84), 'F')
    
    #//*----Black Rectange of Heading Expense and Liability Management
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(px2MM(1499), px2MM(81), px2MM(301), px2MM(82),'F')
    
    pdf.set_xy(px2MM(1519),px2MM(101)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
    pdf.set_text_color(*hex2RGB('#ffffff'))
    pdf.cell(px2MM(206), px2MM(42),'Emergency Planning')
     
    #//*---heading statement
    pdf.set_xy(px2MM(120),px2MM(80))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(589), px2MM(84),'Your Financial Analysis')
    
    
    all_statements = df['Comments']
    #//*----6 Boxes--*//
    main_box_x =main_box_x1 = 120
    heading_label_x = heading_label_x1 = 160
    x_common_gap = x1_common_gap = 577
    score_box_x = score_box_x1 = 160
    score_x = score_x1 = 170
    ideal_range_x = ideal_range_x1 = 403
    all_stat_x = all_stat_x1 = 160
    
    
    ideal_min = df['Ideal']

    for i in range(3):
        
        #//*---vor horizontol Boxes Row 1
        
        pdf.set_fill_color(*hex2RGB('#ffffff'))
        pdf.rect(px2MM(main_box_x), px2MM(204), px2MM(527), px2MM(362),'F')
        
        #//*----Box Headings----*//
        pdf.set_xy(px2MM(heading_label_x),px2MM(244)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(28))
        pdf.set_text_color(*hex2RGB('#2E3034'))
        pdf.cell(px2MM(447), px2MM(39),df["Ratios"][i],align='L')
        
        #//*----Color Score Box---*//
        # if float(df['Actual'][i])>=ideal_mean[i] and float(df['Actual'][i])<=ideal_max[i]:
        #     pdf.set_fill_color(*hex2RGB('#71EBB8'))
        # else:
        #     pdf.set_fill_color(*hex2RGB('#FF937B'))

        if df['Color'][i]=='green':
            pdf.set_fill_color(*hex2RGB('#71EBB8'))
        else:
            pdf.set_fill_color(*hex2RGB('#FF937B'))
            
        pdf.rect(px2MM(score_box_x), px2MM(318), px2MM(90), px2MM(52),'F')
        
        pdf.set_xy(px2MM(score_x),px2MM(322)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.cell(px2MM(70), px2MM(42),str(int(df["Actual"][i]))+str(df["unit2"][i]),align='C')
        
        #//*-----Ideal Ranges
        pdf.set_xy(px2MM(ideal_range_x),px2MM(323)) 
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#313236'))
        ideal_range = str(ideal_min[i]).split('-')
        # pdf.cell(px2MM(204), px2MM(32),'Ideal: '+str(ideal_min[i])+str(df['unit1'][i]),align='C')
        pdf.cell(px2MM(204), px2MM(32),'Ideal: '+str(ideal_range[0])+str(df['unit1'][i])+' - '+str(ideal_range[-1])+str(df['unit1'][i]),align='C')
        
        #//*----Statements----*//
        pdf.set_xy(px2MM(all_stat_x),px2MM(395)) 
        pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.multi_cell(px2MM(447), px2MM(32),all_statements[i],align='L')
        
        main_box_x+= x_common_gap
        score_box_x+=x_common_gap
        score_x+=x_common_gap
        ideal_range_x+=x_common_gap
        heading_label_x+=x_common_gap
        all_stat_x+=x_common_gap
        
        
    #//*----Lower 3 boxes----*//    
    main_box_x1 = 120
    heading_label_x1 = 160
    x1_common_gap = 577
    score_box_x1 = 160
    score_x1 = 170
    ideal_range_x1 = 403
    all_stat_x1 = 160
    
    pdf.set_xy(px2MM(250),px2MM(1019)) 
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(18))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    txt = '''Disclaimer: The red shade denotes a value that falls outside of the suggested range for a given metric, while a green shade indicates a value that falls within that suggested range.'''
    pdf.cell(px2MM(1420), px2MM(21.09),txt,align='C')
    
#//*-------Assets(pIEcHART)-----*//    
def assets_chart(pdf,json_data,c_MoneyS,money_signData,user_data):
    try:
        # df = pd.DataFrame.from_dict(json_data["Snapshot of Holding - Asset"])
        df_table = pd.DataFrame.from_dict(json_data["Asset Snapshot"])
        df_pie = pd.DataFrame.from_dict(json_data["Asset Allocation"])
        print(df_pie)
    except:
        return None
    flag = False
    
    for i in range(len(df_pie['%'])):
        if df_pie['%'].iloc[i] > 0:
            flag = True
            
    if flag == False:
        return None
    
    df.fillna('-')
    df.replace('','-',inplace=True)
    #//*---Page setup----*//
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')
    
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(0, 0, px2MM(964), px2MM(1080),'F')
    
    #//*----Assets----*//
    pdf.set_xy(px2MM(120),px2MM(80)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#ffffff'))
    pdf.multi_cell(px2MM(600), px2MM(84),'Assets',align='L')
    
    #//*---Assets Date----*//
    Day=dt.datetime.now().strftime("%d %b %Y")

    pdf.set_xy(px2MM(314),px2MM(106)) 
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#ffffff'))
    pdf.cell(px2MM(197), px2MM(32),f'As on {str(Day)}',align='L')
    
    #//*---Existing Assets ----*//
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(px2MM(1424), px2MM(81), px2MM(376), px2MM(82),'F')
    
    pdf.set_xy(px2MM(1444),px2MM(101)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
    pdf.set_text_color(*hex2RGB('#ffffff'))
    ext_value = '₹ '+"{:.2f}".format(df_table['Market Value'].iloc[-1])+str(df_table['Values'].iloc[-1])
    pdf.cell(px2MM(336), px2MM(42),f'Existing Assets : {ext_value}',align='C')
    
    #//*-----Assets Table---*//
    #//*----Col1 Assets
    pdf.set_fill_color(*hex2RGB('#ffffff'))
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.rect(px2MM(690), px2MM(204), px2MM(297), px2MM(72),'F')
    
    pdf.set_xy(px2MM(710),px2MM(224)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(257), px2MM(32),'Assets',align='L')
    
    #//*----Col2 %
  
    pdf.rect(px2MM(987), px2MM(204), px2MM(100), px2MM(72),'FD')
    pdf.set_xy(px2MM(1007),px2MM(224)) 
    pdf.cell(px2MM(60), px2MM(32),'%',align='C')
    
    #//*----Col3 Assets Class
    pdf.rect(px2MM(1087), px2MM(204), px2MM(293), px2MM(72),'FD')
    
    pdf.set_xy(px2MM(1107),px2MM(224)) 
    pdf.cell(px2MM(253), px2MM(32),'Asset Class',align='L')
    
    #//*----Col4 Market Value
    pdf.rect(px2MM(1380), px2MM(204), px2MM(177), px2MM(72),'FD')
    
    pdf.set_xy(px2MM(1400),px2MM(224)) 
    pdf.cell(px2MM(137), px2MM(32),'Market Value',align='R')
    
    #//*----Col5 Monthly Investments
    pdf.rect(px2MM(1557), px2MM(204), px2MM(243), px2MM(72),'FD')
    
    pdf.set_xy(px2MM(1577),px2MM(224)) 
    pdf.cell(px2MM(203), px2MM(32),'Monthly Investment',align='R')
    
    #//*---Dynamic y axis---
    rect_y = 276
    rect_gap = 72
    state_y = 296
    state_gap = 72


    y_high = pdf.get_y()+20
    for i in range(len(df_table)-1):
        #//*-----Assets Table---*//
        #//*----Col1 Assets
        if i%2==1:
            pdf.set_fill_color(*hex2RGB('#ffffff'))
        else:
            pdf.set_fill_color(*hex2RGB('#F3F6F9'))
            
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.rect(px2MM(690), px2MM(rect_y), px2MM(297), px2MM(72),'FD')
        
        
        pdf.set_xy(px2MM(710),px2MM(state_y)) 
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.cell(px2MM(257), px2MM(32),str(df_table['Assets'][i]),align='L')
        
        #//*----Col2 %
    
        pdf.rect(px2MM(987), px2MM(rect_y), px2MM(100), px2MM(72),'FD')
        pdf.set_xy(px2MM(1007),px2MM(state_y)) 
        pdf.cell(px2MM(60), px2MM(32),"{:.0f}".format(int(df_table['%'][i]*100))+'%',align='C')
        
        #//*----Col3 Assets Class
        pdf.rect(px2MM(1087), px2MM(rect_y), px2MM(293), px2MM(72),'FD')
        
        pdf.set_xy(px2MM(1107),px2MM(state_y)) 
        pdf.cell(px2MM(253), px2MM(32),str((df_table['Asset Class'][i])),align='L')
        
        #//*----Col4 Market Value
        pdf.rect(px2MM(1380), px2MM(rect_y), px2MM(177), px2MM(72),'FD')
        
        pdf.set_xy(px2MM(1400),px2MM(state_y)) 
        if df_table['Market Value'][i] == '-':
            pdf.cell(px2MM(137), px2MM(32),'-',align='R')
        else:
            if df_table['Market Value'][i]>0:
                pdf.cell(px2MM(137), px2MM(32),'₹ '+"{:.2f}".format((df_table['Market Value'][i]))+str(df_table['Values'][i]),align='R')
            else:
                pdf.cell(px2MM(137), px2MM(32),'₹ '+"{:.2f}".format((df_table['Market Value'][i])),align='R')
        # pdf.cell(px2MM(160), px2MM(32),'₹ '+str(df_table['Market Value'][i])+str(df_table['Values'][i]),align='R')
            
        #//*----Col5 Monthly Investments
        pdf.rect(px2MM(1557), px2MM(rect_y), px2MM(243), px2MM(72),'FD')
        
        pdf.set_xy(px2MM(1577),px2MM(state_y))
        
        if df_table['Monthly Investment'][i] == '-':
            pdf.cell(px2MM(220), px2MM(32),'-',align='R')
        else:
            if df_table['Monthly Investment'][i] > 0 :
                pdf.cell(px2MM(203), px2MM(32),'₹ '+"{:.2f}".format((df_table['Monthly Investment'][i]))+str(df_table['values'][i]),align='R')
            else:
                pdf.cell(px2MM(203), px2MM(32),'₹ '+"{:.2f}".format((df_table['Monthly Investment'][i])),align='R')
        
        rect_y+=rect_gap
        state_y+=state_gap  
        y_high = pdf.get_y() 
    #//*--------------Total Last Line--------------*//
    tot_rect_y = 276+(72*(len(df_table)-1))
    tot_text_y = 288+(72*(len(df_table)-1))

    
    pdf.set_fill_color(*hex2RGB('#ffffff'))
        
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.rect(px2MM(690), px2MM(tot_rect_y), px2MM(297), px2MM(52),'F')
    
    
    pdf.set_xy(px2MM(710),px2MM(tot_text_y)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(257), px2MM(32),df_table['%'].iloc[-1],align='L')
    
    #//*----Col2 %

    pdf.rect(px2MM(987), px2MM(tot_rect_y), px2MM(100), px2MM(52),'F')
    pdf.set_xy(px2MM(1007),px2MM(tot_text_y)) 
    pdf.cell(px2MM(60), px2MM(32),'',align='C')
    
    #//*----Col3 Assets Class
    pdf.rect(px2MM(1087), px2MM(tot_rect_y), px2MM(293), px2MM(52),'F')
    
    pdf.set_xy(px2MM(1107),px2MM(tot_text_y)) 
    pdf.cell(px2MM(253), px2MM(32),'',align='L')
    
    #//*----Col4 Market Value
    pdf.rect(px2MM(1380), px2MM(tot_rect_y), px2MM(177), px2MM(52),'F')
    
    pdf.set_xy(px2MM(1400),px2MM(tot_text_y)) 

    pdf.cell(px2MM(137), px2MM(32),'₹ '+"{:.2f}".format(df_table['Market Value'].iloc[-1])+str(df_table['Values'].iloc[-1]),align='R')
        
    #//*----Col5 Monthly Investments
    pdf.rect(px2MM(1557), px2MM(tot_rect_y), px2MM(243), px2MM(52),'F')
    pdf.set_xy(px2MM(1577),px2MM(tot_text_y))
    if df_table['Monthly Investment'].iloc[-1] == 0 or df_table['Monthly Investment'].iloc[-1] =='-':
        pdf.cell(px2MM(203), px2MM(32),'₹ 0',align='R')
    else:
        pdf.cell(px2MM(203), px2MM(32),'₹ '+"{:.2f}".format(df_table['Monthly Investment'].iloc[-1])+str(df_table['values'].iloc[-1]),align='R')
        

    
        
    #//*----Donut Pie Chart---*//
    font_path = join(cwd,'assets','fonts','Prata')
    font_files = font_manager.findSystemFonts(fontpaths=font_path)
    for font_file in font_files:
        font_manager.fontManager.addfont(font_file)

# set font

    labels = df_pie["Asset Allocation"]
    sizes = df_pie['%']
    
    aut_size = list(str(x) for x in sizes)

    
    # free_colors = ['#A792FF','#82DBC6','#90BEF8','#FFC27E','#3D7DD0']
    free_colors = ['#A792FF','#82DBC6','#90BEF8','#FFC27E','#FFD976']

    colors = free_colors[0:len(df_pie)]
    df_pie['colors'] = colors
    fig, ax0 = plt.subplots(figsize=(6.8, 6.8))
    # ax = plt.pie(sizes, colors = colors, startangle=90 )
    font = {'family': 'prata','color':  'black','weight': 'normal','size': 24,}
    # font = {'color':  'black','weight': 'normal','size': 24,}
    wedges, plt_labels, junk = ax0.pie(sizes, colors = colors,startangle=90,wedgeprops = {"edgecolor" : "black",'linewidth': 2,'antialiased': True},autopct=autopct_generator(9),textprops=font)
    wed_height = [1.08,1,1.08,1,1.08,1,1.08,1]
    plt.rcParams['font.family'] = 'prata'
    
    for i in range(len(wedges)):
        wedges[i].set_radius(wed_height[i])
    
    # plt.pie(sizes, colors = colors, autopct='%1.0f%%', startangle=90,pctdistance=0.6,textprops={'size': '14'})

    centre_circle = plt.Circle((0,0),0.2,color='black')
    fig = plt.gcf()
    fig.patch.set_facecolor('black')
    fig.gca().add_artist(centre_circle)
    
    
    
    #//*-------------------For labeling----------*//
    # bbox_props = dict(boxstyle="square,pad=0.3", fc="w", ec="k", lw=0.72)
    # kw = dict(arrowprops=dict(arrowstyle="-",color='white'),zorder=0, va="center",color='white')

    # for i, p in enumerate(wedges):
    #     ang = (p.theta2 - p.theta1) / 2. + p.theta1
    #     y = np.sin(np.deg2rad(ang))
    #     x = np.cos(np.deg2rad(ang))
    #     yc = np.arcsin(y) / (np.pi / 2)
    #     horizontalalignment = {-1: "right", 1: "left"}[int(np.sign(x))]
    #     connectionstyle = f'angle,angleA=0,angleB={ang}'
    #     kw["arrowprops"].update({"connectionstyle": connectionstyle,"color":colors[i]})
    #     plt.annotate(labels[i], xy=(x, y), xytext=(1.22*np.sign(x), 1.2*y),
    #                 horizontalalignment = horizontalalignment, fontsize = 'x-large', **kw)
    
    
    
    #//*---------------------------------**----------------------*//
    # plt.show()
    plt.tight_layout()
    plt.savefig('asset_chart.png',dpi=450)
    pdf.image('asset_chart.png',px2MM(56), px2MM(195), px2MM(584), px2MM(584))
    
    #//*----Legends---*//
    
    circle_y = 783
    common_gap = 42
    text_y = 777
    
    for i in range(0,len(df_pie)):
            
        pdf.set_fill_color(*hex2RGB(df_pie['colors'].iloc[i]))
        pdf.circle(x=px2MM(165),y=px2MM(circle_y),r=px2MM(20),style='F')
        
        pdf.set_xy(px2MM(205),px2MM(text_y)) 
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#ffffff'))
        pdf.cell(px2MM(230), px2MM(32),str(df_pie["Asset Allocation"].iloc[i])+':',align='L')
        
        pdf.set_xy(px2MM(445),px2MM(text_y))
        pdf.cell(px2MM(80), px2MM(32),"{:.0f}".format(int(df_pie['%'].iloc[i]*100))+'%',align='R')
        
        #//*---Adding double gap to the next value if the current test exceeds the width
        if len(df_pie["Asset Allocation"].iloc[i])>19:
            circle_y+=common_gap
            text_y+=common_gap
            
        circle_y+=common_gap
        text_y+=common_gap
        
def autopct_generator(limit):
    def inner_autopct(pct):
        return ('%.0f' % pct)+'%' if pct > limit else ''
    return inner_autopct  
#//*-------Liabilities(pIEcHART)-----*//    
def liabilities_chart(pdf,json_data,c_MoneyS,money_signData,user_data):
    try:
        # df = pd.DataFrame.from_dict(json_data["Snapshot of Holding - Liability"])
        df_table = pd.DataFrame.from_dict(json_data["Liability Snapshot"])
        df_pie = pd.DataFrame.from_dict(json_data["Liability Allocation"])
    except:
        return None
     
    flag = False 
    for i in range(len(df_pie['%'])):
        if df_pie['%'].iloc[i] > 0:
            flag = True

                    
    if flag == False:
        print('false got')
        return None
        

    #//*---Page setup----*//
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')
    
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(0, 0, px2MM(964), px2MM(1080),'F')
    
    #//*----Snapshot of Holding - Liability----*//
    pdf.set_xy(px2MM(120),px2MM(80)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#ffffff'))
    pdf.cell(px2MM(224), px2MM(84),'Liabilities',align='L')
    
    #//*---Liabilities Date----*//
    Day=dt.datetime.now().strftime("%d %b %Y")
    pdf.set_xy(px2MM(394),px2MM(106)) 
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#ffffff'))
    pdf.cell(px2MM(197), px2MM(32),f'As on {str(Day)}',align='L')
    
    #//*---Existing Liabilities ----*//
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(px2MM(1385), px2MM(81), px2MM(415), px2MM(82),'F')
    
    pdf.set_xy(px2MM(1405),px2MM(101)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
    pdf.set_text_color(*hex2RGB('#ffffff'))
    pdf.cell(px2MM(375), px2MM(42),'Existing Liabilities :'+' ₹ '+str(df_table['Outstanding Amount'].iloc[-1])+str(df_table['Value'].iloc[0]),align='C')
    
    #//*-----Assets Table---*//
    #//*----Col1 Liabilities
    pdf.set_fill_color(*hex2RGB('#ffffff'))
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.rect(px2MM(690), px2MM(256), px2MM(350), px2MM(72),'F')
    
    pdf.set_xy(px2MM(714),px2MM(276)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(310), px2MM(32),'Liabilities',align='L')
    
    #//*----Col2 Pending Months
  
    pdf.rect(px2MM(1040), px2MM(256), px2MM(220), px2MM(72),'FD')
    pdf.set_xy(px2MM(1060),px2MM(276)) 
    pdf.cell(px2MM(180), px2MM(32),'Pending Months',align='C')
    
    #//*----Col3 Outstanding Balance
    pdf.rect(px2MM(1260), px2MM(256), px2MM(270), px2MM(72),'FD')
    
    pdf.set_xy(px2MM(1280),px2MM(276)) 
    pdf.cell(px2MM(230), px2MM(32),'Outstanding Balance',align='R')
    
    #//*----Col4 Monthly EMI Amt.
    pdf.rect(px2MM(1530), px2MM(256), px2MM(270), px2MM(72),'FD')
    
    pdf.set_xy(px2MM(1550),px2MM(276)) 
    pdf.cell(px2MM(230), px2MM(32),'Monthly EMI Amt.',align='R')
    
    
    #//*---Dynamic y axis---
    rect_y = 328
    rect_gap = 72
    state_y = 348
    state_gap = 72    
    for i in range(len(df_table)-1):

        if i%2==1:
            pdf.set_fill_color(*hex2RGB('#ffffff'))
        else:
            pdf.set_fill_color(*hex2RGB('#F3F6F9'))
            
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.rect(px2MM(690), px2MM(rect_y), px2MM(350), px2MM(72),'FD')
        
        
        pdf.set_xy(px2MM(710),px2MM(state_y)) 
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.cell(px2MM(310), px2MM(32),str(df_table['Liability'][i]),align='L')
        
        #//*----Col2 %
    
        pdf.rect(px2MM(1040), px2MM(rect_y), px2MM(220), px2MM(72),'FD')
        pdf.set_xy(px2MM(1060),px2MM(state_y)) 
        if df_table['Pending Tenure (Months)'][i] == '' or df_table['Pending Tenure (Months)'][i] == '0':
            pdf.cell(px2MM(180), px2MM(32),' ',align='C')
        else:
            pdf.cell(px2MM(180), px2MM(32),"{:.0f}".format(df_table['Pending Tenure (Months)'][i]),align='C')
        
        #//*----Col3 Assets Class
        pdf.rect(px2MM(1260), px2MM(rect_y), px2MM(270), px2MM(72),'FD')
        
        pdf.set_xy(px2MM(1280),px2MM(state_y)) 
        if df_table['Outstanding Amount'][i] == 0 or df_table['Outstanding Amount'][i]=='-':
            pdf.cell(px2MM(230), px2MM(32),' ',align='R')
        else:
            pdf.cell(px2MM(230), px2MM(32),'₹ '+"{:.2f}".format((df_table['Outstanding Amount'][i]))+str(df_table['Value'][i]),align='R')
        
        #//*----Col4 Market Value
        pdf.rect(px2MM(1530), px2MM(rect_y), px2MM(270), px2MM(72),'FD')
        pdf.set_xy(px2MM(1550),px2MM(state_y)) 
        if df_table['Monthly EMI Amount'][i] ==0 or df_table['Monthly EMI Amount'][i]=='-':
            pdf.cell(px2MM(230), px2MM(32),' ',align='R')
        else:
            pdf.cell(px2MM(230), px2MM(32),'₹ '+"{:.2f}".format((df_table['Monthly EMI Amount'][i]))+str(df_table['Values'][i]),align='R')

            
      
     
        rect_y+=rect_gap
        state_y+=state_gap
    
    #//*--------------Total Last Line--------------*//
    tot_rect_y = 328+(72*(len(df_table)-1))
    tot_text_y = 338+(72*(len(df_table)-1))

    #//*---Col 1
    pdf.set_fill_color(*hex2RGB('#ffffff'))
        
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.rect(px2MM(690), px2MM(tot_rect_y), px2MM(350), px2MM(52),'F')
    
    
    pdf.set_xy(px2MM(710),px2MM(tot_text_y)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(310), px2MM(32),str(df_table['Liability'].iloc[-1]),align='L')
    
    #//*----Col2 %

    pdf.rect(px2MM(1040), px2MM(tot_rect_y), px2MM(220), px2MM(52),'F')
    pdf.set_xy(px2MM(1060),px2MM(tot_text_y)) 
    # pdf.cell(px2MM(160), px2MM(32),"{:.0f}".format(df_table['Pending Tenure (Months)'].iloc[-1]),align='C')
    pdf.cell(px2MM(180), px2MM(32),'',align='C')
    
    #//*----Col3 Assets Class
    pdf.rect(px2MM(1260), px2MM(tot_rect_y), px2MM(270), px2MM(52),'F')
    
    pdf.set_xy(px2MM(1280),px2MM(tot_text_y)) 
    # oa = "{:.1f}".format(float(df_table['Outstanding Amount'].iloc[-1]))
    pdf.cell(px2MM(230), px2MM(32),'₹ '+"{:.2f}".format((df_table['Outstanding Amount'].iloc[-1]))+str(df_table['Value'].iloc[-1]),align='R')
    # pdf.cell(px2MM(230), px2MM(32),oa+str(df_table['Value'].iloc[-1]),align='R')
    
    #//*----Col4 Market Value
    pdf.rect(px2MM(1530), px2MM(tot_rect_y), px2MM(270), px2MM(52),'F')
    pdf.set_xy(px2MM(1550),px2MM(tot_text_y)) 
    pdf.cell(px2MM(230), px2MM(32),'₹ '+"{:.2f}".format((df_table['Monthly EMI Amount'].iloc[-1]))+str(df_table['Values'].iloc[-1]),align='R')

    # #//*----Donut Pie Chart---*//
    
    # df1 = pd.DataFrame()
    # df1['Liability'] = df['Liability']
    # df1['Pending Tenure (Months)']=df['Pending Tenure (Months)']

    # df2 = pd.DataFrame()

    # df1['sum'] = df1['Liability'].map(df1.groupby('Liability')['Pending Tenure (Months)'].sum())
    # df2 = df1.drop_duplicates(subset = "sum")

    # df2.drop(df2.loc[df2['sum']=='--'].index, inplace=True)

    # labels = df2['Liability'] 
    # sizes = df2['sum']
    # labels = ['Good Liabilities','Bad Liabilities:'] 
    # sizes = [9,91]
    labels = df_pie['Liability Allocation']
    sizes = df_pie['%']
    # df2 = df2[df2.sum != '--']

    free_colors = ['#FFD976','#ffffff','#A792FF','#82DBC6','#90BEF8','#FFC27E','#FFD976','#3D7DD0']
    colors = free_colors[0:len(labels)]
    
    fig, ax0 = plt.subplots(figsize=(6.8, 6.8))
    # plt.subplots(figsize=(6.8, 6.8))
    # ax = plt.pie(sizes, colors = colors, startangle=90 )
    # wedges,texts = ax0.pie(sizes, colors = colors, startangle=90,wedgeprops = {"edgecolor" : "black",'linewidth': 2,'antialiased': True})
    wed_height = [1,0.9,1,1.08,1,1.08,1,1.08,1]
     # ax = plt.pie(sizes, colors = colors, startangle=90 )
    font = {'family': 'prata','color':  'black','weight': 'normal','size': 24,}
    wedges, plt_labels, junk = ax0.pie(sizes, colors = colors,startangle=90,wedgeprops = {"edgecolor" : "black",'linewidth': 2,'antialiased': True},autopct=autopct_generator(9),textprops=font)
    
    for i in range(len(wedges)):
        wedges[i].set_radius(wed_height[i])
        
        
    
    # explode = free_explode[0:len(labels)]
    # ax = plt.pie(sizes, colors = colors, startangle=90)
    centre_circle = plt.Circle((0,0),0.2,color='black')
    fig = plt.gcf()
    fig.patch.set_facecolor('black')
    fig.gca().add_artist(centre_circle)
    plt.tight_layout()
    plt.savefig('liabilities_chart.png',dpi=450)

    pdf.image('liabilities_chart.png',px2MM(56), px2MM(195), px2MM(584), px2MM(584))
    
    #//*----Legends---*//
    circle_y = 803
    common_gap = 42
    text_y = 797

    for i in range(0,len(labels)):
        pdf.set_fill_color(*hex2RGB(colors[i]))
        pdf.circle(x=px2MM(165),y=px2MM(circle_y),r=px2MM(20),style='F')
        
        pdf.set_xy(px2MM(205),px2MM(text_y)) 
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#ffffff'))
        pdf.cell(px2MM(400), px2MM(32),labels[i]+':',align='L')
        
        pdf.set_xy(px2MM(404),px2MM(text_y))
        pdf.cell(px2MM(56), px2MM(32),str(int(sizes[i]*100))+'%',align='R')
        
        circle_y+=common_gap
        text_y+=common_gap
    
    text1 = '''Good liabilities generally are those that serve a useful purpose, have favourable interest rates, reasonable repayment terms, and a high likelihood of successful repayment, while bad liabilities commonly are incurred for non-essential expenses, have unfavourable terms, or carry a high risk of default.'''    
    pdf.set_xy(px2MM(120),px2MM(959)) 
    pdf.set_font('LeagueSpartan-Light', size=px2pts(18))
    pdf.set_text_color(*hex2RGB('#ffffff'))
    pdf.multi_cell(px2MM(774), px2MM(25),text1,align='L')
        
#//*-------Net Worth Projection-----*//    
def net_worth_projection(pdf,json_data,c_MoneyS,money_signData):
    try:
        df = pd.DataFrame.from_dict(json_data["Networth Projection"])
    except:
        return None
        
    ini = 0
    stps = 29
    ini_2 = ini
    last_val = 28
    for tab in range(ini,len(df),stps):
        print(last_val)
        #//*---Page setup----*//
        pdf.add_page()
        pdf.set_fill_color(*hex2RGB('#FCF8ED'))
        pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')

        pdf.image(join(cwd,'assets','images','backgrounds','doubleLine.png'),px2MM(1449),px2MM(0),px2MM(471),px2MM(1080))
        
        #//*----Net Worth Projection----*//
        pdf.set_xy(px2MM(120),px2MM(80)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.cell(px2MM(535), px2MM(84),'Net Worth Projection',align='L')
        
        #//*-----Table White rect
        pdf.set_fill_color(*hex2RGB('#FFFFFF'))

        year_rect_x = 886
        year_state_x = 906
        val_rect_y=val_rect_y2 = 289
        val_state_y=val_state_y2 = 299
        
        current_rect_x = 1025
        current_state_x = 1045
        
        project_rect_x = 1164
        project_state_x = 1184

        pdf.rect(px2MM(846), px2MM(204), px2MM(954), px2MM(755),'F')
        #//*-----Table Headings---*//
        
        #//*---Table 1
        #//*--Col 1
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.set_fill_color(*hex2RGB('#F3F6F9'))
        pdf.set_line_width(px2MM(1))
        pdf.rect(px2MM(886), px2MM(244), px2MM(139), px2MM(45),'FD')
        pdf.set_font('LeagueSpartan-Medium', size=px2pts(18))
        pdf.set_text_color(*hex2RGB('#000000'))
        
        pdf.set_xy(px2MM(906),px2MM(254))
        pdf.cell(px2MM(99), px2MM(25),'Year',align='C')
        
        #//*--Col 2
        pdf.rect(px2MM(1025), px2MM(244), px2MM(139), px2MM(45),'FD')
        pdf.set_xy(px2MM(1045),px2MM(254))
        pdf.cell(px2MM(99), px2MM(25),'Current(Cr)',align='C')
        
        #//*--Col 3

        pdf.rect(px2MM(1164), px2MM(244), px2MM(139), px2MM(45),'FD')
        pdf.set_xy(px2MM(1184),px2MM(254))
        pdf.cell(px2MM(99), px2MM(25),'Projected(Cr)',align='C')
        
        #//*---Table 2
        if len(df)-ini >14:
            #//*--Col 1
            pdf.set_fill_color(*hex2RGB('#F3F6F9'))
            pdf.rect(px2MM(1343), px2MM(244), px2MM(139), px2MM(45),'FD')
            pdf.set_font('LeagueSpartan-Medium', size=px2pts(18))
            
            pdf.set_xy(px2MM(1363),px2MM(254))
            pdf.set_text_color(*hex2RGB('#000000'))
            pdf.cell(px2MM(99), px2MM(25),'Year',align='C')
            
            #//*--Col 2
            pdf.rect(px2MM(1482), px2MM(244), px2MM(139), px2MM(45),'FD')
            pdf.set_xy(px2MM(1502),px2MM(254))
            pdf.cell(px2MM(99), px2MM(25),'Current(Cr)',align='C')
            
            #//*--Col 3
            pdf.rect(px2MM(1621), px2MM(244), px2MM(139), px2MM(45),'FD')
            pdf.set_xy(px2MM(1641),px2MM(254))
            pdf.cell(px2MM(99), px2MM(25),'Projected(Cr)',align='C')
            
        
        #//**--Table x and y settings---**//
        common_gap = 45
        
        #//*---Table value---*//
        
        for i in range(ini_2,last_val):
            try:
        
            #//*----Col 1
                if i%2==0:
                    pdf.set_fill_color(*hex2RGB('#ffffff'))
                else:
                    pdf.set_fill_color(*hex2RGB('#F3F6F9'))
                pdf.set_draw_color(*hex2RGB('#E9EAEE'))
                pdf.set_line_width(px2MM(1))
                pdf.set_font('LeagueSpartan-Regular', size=px2pts(18))
                pdf.set_xy(px2MM(year_state_x),px2MM(val_state_y))
                pdf.set_text_color(*hex2RGB('#000000'))
                if (df['year'][i]):
                    pdf.rect(px2MM(year_rect_x), px2MM(val_rect_y), px2MM(139), px2MM(45),'FD') 
                pdf.cell(px2MM(99), px2MM(25),"{0:.0f}".format(df['year'][i]),align='C')
                
                #//*---Col 2
                pdf.rect(px2MM(current_rect_x), px2MM(val_rect_y), px2MM(139), px2MM(45),'FD') 
                pdf.set_xy(px2MM(current_state_x),px2MM(val_state_y))
                pdf.cell(px2MM(99), px2MM(25),"{0:.2f}".format(df['Current net worth Trajectory'][i]),align='C')
                
                #//*---Col 3
                pdf.rect(px2MM(project_rect_x), px2MM(val_rect_y), px2MM(139), px2MM(45),'FD') 
                pdf.set_xy(px2MM(project_state_x),px2MM(val_state_y))
                pdf.cell(px2MM(99), px2MM(25),"{0:.2f}".format(df['Ideal net worth Trajectory'][i]),align='C')
                
                val_rect_y+=common_gap
                val_state_y+=common_gap
                
                if i==27:
                    ini_2 = 28
                    last_val+=28
        
                #//*---Reiniatilizing x and y axis for 2nd side table
                if i == 13:
                    year_rect_x = 1343
                    year_state_x = 1363
                    current_rect_x = 1482
                    current_state_x = 1502
                    project_rect_x = 1621
                    project_state_x = 1641
                    
                    val_rect_y = val_rect_y2
                    val_state_y= val_state_y2

            except:
                pass
                
        if tab ==0:
            #//*------Line Graph---*//
            font_dir = [join(cwd, 'assets','fonts','League_Spartan','static')]
            font_files2 = font_manager.findSystemFonts(fontpaths=font_dir)
            for files in font_files2:
                font_manager.fontManager.addfont(files)

            # fig,ax = plt.subplots(figsize=(7.8, 8.8))
            fig,ax = plt.subplots()
            # val_a = df['year'].astype(int)
            min_year = df['year'].min()
            max_year = df['year'].max()
            
            a = df['year'].astype(int)
            b = df['Ideal net worth Trajectory'].astype(float)
            c = df['Current net worth Trajectory'].astype(float)  

            # a = df['year'].astype(int)
            # b = df['Ideal net worth Trajectory'].astype(float)
            # c = df['Current net worth Trajectory'].astype(float)
            
                
            # for x in range(0,len(df),step):
            #     # a.append(df['year'][x].astype(int))
            #     b.append(df['Ideal net worth Trajectory'][x].astype(float))
            #     c.append(df['Current net worth Trajectory'][x].astype(float))
            
            
            #//*----------case 1-------------------*//
            pp = math.ceil(len(df['year'])/8)
            print(pp)
            rem = pp%8
            # x_tick_val = np.arange(start=df['year'].min(), stop=df['year'].max(), step=pp)

            # if x_tick_val[-1] == df['year'].iloc[-1]:
            #     pass
            # else:
            #     x_tick_val = list(x_tick_val)
            #     x_tick_val.append(int(df['year'].iloc[-1]))    
            
            # if len(a)>3:
            #     diff = (x_tick_val[-1]-x_tick_val[-3])/2
            #     in_diff = (x_tick_val[-2]-x_tick_val[-3])
                
            #     if in_diff>diff:
            #         x_tick_val.pop(-2)
            #         print('old',x_tick_val)
            #         mid_range_tick = list(int(x) for x in range(int(x_tick_val[-2]),int(x_tick_val[-1])))
            #         mid = math.ceil(len(mid_range_tick)/2)
            #         print('mid',mid_range_tick[mid])
            #         x_tick_val.insert(-1,mid_range_tick[mid])
            #         print('new',x_tick_val)
                
                
            # for j in a:
            #     b.append(df.loc[df['year'] == j, 'Ideal net worth Trajectory'].iloc[0])
            #     c.append(df.loc[df['year'] == j, 'Current net worth Trajectory'].iloc[0])
                
            # if a[-1] == df['year'].iloc[-1]:
            #     pass
            # else:
            #     a = list(a)
            #     a.append(int(df['year'].iloc[-1]))    
            #     b.append(df['Ideal net worth Trajectory'].iloc[-1])    
            #     c.append(df['Current net worth Trajectory'].iloc[-1]) 
  
            
              

                
            if max(c)>max(b):
                color_a = '#FF7051'
                color_b =  '#43D195'
            else:
                color_a =  '#43D195'
                color_b = '#FF7051'

            default_x_ticks = range(len(a))
            
            # ax = sns.lineplot(x = a,y=b, data=df)
            ax = sns.lineplot(x = a,y=b)
            # ax.xaxis.set_major_locator(plt.MaxNLocator(8))
            # ax1 = sns.lineplot(x = a,y=c)
            plt.plot(a,b,color=color_a,ms = 3 ,lw = 1)
            plt.plot(a,c,color=color_b,ms = 3 ,lw = 1)
            # plt.plot(min(a),min(b),color='black',ms = 5)
            ax.yaxis.set_major_formatter(tick.FuncFormatter(y_fmt))
    
            #FF7051


            # plt.fill_between(a, c,color='g',alpha=.8,zorder=20)
            # plt.fill_between(a, b,color='r',alpha=.9)
            
            # plt.xlabel('Years',labelpad = 8,fontdict= {'fontsize':8,'weight': 'bold',})
            # plt.ylabel('₹ in Cr',labelpad = 8,fontdict= {'fontsize':8,'weight': 'bold'})
            plt.xlabel('')
            plt.ylabel('')
            
            # ax.set_xticks(df['year'],fontdict={'fontsize':6,'weight': 'normal','family': 'LeagueSpartan-Bold',})
            

            min_year = df['year'].min()
            max_year = df['year'].max()
            
            # min_ideal = df['Current net worth Trajectory'].min()
            # max_ideal = df['Ideal net worth Trajectory'].max()
            
            min_ideal = min(b)
            max_ideal = max(b)
            
            # z = max_ideal/3
            # max_ideal = max_ideal +z
            # plt.xlim(min_year-1,max_year)
            # plt.ylim(min_ideal-2,max_ideal)
            
            if max(c)>max(b):
                min_ideal = min(c)
                max_ideal = max(c)
            
                z = max_ideal/3
                max_ideal = max_ideal +z
                ymin = min(c)
                ymx = max(c)
            else:
                min_ideal = min(b)
                max_ideal = max(b)
            
                z = max_ideal/3
                max_ideal = max_ideal+z 
                ymin = min(b)
                ymx = max(b)
                
            plt.xlim(min(a),max(a))
            plt.ylim(min_ideal,max_ideal)
            # ax.set_xticklabels(df['year'])
            
            # plt.fill_between(a, b,c,color= '#D4FFED',alpha=.5,zorder=100)
            # plt.fill_between(a, b,color= '#FFD4CB',alpha=.5,zorder=100)
            
            lop_list = []
            
                
            max_c = max(c)
            red_lp = np.linspace(max_c,0,100)
            for i in red_lp:
                plt.fill_between(a,i,c,color= '#FFD4CB',alpha=0.03) 
            
            NbData = len(a)  
            max_a = max(a)  
            red_MaxBL = [[MaxBL] * NbData for MaxBL in range(max_a)]
            Max = [np.asarray(red_MaxBL[x]) for x in range(max_a)]
            
            for x in range (math.ceil(max(c)),max_a):
                plt.fill_between(a,Max[x],c, facecolor='white', alpha=1) 
                
            
            plt.fill_between(a, b,c,color= '#D4FFED',alpha=.9)
            # green_lp = np.linspace(max(b),0,100)
            # for i in green_lp:
            #     plt.fill_between(a,i,b,color='#D4FFED', alpha=0.3)  
                
                    
            
            # NbData = len(a)    
            # MaxBL = [[MaxBL] * NbData for MaxBL in range(100)]
            # Max = [np.asarray(MaxBL[x]) for x in range(100)]
            
            # # for x in range (int(max_b),0,-1):
            # #     plt.fill_between(a,Max[x],b,facecolor='#FFD4CB', alpha=0.2)

            # for x in range (int(max(c)+10),100):
            #     plt.fill_between(a,Max[x],b, facecolor='white', alpha=1)
            
            # plt.xticks(x_tick_val)
            # plt.xticks(a)
            
            #//****Case 5--*//
            
            # rmd = 0
            
            # for l in range(1,len(df)+1):
            #     if len(df)%l ==0 and rmd<4:
            #         rmd = l
            #         print(rmd)
            # plt.xticks(np.arange(df['year'].min(), df['year'].max(), rmd))
            # # //*-------Case2------*//
            pp = math.ceil(len(a)/8)
            rem = pp%8
            arg = np.arange(start=df['year'].min(), stop=df['year'].max(), step=pp)
            plt.xticks(np.arange(min(a)+1, max(a),pp))
            
            if len(a)>0 and len(a)<=8:
                plt.xticks(np.arange(min(a)+1, max(a), 1))
            elif len(a)>8 and len(a)<=16:
                plt.xticks(np.arange(min(a)+1, max(a), 2))
            elif len(a)>16 and len(a)<=24:
                plt.xticks(np.arange(min(a)+1, max(a), 3))
            elif len(a)>24 and len(a)<=32:
                plt.xticks(np.arange(min(a)+1, max(a)+1, 4))
            elif len(a)>32and len(a)<=40:
                plt.xticks(np.arange(min(a)+1, max(a)+1,5))
                
                
           
            # # //*---Case 3------*//        
            # plt.xticks(a)
            # if len(a)>0 and len(a)<=8:
            #     plt.xticks(np.arange(min(a), max(a)+1, 1))
            # elif len(a)>8 and len(a)<=16:
            #     kl = max(a)%2
            #     mx_a = max(a)+(2-kl)
            #     plt.xticks(np.arange(min(a), mx_a, 2))
            # elif len(a)>16 and len(a)<=24:
            #     kl = max(a)%3
            #     mx_a = max(a)+(3-kl)
            #     plt.xticks(np.arange(min(a), mx_a, 3))
            # elif len(a)>24 and len(a)<=32:
            #     kl = max(a)%4
            #     mx_a = max(a)+(4-kl)
            #     plt.xticks(np.arange(min(a), mx_a, 4))
            # elif len(a)>32and len(a)<=40:
            #     kl = max(a)%5
            #     mx_a = max(a)+(5-kl)
            #     plt.xticks(np.arange(min(a), mx_a, 5))
            


            #//*---X tick Rotation
            plt.yticks(fontname = "Arial")  
            plt.xticks(fontname = "Arial")  
            ax.tick_params(axis='x', labelrotation = 00)
            ax.tick_params(axis='both',labelsize=10,colors='#65676D')
            ax.tick_params(axis='y',labelsize=10)
            ax.grid(color='#DCDCDC', linestyle='-', linewidth=0.15)
            ax.yaxis.grid(True) 
            ax.xaxis.grid(True)
            ax.spines[['right', 'top','left','bottom']].set_visible(False)
            plt.tick_params(left = False,bottom = False)

            # ax.spines['bottom'].set_color('#DCDCDC')

            plt.savefig('acutal_networth_chart.png',dpi=250)
        # plt.show()
        
        color_a = '#FF7051'
        color_b =  '#43D195'
        #//*----Legend and Graph plotting---*//
        pdf.set_fill_color(*hex2RGB('#ffffff'))
        pdf.rect(px2MM(120),px2MM(204),px2MM(686),px2MM(623),'F')    
        pdf.image('acutal_networth_chart.png',px2MM(160),px2MM(204),px2MM(606),px2MM(400))
        
        pdf.set_fill_color(*hex2RGB(color_a))
        pdf.rect(px2MM(169),px2MM(629),px2MM(12),px2MM(12),'F')   
        
        pdf.set_font('LeagueSpartan-Medium', size=px2pts(24))
        pdf.set_xy(px2MM(196),px2MM(619))
        pdf.set_text_color(*hex2RGB('#000000'))
        # pdf.cell(px2MM(295), px2MM(32),str(df['Current net worth Trajectory'].max()),align='C') 
        pdf.cell(px2MM(295), px2MM(32),'Current Net Worth Trajectory',align='C')
        
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_xy(px2MM(196),px2MM(656))
        pdf.set_text_color(*hex2RGB('#898B90'))
        # pdf.cell(px2MM(295), px2MM(32),str(df['Current net worth Trajectory'].max()),align='C') 
        max_curr = "{0:.2f}".format(float(df['Current net worth Trajectory'].max()))
        today = datetime.now()
        mnth = today.strftime("%B")
        data1 = mnth+' '+str(int(max_year))+' | ₹'+max_curr+' Cr'
        pdf.cell(px2MM(290), px2MM(32),data1,align='L')
        
        
        #//*------
        pdf.set_fill_color(*hex2RGB(color_b))
        pdf.rect(px2MM(169),px2MM(728),px2MM(12),px2MM(12),'F')   
        
        pdf.set_font('LeagueSpartan-Medium', size=px2pts(24))
        pdf.set_xy(px2MM(196),px2MM(718))
        pdf.set_text_color(*hex2RGB('#000000'))
        # pdf.cell(px2MM(295), px2MM(32),str(df['Current net worth Trajectory'].max()),align='C') 
        pdf.cell(px2MM(403), px2MM(32),'Net Worth with Right Financial Planning',align='C')
        
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_xy(px2MM(196),px2MM(755))
        pdf.set_text_color(*hex2RGB('#898B90'))
        # pdf.cell(px2MM(295), px2MM(32),str(df['Current net worth Trajectory'].max()),align='C')
        max_net_worth = "{0:.2f}".format(float(df['Ideal net worth Trajectory'].max())) 

        data2 = mnth+' '+str(int(max_year))+' | ₹'+max_net_worth+' Cr'
        pdf.cell(px2MM(225), px2MM(32),data2,align='L')
        
        ini = 29
        stps = 29 
        
def y_fmt(x, y):
    return f'₹ {int(x)}Cr'.format(x)
#//*---Structure for Term and Health Insurance---*// 
def term_health_features(pdf,json_data,c_MoneyS,money_signData,insurer,plan,pros,cons,pg_name,max_row):
    #/*---line count for pros
    pros_each_length =[]
    for len_tot_pros in range(len(pros)):
        line_count = 0
        
        for len_each_pros in range(len(pros[len_tot_pros])):
            lines = int(len((pros[len_tot_pros][len_each_pros])))/54
            try:
                if lines%int(lines)!= 0:
                    lines = int(lines)+1
            except:
                lines = int(lines)+1
            line_count+= lines
        line_count=line_count*32
        pros_each_length.append(line_count)
        
    
    
    #/*---line count for cons
    cons_each_length =[]
    for len_tot_cons in range(len(cons)):
        line_count = 0
        
        for len_each_cons in range(len(cons[len_tot_cons])):
            lines = int(len((cons[len_tot_cons][len_each_cons])))/40
            try:
                if lines%int(lines)!= 0:
                    lines = int(lines)+1
            except:
                lines = int(lines)+1
            line_count+= lines
        line_count=line_count*32
        cons_each_length.append(line_count)
    tot_lines = []    
    for i in range(len(insurer)):
        tot_lines.append([pros_each_length[i],cons_each_length[i]])         
    tot_lines = list(max(x) for x in tot_lines)  
    
    ini = 0
    steps = max_row 
           
    for i in range(0,len(insurer),max_row):
        pdf.add_page()
        pdf.set_fill_color(*hex2RGB('#FCF8ED'))
        pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')

        #//*----Featured List of Financial Products----*//
        pdf.set_xy(px2MM(120),px2MM(80)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.cell(px2MM(877), px2MM(84),'Financial Products Featured List',align='L')
        
        #//*---Top Black box
        
        
        if pg_name =='Term Insurance Plans':
            pdf.set_fill_color(*hex2RGB('#313236'))
            pdf.rect(px2MM(126), px2MM(204), px2MM(242), px2MM(42),'F')
            pdf.set_xy(px2MM(141),px2MM(209)) 
            pdf.set_font('LeagueSpartan-Medium', size=px2pts(24))
            pdf.set_text_color(*hex2RGB('#ffffff'))
            pdf.cell(px2MM(212), px2MM(32),pg_name,align='C')
        else:
            pdf.set_fill_color(*hex2RGB('#313236'))
            pdf.rect(px2MM(126), px2MM(204), px2MM(259), px2MM(42),'F')
            pdf.set_xy(px2MM(141),px2MM(209)) 
            pdf.set_font('LeagueSpartan-Medium', size=px2pts(24))
            pdf.set_text_color(*hex2RGB('#ffffff'))
            pdf.cell(px2MM(229), px2MM(32),pg_name,align='C')
        #//*---black line height---*//
        bl_height = 114
        #//*---Table Value
        
        #//*--Col 1
        pdf.set_fill_color(*hex2RGB('#ffffff'))
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.rect(px2MM(126), px2MM(246), px2MM(220), px2MM(72),'FD')
        
        pdf.set_xy(px2MM(146),px2MM(266)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.multi_cell(px2MM(180), px2MM(32),'Insurer',align='L')
        
        #//*---Col 2
        pdf.rect(px2MM(346), px2MM(246), px2MM(240), px2MM(72),'FD')
        pdf.set_xy(px2MM(366),px2MM(266)) 
        pdf.multi_cell(px2MM(200), px2MM(32),'Plan',align='L')
        
        #//*---Col 3
        pdf.rect(px2MM(586), px2MM(246), px2MM(650), px2MM(72),'FD')
        pdf.set_xy(px2MM(606),px2MM(266)) 
        pdf.multi_cell(px2MM(610), px2MM(32),'Pros',align='L')
        
        #//*---Col 4
        pdf.rect(px2MM(1236), px2MM(246), px2MM(564), px2MM(72),'FD')
        pdf.set_xy(px2MM(1256),px2MM(266)) 
        pdf.multi_cell(px2MM(524), px2MM(32),'Cons',align='L')
        
        # text_y = 338
        # rect_y = 318
        line_no = 0
        text_y = 338
        rect_y = 318

        try:
            for row in range(ini,steps):  
                line_no +=row
                #//*---rows
                #//*--Col 1
                if row%2==0:
                    pdf.set_fill_color(*hex2RGB('#F3F6F9'))
                    row_color = '#F3F6F9'
                else:
                    pdf.set_fill_color(*hex2RGB('#ffffff'))
                    row_color = '#FFFFFF'
                if pg_name=='Term Insurance Plans': 
                    if row==3:
                        pdf.set_fill_color(*hex2RGB('#F3F6F9'))
                        row_color = '#F3F6F9'
                    elif row==4:
                        pdf.set_fill_color(*hex2RGB('#ffffff'))
                        row_color = '#FFFFFF'
                        
                pdf.set_draw_color(*hex2RGB('#E9EAEE'))
                pdf.rect(px2MM(126), px2MM(rect_y), px2MM(220), px2MM(tot_lines[row]+40),'FD')
                bl_height+=tot_lines[row]+40
                pdf.set_xy(px2MM(146),px2MM(text_y)) 
                pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
                pdf.set_text_color(*hex2RGB('#1A1A1D'))
                pdf.multi_cell(px2MM(180), px2MM(32),insurer[row],align='L')
                
                #//*---Col 2
                pdf.rect(px2MM(346), px2MM(rect_y), px2MM(240), px2MM(tot_lines[row]+40),'FD')
                pdf.set_xy(px2MM(366),px2MM(text_y)) 
                pdf.multi_cell(px2MM(200), px2MM(32),plan[row],align='L')
                
                #//*---Col 3
                pdf.rect(px2MM(586), px2MM(rect_y), px2MM(650), px2MM(tot_lines[row]+40),'FD')
                
            

                for j in range(0,len(pros[row])):
                    if pros[row][j]=='':
                            continue
                    if j==0:
                        pdf.set_fill_color(*hex2RGB('#000000'))
                        pdf.circle(x=px2MM(620),y=px2MM(text_y+15),r=px2MM(6),style='F')
                        
                        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
                        pdf.set_xy(px2MM(646),px2MM(text_y)) 
                        pdf.multi_cell(px2MM(580), px2MM(32),pros[row][j],align='L')
                            
                    else:
                        p_y = mm2PX(pdf.get_y())
                        pdf.set_fill_color(*hex2RGB('#000000'))
                        pdf.circle(x=px2MM(620),y=px2MM(p_y+15),r=px2MM(6),style='F')
                
                        pdf.set_xy(px2MM(646),px2MM(p_y)) 
                        pdf.multi_cell(px2MM(580), px2MM(32),pros[row][j],align='L')
                
                #//*---Col 4
                pdf.set_fill_color(*hex2RGB(row_color))
                pdf.rect(px2MM(1236), px2MM(rect_y), px2MM(564), px2MM(tot_lines[row]+40),'FD')
                
                if len(cons[row])==1:
                    pdf.set_fill_color(*hex2RGB('#000000'))
                    pdf.circle(x=px2MM(1276),y=px2MM(text_y+15),r=px2MM(6),style='F')
                    
                    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
                    pdf.set_xy(px2MM(1296),px2MM(text_y)) 
                    pdf.multi_cell(px2MM(484), px2MM(32),cons[row][0],align='L')
                else:
                    for k in range(0,len(cons[row])):
                        if cons[row][k]=='':
                            continue
                        if k ==0:
                            pdf.set_fill_color(*hex2RGB('#000000'))
                            pdf.circle(x=px2MM(1276),y=px2MM(text_y+15),r=px2MM(6),style='F')
                            
                            pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
                            pdf.set_xy(px2MM(1296),px2MM(text_y)) 
                            pdf.multi_cell(px2MM(484), px2MM(32),cons[row][k],align='L')
                        else:    
                            p_y = mm2PX(pdf.get_y())
                            pdf.set_fill_color(*hex2RGB('#000000'))
                            pdf.circle(x=px2MM(1276),y=px2MM(p_y+15),r=px2MM(6),style='F')
                            
                            pdf.set_xy(px2MM(1296),px2MM(p_y)) 
                            pdf.multi_cell(px2MM(484), px2MM(32),cons[row][k],align='L')
                            
                
                text_y =text_y+ tot_lines[row]+40
                rect_y =rect_y+tot_lines[row]+40
                            
            
           
        except:
            e,p,t= sys.exc_info()
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_xy(px2MM(597),px2MM(1008)) 
        pdf.cell(px2MM(727), px2MM(32),"Note: The above featured list is based on 1 Finance's proprietary research.",align='L')
        
        pdf.set_fill_color(*hex2RGB('#313236'))
        pdf.rect(px2MM(120), px2MM(204), px2MM(6), px2MM(bl_height),'F')
        ini += max_row
        steps += max_row
        
#//*------Health Term Insurance -----*//        
def term_insurance(pdf,json_data,c_MoneyS,money_signData):
    
    #//*------For Term Insurance Plans-----*//
    insurer = ['Hdfc Life','Max Life','TATA-AIA Life','Aditya Birla Sun Life','ICICI Pru Life']
    plan=['Hdfc Life Click 2 protect Life','Max Life Smart Secure Plus','Sampoorna Raksha Supreme','Life Shield','iProtect Smart Term Plan']
    pros = [['Claim Settlement Ratio (CSR) of 98.66%','Flexibility to change premium payment terms from regular to limited pay','Change of payment frequency during premium paying term available at no cost','Good track record of settling claims faster than other insurers','3 complaints per 10,000 claims'],
            ['Claim Settlement ratio of 99.34%','Slightly less expensive than competing products on the market','Terminal illness coverage at no cost','Coverage of 64 Critical illness, which is a separate cover in addition to base cover amount'],
            ['Claim settlement ratio of 98.53%','In-built terminal illness cover','Option to increase coverage at key life stages'],
            ['Claim settlement ratio of 98.07%','Option to cover your spouse under the same policy','Terminal illness benefit is inbuilt in the plan',"Excellent track record of settling claims faster than other insurer's"],
            ['Inbuilt terminal illness benefit','18% discount for women policyholders']
            ]
    cons = [['Expensive than other competing products'],
            ['There are few options for customising claim payouts'],
            ['No discount on online purchase of policy'],
            ['When compared to other insurers, received a higher number of complaints during policy purchase'],
            ['Much more expensive than competing products on the market','Critical illness rider is not an additional cover, it pays out from the life cover amount','A track record of slow claim settlement']
            ]
    
    pg_name = 'Term Insurance Plans'
    term_health_features(pdf,json_data,c_MoneyS,money_signData,insurer,plan,pros,cons,pg_name,3)
    
#//*------Health Insurance Plans-----*//

def health_insurance(pdf,json_data,c_MoneyS,money_signData):
    insurer = ['Manipal Cigna','Niva Bupa Health','Care Health','Aditya Birla Health','Navi General']
    plan=['ProHealth Prime Protect','Health ReAssure','Care Plus','Activ Health Platinum Enhanced','Navi Health Plan 5']
    pros = [
        ['No room rent sublimits','No Claim Bonus (NCB) benefit is greater than 150% over time','Sum insured can be recharged up to the cover amount for related diseases','No co-pay in the policy','Rewards for healthy behaviour are available'],
       ['No room rent sublimits','No Claim Bonus (NCB) benefit over time is up to 100%','Sum insured can be recharged up to the cover amount for related diseases','No co-pay in the policy','Rewards for healthy behaviour are available','8600 cashless network hospitals'],
        ['No room rent sublimits','No Claim Bonus (NCB) benefit is greater than 150% over time','Recharge of sum insured upto the cover amount is available for an unlimited number of times for unrelated diseases','No co-pay in the policy','Rewards for healthy behaviour are available','There are 19000 cashless network hospitals, which is the most of any insurer'],
        ['No room rent sublimits','No Claim Bonus (NCB) benefit can be up to 100% over time','No co-pay in the policy','Rewards for healthy behaviour are available','10,051 cashless network hospitals','Sum insured can be recharged once per policy year up to the cover amount'],
       ['No Claim Bonus (NCB) benefit is greater than 150% over time','Recharge of sum insured upto the cover amount is available for an unlimited number of times for unrelated diseases','Pre-existing diseases are covered after 1 year','No Co-pay in the policy','Rewards for healthy behaviour are available','10,000 Cashless Network Hospitals']
    ]
    cons = [
        ['6500 Cashless Network Hospitals'],
        ['Pre-existing diseases are covered after 3 years'],
        ['Pre-existing diseases are covered after 3 years'],
        ['Pre-existing diseases are covered after 3 years'],
        ['Low Claim Settlement Ratio (CSR) of 83%']
    ]
    pg_name = 'Health Insurance Plans'
    
    term_health_features(pdf,json_data,c_MoneyS,money_signData,insurer,plan,pros,cons,pg_name,2)
    
#//*----New Mutual Fund-------------*//      
def mutual_fund(pdf,json_data,c_MoneyS,money_signData):
    category = ['Large Cap Index','Large Cap Index','Large Cap Index','Flexicap fund','Flexicap fund','Flexicap fund']
    fund_scheme = ['HDFC Index Fund-S&P BSE Sensex','Navi Nifty 50 Index Fund','Axis Nifty 50 Index Fund','Parag Parekh Flexi Cap Fund','Kotak Flexicap Fund','Franklin India Flexi Cap Fund']
    of_score = [96,90,90,99,97,94]
    strength = [
        ['Accurate tracking of performance of underlying benchmark index with minimal deviation','Large AUM of ₹ 4156 cr'],
        ['Accurate tracking of performance of underlying benchmark index with minimal deviation','Low expense ratio of 0.06% as compared with category average of 0.22%.'],
        ['Accurate tracking of performance of underlying benchmark index with minimal deviation','Low expense ratio of 0.12% as compared with category average of 0.22%.'],
        ['Consistency in delivering high risk-adjusted returns by actively managing funds between large cap, midcap and smallcap','Strong ability to outperform benchmark returns'],
        ['Consistency in delivering high risk-adjusted returns by actively managing funds between large cap, midcap and smallcap','Low expense ratio of 0.67% versus category average of 0.84%'],
        ['Consistency in delivering high risk-adjusted returns by actively managing funds between large cap, midcap and smallcap']
    ]
    weakness = [
        [''],
        ['Only 1 year of track record','Small AUM of ₹ 96 cr'],
        ['Only 1 year of track record'],
        [''],
        [''],
        ['High expense ratio of 1.09% versus category average of 0.84%'],
    ]
    
    #//**----Line Count for the Strength
    str_each_length=[]
    
    for len_tot_str in range(len(category)):
        line_count = 0
        
        for len_each_str in range(len(strength[len_tot_str])):
            lines = int(len((strength[len_tot_str][len_each_str])))/42
            try:
                if lines%int(lines)!= 0:
                    lines = int(lines)+1
            except:
                lines = int(lines)+1
            line_count+= lines
        line_count=line_count*32
        str_each_length.append(line_count)
    #//**----Line Count for the weakness
    weak_each_length=[]
    
    for len_tot_weak in range(len(category)):
        line_count = 0
        
        for len_each_weak in range(len(weakness[len_tot_weak])):
            lines = int(len((weakness[len_tot_weak][len_each_weak])))/25
            try:
                if lines%int(lines)!= 0:
                    lines = int(lines)+1
            except:
                lines = int(lines)+1
            line_count+= lines
        line_count=line_count*32
        weak_each_length.append(line_count)

    
    tot_lines = []    
    for i in range(len(category)):
        tot_lines.append([str_each_length[i],weak_each_length[i]])        
    tot_lines = list(max(x) for x in tot_lines)         
    ini = 0
    steps = 3
    for i in range(0,len(category),3):
        pdf.add_page()
        pdf.set_fill_color(*hex2RGB('#FCF8ED'))
        pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')

        #//*----Featured List of Financial Products----*//
        pdf.set_xy(px2MM(120),px2MM(80)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.cell(px2MM(877), px2MM(84),'Financial Products Featured List',align='L')
        
        #//*---Top Black box
        pdf.set_fill_color(*hex2RGB('#313236'))
        pdf.rect(px2MM(126), px2MM(204), px2MM(165), px2MM(42),'F')
        
        pdf.set_xy(px2MM(141),px2MM(209)) 
        pdf.set_font('LeagueSpartan-Medium', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#ffffff'))
        pdf.cell(px2MM(135), px2MM(32),"Mutual Funds",align='C')
        
        #//*---black line height---*//
        bl_height = 114
        
        #//*-----Table Header Creation----*//
        #//*---Col 1
        
        pdf.set_fill_color(*hex2RGB('#ffffff'))
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.rect(px2MM(126), px2MM(246), px2MM(240), px2MM(72),'FD')
        
        pdf.set_xy(px2MM(146),px2MM(266)) 
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.multi_cell(px2MM(200), px2MM(32),'Category',align='L')
        
        #//*---Col 2
        pdf.rect(px2MM(366), px2MM(246), px2MM(250), px2MM(72),'FD')
        pdf.set_xy(px2MM(386),px2MM(266)) 
        pdf.multi_cell(px2MM(230), px2MM(32),'Fund Scheme',align='L')
       
       #//*---Col 3
        pdf.rect(px2MM(616), px2MM(246), px2MM(200), px2MM(72),'FD')
        pdf.set_xy(px2MM(636),px2MM(266)) 
        pdf.multi_cell(px2MM(160), px2MM(32),'1 Finance score',align='L')
        
        #//*---Col 4
        pdf.rect(px2MM(816), px2MM(246), px2MM(610), px2MM(72),'FD')
        pdf.set_xy(px2MM(836),px2MM(266)) 
        pdf.multi_cell(px2MM(570), px2MM(32),'Strength',align='L')
        
        #//*---Col 5
        pdf.rect(px2MM(1426), px2MM(246), px2MM(374), px2MM(72),'FD')
        pdf.set_xy(px2MM(1446),px2MM(266)) 
        pdf.multi_cell(px2MM(334), px2MM(32),'Weakness',align='L')
        
        line_no = 0
        rect_y = 318
        text_y = 338
        
        try:
            for row in range(ini,steps):
                line_no+=row
                
                #//*----Column 1 Value
                if row%2==0:
                    pdf.set_fill_color(*hex2RGB('#F3F6F9'))
                    row_color = '#F3F6F9'
                else:
                    pdf.set_fill_color(*hex2RGB('#ffffff'))
                    row_color = '#FFFFFF'
                    
                if row==3 or row==5:
                    pdf.set_fill_color(*hex2RGB('#F3F6F9'))
                    row_color = '#F3F6F9'
                elif row ==4:
                    pdf.set_fill_color(*hex2RGB('#ffffff'))
                    row_color = '#FFFFFF' 
                    
                    
                pdf.set_draw_color(*hex2RGB('#E9EAEE'))
                pdf.rect(px2MM(126), px2MM(rect_y), px2MM(240), px2MM(tot_lines[row]+25),'FD')
                bl_height+=tot_lines[row]+25
                pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
                pdf.set_text_color(*hex2RGB('#1A1A1D'))
                pdf.set_xy(px2MM(146),px2MM(text_y))
                pdf.multi_cell(px2MM(180), px2MM(32),category[row],align='L') 
                
                #//*---Column 2 Value
                pdf.rect(px2MM(366), px2MM(rect_y), px2MM(250), px2MM(tot_lines[row]+25),'FD')
                pdf.set_xy(px2MM(386),px2MM(text_y)) 
                pdf.multi_cell(px2MM(210), px2MM(32),fund_scheme[row],align='L')
                
                #//*---Column 3 Value
                pdf.rect(px2MM(616), px2MM(rect_y), px2MM(250), px2MM(tot_lines[row]+25),'FD')
                pdf.set_xy(px2MM(636),px2MM(text_y)) 
                pdf.multi_cell(px2MM(160), px2MM(32),str(of_score[row]),align='C')
                
                #//*----Column 4 Value
                
                pdf.rect(px2MM(816), px2MM(rect_y), px2MM(610), px2MM(tot_lines[row]+25),'FD')
                
                for j in range(0,len(strength[row])):
                    if j==0:
                        pdf.set_fill_color(*hex2RGB('#000000'))
                        pdf.circle(x=px2MM(856),y=px2MM(text_y+15),r=px2MM(6),style='F')
                        
                        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
                        pdf.set_xy(px2MM(876),px2MM(text_y)) 
                        pdf.multi_cell(px2MM(530), px2MM(32),strength[row][j],align='L')
                        
                    else:
                        p_y = mm2PX(pdf.get_y())
                        pdf.set_fill_color(*hex2RGB('#000000'))
                        pdf.circle(x=px2MM(856),y=px2MM(p_y+15),r=px2MM(6),style='F')
                
                        pdf.set_xy(px2MM(876),px2MM(p_y)) 
                        pdf.multi_cell(px2MM(530), px2MM(32),strength[row][j],align='L')
                        
                #//*----Column 5 Value
                pdf.set_fill_color(*hex2RGB(row_color))
                pdf.rect(px2MM(1426), px2MM(rect_y), px2MM(374), px2MM(tot_lines[row]+25),'FD')
                
                for j in range(0,len(weakness[row])):
 
                    if weakness[row][j]=='':
                        continue
                    if j==0:
                        pdf.set_fill_color(*hex2RGB('#000000'))
                        pdf.circle(x=px2MM(1466),y=px2MM(text_y+15),r=px2MM(6),style='F')
                        
                        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
                        pdf.set_xy(px2MM(1486),px2MM(text_y)) 
                        pdf.multi_cell(px2MM(294), px2MM(32),weakness[row][j],align='L')
                        
                    else:
                        p_y = mm2PX(pdf.get_y())
                        pdf.set_fill_color(*hex2RGB('#000000'))
                        pdf.circle(x=px2MM(1466),y=px2MM(p_y+15),r=px2MM(6),style='F')
                        pdf.set_xy(px2MM(1486),px2MM(p_y)) 

                            
                        pdf.multi_cell(px2MM(294), px2MM(32),weakness[row][j],align='L')     
                text_y =text_y+ tot_lines[row]+25
                rect_y =rect_y+tot_lines[row]+25      
                
          
        except:
            e,p,t= sys.exc_info()
        
        #//*---Note---*//    
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_xy(px2MM(120),px2MM(902)) 
        pdf.cell(px2MM(54), px2MM(32),"Note: ",align='L')
        
        note_stat = ['All the above schemes are Growth-Direct plans','1 Finance score ranges from 0-100',"The above featured list is based on 1 Finance's proprietary research"]
        
        for nt in range(3):
            pdf.set_fill_color(*hex2RGB('#000000'))
            pdf.circle(x=px2MM(140),y=px2MM((nt*32)+955),r=px2MM(6),style='F')  
            
            pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
            pdf.set_xy(px2MM(160),px2MM((nt*32)+940)) 
            pdf.cell(px2MM(650), px2MM(32),note_stat[nt],align='L')
            
        #//*----Black VerticaL Line
        pdf.set_fill_color(*hex2RGB('#313236'))
        pdf.rect(px2MM(120), px2MM(204), px2MM(6), px2MM(bl_height),'F')
            
        ini +=3
        steps +=3
            
#//*-----Bureau Report Summary---*//
#//*-----Bureau Report Summary---*//
def bureao_report(pdf,json_data,c_MoneyS,money_signData):
    try:
        csa = pd.DataFrame.from_dict(json_data["Credit_score_analysis"])
        cft = pd.DataFrame.from_dict(json_data["Credit_facility_taken"])
    except:
        print('\n\n\n\n\nno page')
        return None
    
    if cft.empty:
        return None
    try:
        type_facility = cft["Type of Facility"].tolist()
        tot_record = cft["Total Records"].tolist()
        active_acc = cft["Active Accounts"].tolist()
        clsd_acc = cft["Closed Accounts"].tolist()
        acc_neg_hist = cft["Accounts with Negative History"].tolist()
    except:
        return None

    
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')

    #//*----Featured List of Financial Products----*//
    pdf.set_xy(px2MM(120),px2MM(80)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(877), px2MM(84),'Bureau Report Summary',align='L')
    
    #//*---Top Black box
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84),'F')
    
    #//*---Credit Score Analysis
    pdf.set_fill_color(*hex2RGB('#313236'))
    pdf.rect(px2MM(126), px2MM(204), px2MM(243), px2MM(42),'F')
    
    pdf.set_fill_color(*hex2RGB('#313236'))
    pdf.rect(px2MM(120), px2MM(204), px2MM(6), px2MM(218),'F')
    
    pdf.set_xy(px2MM(141),px2MM(209)) 
    pdf.set_font('LeagueSpartan-Medium', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#ffffff'))
    pdf.cell(px2MM(213), px2MM(32),'Credit Score Analysis',align='C')
    
    #//*---Table Header----*//
    pdf.set_fill_color(*hex2RGB('#ffffff'))
    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.set_line_width(px2MM(0.5))
    #//*---Col 1
    pdf.rect(px2MM(126), px2MM(246), px2MM(240), px2MM(72),'FD')
    pdf.set_xy(px2MM(146),px2MM(266)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(200), px2MM(32),'Your Credit Score',align='C')
     #//*---Col 2
    pdf.rect(px2MM(366), px2MM(246), px2MM(320), px2MM(72),'FD')
    pdf.set_xy(px2MM(386),px2MM(266)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.cell(px2MM(600), px2MM(32),'Our Evaluation',align='L')
     #//*---Col 3
    pdf.rect(px2MM(686), px2MM(246), px2MM(1114), px2MM(72),'FD')
    pdf.set_xy(px2MM(706),px2MM(266)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.cell(px2MM(1074), px2MM(32),'Commentary',align='L')
    
    
    #//*---Table Value---*//
    
    pdf.set_fill_color(*hex2RGB('#F3F6F9'))
    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.set_line_width(px2MM(0.5))
    #//*---Col 1
    pdf.rect(px2MM(126), px2MM(318), px2MM(240), px2MM(104),'FD')
    pdf.set_xy(px2MM(146),px2MM(338)) 
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(200), px2MM(34),str(csa['Your Credit Score'].iloc[0]),align='C')
     #//*---Col 2
    pdf.rect(px2MM(366), px2MM(318), px2MM(320), px2MM(104),'FD')
    pdf.set_xy(px2MM(386),px2MM(338)) 
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    pdf.multi_cell(px2MM(600), px2MM(32),csa['Our Evaluation'].iloc[0],align='L')
     #//*---Col 3
    pdf.rect(px2MM(686), px2MM(318  ), px2MM(1114), px2MM(104),'FD')
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.circle(x=px2MM(728),y=px2MM(351),r=px2MM(5),style='F')
    pdf.set_xy(px2MM(746),px2MM(338)) 
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    pdf.multi_cell(px2MM(1044), px2MM(32),csa["Commentary"].iloc[0],align='L')
    
    
    
    #//*---Credit Facilities Taken---*//
    
    pdf.set_fill_color(*hex2RGB('#313236'))
    pdf.rect(px2MM(126), px2MM(502), px2MM(248), px2MM(42),'F')
    
    pdf.set_xy(px2MM(141),px2MM(507)) 
    pdf.set_font('LeagueSpartan-Medium', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#ffffff'))
    pdf.cell(px2MM(218), px2MM(32),'Credit Facilities Taken',align='C')
    
    bl_hight = 114
    
    
    #//*---Table Header----*//
    pdf.set_fill_color(*hex2RGB('#ffffff'))
    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.set_line_width(px2MM(0.5))
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    
    #//*---Col 1
    pdf.rect(px2MM(126), px2MM(544), px2MM(290), px2MM(72),'FD')
    pdf.set_xy(px2MM(146),px2MM(564)) 
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(250), px2MM(32),'Type of Facility',align='L')
    #//*---Col 2
    pdf.rect(px2MM(416), px2MM(544), px2MM(290), px2MM(72),'FD')
    pdf.set_xy(px2MM(436),px2MM(564)) 
    pdf.cell(px2MM(250), px2MM(32),'Total Records',align='C')
    #//*---Col 3
    pdf.rect(px2MM(706), px2MM(544), px2MM(290), px2MM(72),'FD')
    pdf.set_xy(px2MM(726),px2MM(564)) 
    pdf.cell(px2MM(250), px2MM(32),'Active Accounts',align='C')
    #//*---Col 4
    pdf.rect(px2MM(996), px2MM(544), px2MM(290), px2MM(72),'FD')
    pdf.set_xy(px2MM(1016),px2MM(564)) 
    pdf.cell(px2MM(250), px2MM(32),'Closed Accounts',align='C')
    #//*---Col 5
    pdf.rect(px2MM(1286), px2MM(544), px2MM(514), px2MM(72),'FD')
    pdf.set_xy(px2MM(1306),px2MM(564)) 
    pdf.cell(px2MM(474), px2MM(32),'Accounts with Negative History',align='C')
    
    for i in range(len(type_facility)-1):
        #//*---Table Header----*//
        if i%2==0:
            pdf.set_fill_color(*hex2RGB('#F3F6F9'))
            pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        else:
            pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
            pdf.set_fill_color(*hex2RGB('#ffffff'))
            
        if i==len(type_facility)-1:
            pdf.set_fill_color(*hex2RGB('#ffffff'))
            pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
            
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.set_line_width(px2MM(0.5))
        #//*---Col 1
        pdf.rect(px2MM(126), px2MM(616+(i*52)), px2MM(290), px2MM(52),'FD')
        pdf.set_xy(px2MM(146),px2MM(626+(i*52))) 
        
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.cell(px2MM(250), px2MM(32),str(type_facility[i]),align='L')
        #//*---Col 2
        pdf.rect(px2MM(416), px2MM(616+(i*52)), px2MM(290), px2MM(52),'FD')
        pdf.set_xy(px2MM(436),px2MM(626+(i*52))) 
        pdf.cell(px2MM(250), px2MM(32),str(tot_record[i]),align='C')
        #//*---Col 3
        pdf.rect(px2MM(706), px2MM(616+(i*52)), px2MM(290), px2MM(52),'FD')
        pdf.set_xy(px2MM(726),px2MM(626+(i*52))) 
        pdf.cell(px2MM(250), px2MM(32),str(active_acc[i]),align='C')
        #//*---Col 4
        pdf.rect(px2MM(996), px2MM(616+(i*52)), px2MM(290), px2MM(52),'FD')
        pdf.set_xy(px2MM(1016),px2MM(626+(i*52))) 
        pdf.cell(px2MM(250), px2MM(32),str(clsd_acc[i]),align='C')
        #//*---Col 5
        pdf.rect(px2MM(1286), px2MM(616+(i*52)), px2MM(514), px2MM(52),'FD')
        pdf.set_xy(px2MM(1306),px2MM(626+(i*52))) 
        pdf.cell(px2MM(474), px2MM(32),str(acc_neg_hist[i]),align='C')
        
    
        bl_hight+=52
    #//*---Total----*// 
    tot_height = pdf.get_y()   
    pdf.set_fill_color(*hex2RGB('#ffffff'))
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.set_draw_color(*hex2RGB('#B9BABE'))
    pdf.set_line_width(px2MM(1))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    #//*---Col 1

    pdf.rect(px2MM(126), px2MM(mm2PX(tot_height)+43), px2MM(290), px2MM(52),'F') 
    pdf.set_xy(px2MM(146),px2MM(mm2PX(tot_height)+53)) 
    pdf.cell(px2MM(290), px2MM(32),str(type_facility[-1]),align='L')
    #//*---Col 2
    pdf.rect(px2MM(416), px2MM(mm2PX(tot_height)+43), px2MM(290), px2MM(52),'F')
    pdf.set_xy(px2MM(416),px2MM(mm2PX(tot_height)+53)) 
    pdf.cell(px2MM(290), px2MM(32),str(tot_record[-1]),align='C')
    #//*---Col 3
    pdf.rect(px2MM(706), px2MM(mm2PX(tot_height)+43), px2MM(290), px2MM(52),'F')
    pdf.set_xy(px2MM(706),px2MM(mm2PX(tot_height)+53)) 
    pdf.cell(px2MM(290), px2MM(32),str(active_acc[-1]),align='C')
    #//*---Col 4
    pdf.rect(px2MM(996), px2MM(mm2PX(tot_height)+43), px2MM(290), px2MM(52),'F')
    pdf.set_xy(px2MM(996),px2MM(mm2PX(tot_height)+53)) 
    pdf.cell(px2MM(290), px2MM(32),str(clsd_acc[-1]),align='C')
    #//*---Col 5
    pdf.rect(px2MM(1286), px2MM(mm2PX(tot_height)+43), px2MM(514), px2MM(52),'F')
    pdf.set_xy(px2MM(1286),px2MM(mm2PX(tot_height)+53)) 
    pdf.cell(px2MM(514), px2MM(32),str(acc_neg_hist[-1]),align='C')
    
    bl_hight+=52
    
    pdf.set_fill_color(*hex2RGB('#313236'))
    pdf.rect(px2MM(120), px2MM(502), px2MM(6), px2MM(bl_hight),'F')
    
            
#//*-----Liability Management 1---*//
def libility_management_1(pdf,json_data,c_MoneyS,money_signData):
    try:
        aff_check = pd.DataFrame.from_dict(json_data["affordibility_check"])
        aff_comment = pd.DataFrame.from_dict(json_data["affordibility_check_comment"])
    except:
        print('\n\n\n\n\nno page')
        return None
    
    if aff_check.empty:
        return None
    
    try:
        lib_type = aff_check["Liability Type"].tolist()
        outstanding = aff_check['Outstanding'].tolist()
        out_emi = aff_check['EMI'].tolist()
        balance = aff_check['Loan Size'].tolist()
        bal_emi = aff_check['emi'].tolist()
        

    except:
        return None
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')

    #//*----Featured List of Financial Products----*//
    pdf.set_xy(px2MM(120),px2MM(80)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(877), px2MM(84),'Liability Management',align='L')
    
    #//*---Top Black box
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84),'F')    
    
    
    pdf.set_fill_color(*hex2RGB('#313236'))
    pdf.rect(px2MM(126), px2MM(204), px2MM(224), px2MM(42),'F')
    
    pdf.set_xy(px2MM(141),px2MM(209)) 
    pdf.set_font('LeagueSpartan-Medium', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#ffffff'))
    pdf.cell(px2MM(194), px2MM(32),'Affordability Check',align='C')
    
    bl_height = 146
    #//*------Affordability Check----*//
    #//*---Col 1
    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.set_line_width(px2MM(0.5))
    pdf.set_fill_color(*hex2RGB('#ffffff'))
    pdf.rect(px2MM(126), px2MM(246), px2MM(290), px2MM(104),'FD')

    pdf.set_xy(px2MM(146),px2MM(280)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(194), px2MM(32),'Liability Type',align='L')

    #//*----Col 1/1-----*//
    pdf.rect(px2MM(416), px2MM(246), px2MM(692), px2MM(52),'FD')
    pdf.set_xy(px2MM(436),px2MM(256)) 
    pdf.cell(px2MM(652), px2MM(32),'Current Liability Distribution',align='C')
    
    #//*----Col 1/1-1-----*//
    pdf.rect(px2MM(416), px2MM(298), px2MM(346), px2MM(52),'FD')
    pdf.set_xy(px2MM(436),px2MM(308)) 
    pdf.cell(px2MM(306), px2MM(32),'Outstanding',align='C')
    
    #//*----Col 1/1-2-----*//
    pdf.rect(px2MM(762), px2MM(298), px2MM(346), px2MM(52),'FD')
    pdf.set_xy(px2MM(782),px2MM(308)) 
    pdf.cell(px2MM(306), px2MM(32),'EMI',align='C')
    
     #//*----Col 2/1-----*//
    pdf.rect(px2MM(1108), px2MM(246), px2MM(692), px2MM(52),'FD')
    pdf.set_xy(px2MM(1128),px2MM(256)) 
    pdf.cell(px2MM(652), px2MM(32),'Suggested Range',align='C')
    
    #//*----Col 2/1-1-----*//
    pdf.rect(px2MM(1108), px2MM(298), px2MM(346), px2MM(52),'FD')
    pdf.set_xy(px2MM(1128),px2MM(308)) 
    pdf.cell(px2MM(306), px2MM(32),'Balance',align='C')
    
    #//*----Col 2/1-2-----*//
    pdf.rect(px2MM(1454), px2MM(298), px2MM(346), px2MM(52),'FD')
    pdf.set_xy(px2MM(1474),px2MM(308)) 
    pdf.cell(px2MM(306), px2MM(32),'EMI',align='C')
    
    
    #//*---Table Data---*//
    
    
    rect_y = 350
    text_y = 365
    common_gap = 62
    
    for i in range(len(lib_type)-1):
        if i%2==0:
            pdf.set_fill_color(*hex2RGB('#F3F6F9'))
        else:
            pdf.set_fill_color(*hex2RGB('#ffffff'))
            
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.set_line_width(px2MM(0.5))
        pdf.rect(px2MM(126), px2MM(rect_y), px2MM(290), px2MM(62),'FD')

        #//*--Col 1---*/
        pdf.set_xy(px2MM(146),px2MM(text_y)) 
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.cell(px2MM(194), px2MM(32),str(lib_type[i]),align='L')
        
        #//*---Col 2---*/
        pdf.rect(px2MM(416), px2MM(rect_y), px2MM(346), px2MM(62),'FD')
        pdf.set_xy(px2MM(436),px2MM(text_y)) 
        pdf.cell(px2MM(306), px2MM(32),str(outstanding[i]*100)+'%',align='C')
        
        #//*---Col 3---*/
        pdf.rect(px2MM(762), px2MM(rect_y), px2MM(346), px2MM(62),'FD')
        pdf.set_xy(px2MM(782),px2MM(text_y)) 
        pdf.cell(px2MM(306), px2MM(32),str(out_emi[i]*100)+'%',align='C')
        
        #//*---Col 4---*/
        pdf.rect(px2MM(1108), px2MM(rect_y), px2MM(346), px2MM(62),'FD')
        pdf.set_xy(px2MM(1128),px2MM(text_y)) 
        pdf.cell(px2MM(306), px2MM(32),balance[i],align='C')
        
        #//*---Col 5---*/
        pdf.rect(px2MM(1454), px2MM(rect_y), px2MM(346), px2MM(62),'FD')
        pdf.set_xy(px2MM(1474),px2MM(text_y)) 
        pdf.cell(px2MM(306), px2MM(32),bal_emi[i],align='C')
        
        rect_y+=common_gap
        text_y+=common_gap
        bl_height+=common_gap
        
        
    #//*---Total-----*//
    pdf.set_fill_color(*hex2RGB('#FFFFFF'))
        
    pdf.set_draw_color(*hex2RGB('#FFFFFF'))
    pdf.rect(px2MM(126), px2MM(rect_y+1), px2MM(290), px2MM(52),'F')

    #//*--Col 1---*/
    pdf.set_xy(px2MM(146),px2MM(text_y)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(194), px2MM(32),aff_check["Liability Type"].iloc[-1],align='L')
    
    #//*---Col 2---*/
    pdf.rect(px2MM(416), px2MM(rect_y+1), px2MM(346), px2MM(52),'F')
    pdf.set_xy(px2MM(436),px2MM(text_y)) 
    pdf.cell(px2MM(306), px2MM(32),'₹  '+str(aff_check["Outstanding"].iloc[-1])+' '+aff_check["Value"].iloc[-1],align='C')
    
    #//*---Col 3---*/
    pdf.rect(px2MM(762), px2MM(rect_y+1), px2MM(346), px2MM(52),'F')
    pdf.set_xy(px2MM(782),px2MM(text_y)) 
    pdf.cell(px2MM(306), px2MM(32),'₹  '+str(aff_check["EMI"].iloc[-1])+' '+aff_check["value"].iloc[-1],align='C')
    
    #//*---Col 4---*/
    pdf.rect(px2MM(1108), px2MM(rect_y+1), px2MM(346), px2MM(52),'F')
    pdf.set_xy(px2MM(1128),px2MM(text_y)) 
    pdf.cell(px2MM(306), px2MM(32),str(aff_check["Loan Size"].iloc[-1]),align='C')
    
    #//*---Col 5---*/
    pdf.rect(px2MM(1454), px2MM(rect_y+1), px2MM(346), px2MM(52),'F')
    pdf.set_xy(px2MM(1474),px2MM(text_y)) 
    pdf.cell(px2MM(306), px2MM(32),str(aff_check["emi"].iloc[-1]),align='C')
    
    #//*---Long Black vertical line
    pdf.set_fill_color(*hex2RGB('#313236'))
    pdf.rect(px2MM(120), px2MM(204), px2MM(6), px2MM(bl_height+53),'F')
    
    comment_y = pdf.get_y()
    
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(40))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.set_xy(px2MM(120),px2MM(mm2PX(comment_y)+122)) 
    pdf.cell(px2MM(170), px2MM(56),'Comments',align='L')
    
    for_stat = 682
    statement = ['Your EMI Burden Ratio (EMI by gross monthly income) is very high at 121%.','Consider bringing down your liabilities and EMI to our suggested range.']
    try:
        statement = aff_comment['Comments:'].tolist()
        
    except:
        print('no data')
        return None
    
    for i in range(len(statement)):        
        pdf.set_fill_color(*hex2RGB('#000000'))
        pdf.rect(px2MM(120), px2MM(for_stat+20), px2MM(10), px2MM(10),'F')
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.set_xy(px2MM(150),px2MM(for_stat)) 
        pdf.cell(px2MM(1304), px2MM(42),statement[i],align='L')
        
        for_stat+=52

#//*-----Liability Management 2---*//
def libility_management_2(pdf,json_data,c_MoneyS,money_signData):
    try:
        rate_reduct = pd.DataFrame.from_dict(json_data["rate_reduction_oppurtunities"])
    except:
        print('\n\n\n\n\nno page')
        return None
    
    if rate_reduct.empty:
        return None
    
    try:
        liability = rate_reduct['Liability'].tolist()
        pend_tenure = rate_reduct["Pending Tenure (Months)"].tolist()
        out_amt = rate_reduct['Outstanding Amount'].tolist()
        emi = rate_reduct['EMI'].tolist()
        est_int_rate = rate_reduct['Estimated Interest Rate'].tolist()
        scope_red = rate_reduct['Scope of Rate Reduction'].tolist()
        poss_saving = rate_reduct['Possible Savings'].tolist()
        rep_priorty = rate_reduct['Repayment Priority'].tolist()
    except:
        return None
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')

    #//*----Featured List of Financial Products----*//
    pdf.set_xy(px2MM(120),px2MM(80)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(877), px2MM(84),'Liability Management',align='L')
    
    #//*---Top Black box
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84),'F')    
    
    
    pdf.set_fill_color(*hex2RGB('#313236'))
    pdf.rect(px2MM(126), px2MM(204), px2MM(327), px2MM(42),'F')
    
    pdf.set_xy(px2MM(141),px2MM(209)) 
    pdf.set_font('LeagueSpartan-Medium', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#ffffff'))
    pdf.cell(px2MM(297), px2MM(32),'Rate Reduction Opportunities',align='C')  
    
    #//*---Table Heading----*//
    
    #//*---Col 1
    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.set_line_width(px2MM(0.5))
    pdf.set_fill_color(*hex2RGB('#ffffff'))
    pdf.rect(px2MM(126), px2MM(246), px2MM(240), px2MM(104),'FD')
    
    pdf.set_xy(px2MM(146),px2MM(266)) 
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(200), px2MM(32),'Liability',align='L')  
    
    #//*---Col 2
    pdf.rect(px2MM(366), px2MM(246), px2MM(200), px2MM(104),'FD')
    pdf.set_xy(px2MM(380),px2MM(266)) 
    pdf.multi_cell(px2MM(170), px2MM(32),'Pending Tenure (Months)',align='C')
    
    #//*---Col 3
    pdf.rect(px2MM(566), px2MM(246), px2MM(200), px2MM(104),'FD')
    pdf.set_xy(px2MM(586),px2MM(266)) 
    pdf.multi_cell(px2MM(160), px2MM(32),'Outstanding Amount',align='R')
    
    #//*---Col 4
    pdf.rect(px2MM(766), px2MM(246), px2MM(200), px2MM(104),'FD')
    pdf.set_xy(px2MM(786),px2MM(282)) 
    pdf.multi_cell(px2MM(160), px2MM(32),'EMI',align='C') 
    
    #//*---Col 5
    pdf.rect(px2MM(966), px2MM(246), px2MM(200), px2MM(104),'FD')
    pdf.set_xy(px2MM(986),px2MM(266)) 
    pdf.multi_cell(px2MM(160), px2MM(32),'Estimated Interest Rate',align='C')  
    
    #//*---Col 6
    pdf.rect(px2MM(1166), px2MM(246), px2MM(200), px2MM(104),'FD')
    pdf.set_xy(px2MM(1186),px2MM(266)) 
    pdf.multi_cell(px2MM(160), px2MM(32),'Scope of Rate Reduction',align='C')         
        
        
    #//*---Col 7
    pdf.rect(px2MM(1366), px2MM(246), px2MM(230), px2MM(104),'FD')
    pdf.set_xy(px2MM(1386),px2MM(282)) 
    pdf.multi_cell(px2MM(190), px2MM(32),'Possible Savings',align='C')
    
    #//*---Col 8
    pdf.rect(px2MM(1596), px2MM(246), px2MM(204), px2MM(104),'FD')
    pdf.set_xy(px2MM(1616),px2MM(266)) 
    pdf.multi_cell(px2MM(164), px2MM(32),'Repayment Priority',align='C')  
    
    bl_height=146
    
    
    
    rect_y = 350
    text_y = 365
    common_gap = 62
    
    
    #//*---Table value----*//
    
    for i in range(len(liability)):
        
        #//*---Col 1
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.set_line_width(px2MM(0.5))
        if i%2==0:
            pdf.set_fill_color(*hex2RGB('#F3F6F9'))
        else:
            pdf.set_fill_color(*hex2RGB('#ffffff'))
        pdf.rect(px2MM(126), px2MM(rect_y), px2MM(240), px2MM(62),'FD')
        
        pdf.set_xy(px2MM(146),px2MM(text_y)) 
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#1A1A1D'))
        pdf.cell(px2MM(200), px2MM(32),liability[i],align='L')

        #//*----Col 2
        pdf.rect(px2MM(366), px2MM(rect_y), px2MM(200), px2MM(62),'FD')
        pdf.set_xy(px2MM(386),px2MM(text_y))
        pdf.cell(px2MM(160), px2MM(32),str(pend_tenure[i]),align='C')
        
        #//*----Col 3
        pdf.rect(px2MM(566), px2MM(rect_y), px2MM(200), px2MM(62),'FD')
        pdf.set_xy(px2MM(586),px2MM(text_y))
        pdf.cell(px2MM(160), px2MM(32),str(out_amt[i]),align='R')
        
        #//*----Col 4
        pdf.rect(px2MM(766), px2MM(rect_y), px2MM(200), px2MM(62),'FD')
        pdf.set_xy(px2MM(786),px2MM(text_y))
        pdf.cell(px2MM(160), px2MM(32),str(emi[i]),align='R')

        #//*----Col 5
        pdf.rect(px2MM(966), px2MM(rect_y), px2MM(200), px2MM(62),'FD')
        pdf.set_xy(px2MM(986),px2MM(text_y))
        if est_int_rate[i] =='-':
            pdf.cell(px2MM(160), px2MM(32),str(est_int_rate[i]),align='C')
        else:    
            pdf.cell(px2MM(160), px2MM(32),str(est_int_rate[i])+'%',align='C')
        
        #//*----Col 6
        pdf.rect(px2MM(1166), px2MM(rect_y), px2MM(200), px2MM(62),'FD')
        pdf.set_xy(px2MM(1186),px2MM(text_y))
        if scope_red[i]=='-':
            pdf.cell(px2MM(160), px2MM(32),str(scope_red[i]),align='C') 
        else:        
            pdf.cell(px2MM(160), px2MM(32),str(scope_red[i])+'%',align='C') 
        
        #//*----Col 7
        pdf.rect(px2MM(1366), px2MM(rect_y), px2MM(230), px2MM(62),'FD')
        pdf.set_xy(px2MM(1386),px2MM(text_y))
        pdf.cell(px2MM(190), px2MM(32),poss_saving[i],align='C') 
        
        #//*----Col 8
        pdf.rect(px2MM(1596), px2MM(rect_y), px2MM(204), px2MM(62),'FD')
        pdf.set_xy(px2MM(1616),px2MM(text_y))
        if rep_priorty[i] == '-':
            pdf.cell(px2MM(164), px2MM(32),rep_priorty[i],align='C')     
        else :  
            pdf.cell(px2MM(164), px2MM(32),str(int(rep_priorty[i])),align='C')  
        
        rect_y+=common_gap
        text_y+=common_gap
        bl_height+=common_gap
        
    
    #//*---Long Black vertical line
    pdf.set_fill_color(*hex2RGB('#313236'))
    pdf.rect(px2MM(120), px2MM(204), px2MM(6), px2MM(bl_height),'F')
        
        
#//*----Page Written By Gurutirth 


#//*----Your Financial Profile---*//  
#//*----Your Financial Profile---*//  
def fin_profile(pdf, json_data,c_MoneyS,money_signData,user_data):
     # //*---User Values---*//

    money_sign_desc = {
        "Eagle":"Far-Sighted Eagle", "Horse":"Persistent Horse",
        "Tiger":"Tactical Tiger", "Lion":"Opportunistic Lion",
        "Elephant":"Virtuous Elephant", "Turtle":"Vigilant Turtle",
        "Whale":"Enlightened Whale", "Shark":"Stealthy Shark"
    }
    

    #  generation
    try:
        # df = pd.DataFrame.from_dict(json_data["Generation Profile"])
        fin_score=json_data['Score'][0]['C_Score']
    except:
        return None
    gen_profile = json_data['Genration'][0]["Gen Profile"] 
    if gen_profile=='Gen 1':
        generation = 'Generation 1'
         # card 3 data
        generation_desc = "Determined individual who, despite limited education, provides for family's basic necessities and is the primary bread-earner."

    elif gen_profile=='Gen 2':
        generation = 'Generation 2'
         # card 3 data
        generation_desc = "Skilled professional with a steady income and cautious outlook, who wants to improve the standard of living."

    elif gen_profile=='Gen 3':
        generation = 'Generation 3'
         # card 3 data
        generation_desc = "Progressive individual with a large financial safety net as a result of self-accomplishment or that of the prior generation(s)."

    else:
        generation = 'None'
         # card 3 data
        generation_desc = "No Data"
        
    # card 4 data
    age_range = json_data['Life Stage'][0]['Life Stage']
    
    if age_range =='26-35':
        phase = "Building phase"
        life_stage_pts = ['Focusing on building a strong foundation', 
        'Growing professional skills and expertise',
        'Taking on more responsibilities',
        'Building financial stability']
    elif age_range =='36-45':
        phase = "Growth phase"
        life_stage_pts = ['Consolidating professional knowledge', 
        'Exploring opportunities to broaden skills and experience',
        'Building wealth by taking calculated risks',
        'Balancing work and family life']
    elif age_range =='46-55':
        phase = " Sustainability phase"
        life_stage_pts = ['Sustaining professional growth', 
        'Continual learning and development',
        'Building a strong financial base',
        'Preparing for retirement']
    elif age_range =='56-60':
        phase = "Pre-Retirement phase"
        life_stage_pts = ['Reviewing retirement plan', 
        'Reducing financial risks',
        'Preparing for a transition to a more relaxed lifestyle',
        'Reviewing estate planning and insurance needs Need age range corresponding to Life stage']
    else:
        phase = "Pre-Retirement phase"
        life_stage_pts = ['Reviewing retirement plan', 
        'Reducing financial risks',
        'Preparing for a transition to a more relaxed lifestyle',
        'Reviewing estate planning and insurance needs Need age range corresponding to Life stage']
        
    age_range_color = money_signData[c_MoneyS]['fin_profile'][0] 
    meter_stick_xpos_dict = {20:(0, 63), 40:(70, 137), 60:(144, 211), 80:(218, 285), 100:(292, 359)}
    meter_img_dict = {20:'meter_1_20.png', 40:'meter_20_40.png', 
    60:'meter_40_60.png', 80:'meter_60_80.png', 100:'meter_80_100.png'}
    for val in meter_img_dict:
        if fin_score <= val:
            meter_img = meter_img_dict[val]
            meter_stick_xpos = (meter_stick_xpos_dict[val][1] - meter_stick_xpos_dict[val][0])/20*(fin_score - (val-20))
            meter_stick_xpos = (201 + meter_stick_xpos_dict[val][0] + meter_stick_xpos)
            if fin_score <= 20:
                score_box_xpos = 190
            elif fin_score >=81:
                score_box_xpos = 416
            else:
                score_box_xpos = meter_stick_xpos - 74
            break
    
    your_money_sign = c_MoneyS.capitalize()

   
   
    #//*---Page setup----*//
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')

    vl_color = money_signData[c_MoneyS]['content'][2]
    # purple rectangle
    pdf.set_fill_color(*hex2RGB(vl_color))
    pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84), 'F')
    

    # cards
    ms_tm = 'MoneySign<sup>TM</sup>'
    card_titles = ['Financial Behaviour Score', 'MoneySign', 'Generation Profile']
    for card_num in range(3):
        # card background
        pdf.set_fill_color(*hex2RGB('#FCF8ED'))
        if card_num == 2:
            card_height = 309
        else:
            card_height = 786
        pdf.rect(px2MM(120+(card_num*577)), px2MM(214), px2MM(527), px2MM(card_height), 'F')

        if card_num == 1:
        # card 2 background
            pdf.image(join(cwd,'assets','images','MoneySign',f'{your_money_sign}_text.svg'),px2MM(697), px2MM(216), px2MM(527), px2MM(784))
            pdf.image(join(cwd,'assets','images','MoneySign','cream_bg_mask.png'),px2MM(697), px2MM(216), px2MM(527), px2MM(784))
            pdf.image(join(cwd,'assets','images','MoneySign',f'{your_money_sign}.svg'),px2MM(810), px2MM(422), px2MM(300), px2MM(300))
            # black boxes to hide your_money_sign_bg.svg vertical overflow
            pdf.set_fill_color(*hex2RGB('#000000'))
            pdf.rect(px2MM(697), px2MM(0), px2MM(527), px2MM(216), 'F')
            pdf.rect(px2MM(647), px2MM(0), px2MM(50), px2MM(1080), 'F')
            pdf.rect(px2MM(697), px2MM(1000), px2MM(527), px2MM(80), 'F')

        # card titles
        pdf.set_xy(px2MM(168+(card_num*577)), px2MM(254))  
        pdf.set_font('LeagueSpartan-Medium', size=px2pts(40))
        pdf.set_text_color(*hex2RGB('#000000'))
        # if card_num==1:
        #     pdf.cell(px2MM(431), px2MM(56),pdf.write_html(ms_tm), align='C')
        # else:

        pdf.cell(px2MM(431), px2MM(56), card_titles[card_num], align='C')
        
        pdf.set_xy(px2MM(1048), px2MM((266)))
        pdf.set_font('LeagueSpartan-Medium', size=16)
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.cell(px2MM(16), px2MM(8), 'TM')

    # ------------- meter ----------------
    
    # fin score stick
    pdf.set_fill_color(*hex2RGB("#FFFFFF"))
    # pdf.set_draw_color(*hex2RGB("#E9EAEE"))
    pdf.rect(px2MM(meter_stick_xpos), px2MM(492), px2MM(6), px2MM(95), 'F')
    # fin score box
    pdf.rect(px2MM(score_box_xpos), px2MM(380), px2MM(160), px2MM(148), 'F')

    # fin score 
    pdf.set_xy(px2MM(score_box_xpos), px2MM(416))  
    pdf.set_font('Prata', size=px2pts(64))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(160), px2MM(87), str(fin_score), align='C')

    # Actual meter image
    pdf.image(join(cwd,'assets','images','BehaviourMeter', meter_img),
    px2MM(190), px2MM(575), px2MM(386), px2MM(74))

    # -------------meter labels------------

    # 0 label
    pdf.set_xy(px2MM(190), px2MM(669))  
    pdf.set_font('LeagueSpartan-semiBold', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#898B90'))
    pdf.multi_cell(px2MM(40), px2MM(32), '0', align='L')

    # 100 label
    pdf.set_xy(px2MM(540), px2MM(669))  
    pdf.set_font('LeagueSpartan-semiBold', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#898B90'))
    pdf.multi_cell(px2MM(50), px2MM(32), '100', align='L')

    # card 1 footer
    # card footer range
    pdf.set_font('LeagueSpartan-Medium', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.set_xy(px2MM(160), px2MM(761))
    pdf.cell(px2MM(63), px2MM(32), '0-50 : ', align='L')
    pdf.set_xy(px2MM(160), px2MM(808))
    pdf.cell(px2MM(74), px2MM(32), '50-75 : ', align='L')
    pdf.set_xy(px2MM(160), px2MM(855))
    pdf.cell(px2MM(83), px2MM(32), '75-100 : ', align='L')
    # card footer range descriptions
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.set_xy(px2MM(233), px2MM(761))
    pdf.cell(px2MM(213), px2MM(32), 'Financially vulnerable', align='L')
    pdf.set_xy(px2MM(244), px2MM(808))
    pdf.cell(px2MM(175), px2MM(32), 'Financially coping', align='L')
    pdf.set_xy(px2MM(253), px2MM(855))
    pdf.cell(px2MM(131), px2MM(32), 'Financially fit', align='L')

    # card 2 footer
    pdf.set_xy(px2MM(766), px2MM(822))
    pdf.set_font('Prata', size=px2pts(42))
    pdf.set_text_color(*hex2RGB('#000000'))
    # pdf.cell(px2MM(400), px2MM(66), money_sign_desc[your_money_sign], align='C')
    pdf.cell(px2MM(400), px2MM(66), user_data['moneySign'], align='C')

    # card 3 content
    # --Titles
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    # generation 
    pdf.set_xy(px2MM(1313), px2MM(330))
    pdf.cell(px2MM(447), px2MM(42), generation)

    # content
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#313236'))
    pdf.set_xy(px2MM(1313), px2MM(387))
    pdf.multi_cell(px2MM(447), px2MM(32), 
    generation_desc, 
    align='L')

    # -----card 4
    # background
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(px2MM(1273), px2MM(563), px2MM(527), px2MM(437), 'F')

    # title
    pdf.set_xy(px2MM(1454), px2MM(603))  
    pdf.set_font('LeagueSpartan-Medium', size=px2pts(40))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(164), px2MM(56), 'Life stage', align='C')

    # subtitle
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
    pdf.set_text_color(*hex2RGB('#1A1A1D')) 
    pdf.set_xy(px2MM(1313), px2MM(679))
    pdf.cell(px2MM(250), px2MM(42), phase, align='L')

    # label
    pdf.set_fill_color(*hex2RGB(age_range_color))
    pdf.rect(px2MM(1605), px2MM(682.5), px2MM(166), px2MM(35), 'F')
    # label text
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(18))
    pdf.set_text_color(*hex2RGB('#000000')) 
    pdf.set_xy(px2MM(1610), px2MM(687.5))
    pdf.cell(px2MM(154), px2MM(25), f'Age Range: {age_range}', align='C')

    y_h = pdf.get_y()+13
    # bullet points
    for idx, point in enumerate(life_stage_pts):
        pdf.set_fill_color(*hex2RGB('#313236'))
        # pdf.rect(px2MM(1295), px2MM(mm2PX(y_h)+15), px2MM(5), px2MM(5), 'F')
        pdf.circle(x=px2MM(1333), y=px2MM(mm2PX(y_h)+14), r=px2MM(5), style='F')
        # text
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.set_text_color(*hex2RGB('#313236')) 
        pdf.set_xy(px2MM(1353), px2MM(mm2PX(y_h)))
        pdf.multi_cell(px2MM(427), px2MM(32), point,align='L')
        y_h = pdf.get_y()
        # pdf.cell(px2MM(1334), px2MM(pdf.get_y()+32), point, align='L')

    # page tile 
    pdf.set_xy(px2MM(120), px2MM(80))  
    pdf.set_font('LeagueSpartan-Medium', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.cell(px2MM(792), px2MM(84), 'Your Financial Profile')
 

#//*----Our Assumptions------*//
def assumptions(pdf,json_data,c_MoneyS,money_signData):
    try:
        df = pd.DataFrame.from_dict(json_data["Our Assumption"])
        df2 = pd.DataFrame.from_dict(json_data["OA Expected Income"])
        df3 = pd.DataFrame.from_dict(json_data["OA Expected Interest"])
        
    except:
        return None
   
    #//*---Page setup----*//
    pdf.add_page()

    # pg background color
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080), 'F')
    pdf.image(join(cwd,'assets','images','backgrounds','doubleLine.png'),px2MM(1449),px2MM(0),px2MM(471),px2MM(1080))
    # black rectangle besides title
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(80), 'F')

    # Page title
    pdf.set_xy(px2MM(120), px2MM(80))  
    pdf.set_font('LeagueSpartan-Medium', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.multi_cell(px2MM(441), px2MM(84), 'Our Assumptions', align='L')

    # ------cards--------------
    pdf.set_fill_color(*hex2RGB('#FFFFFF'))
    # card 1: Asset class risk level table  card
    pdf.rect(px2MM(120), px2MM(204), px2MM(820), px2MM(700), 'F')
    # card 2: income/expense table card
    pdf.rect(px2MM(980), px2MM(184), px2MM(820), px2MM(462), 'F')
    # card 3: Interest rate table
    pdf.rect(px2MM(980), px2MM(676), px2MM(540), px2MM(357), 'F')

    # card 2 title
    pdf.set_xy(px2MM(1020), px2MM(224))  
    pdf.set_font('LeagueSpartan-Medium', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.multi_cell(px2MM(400), px2MM(32), 'Expected Income/Expense YoY Growth', align='L')

    # card 2 title
    pdf.set_xy(px2MM(1020), px2MM(716))  
    pdf.set_font('LeagueSpartan-Medium', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.multi_cell(px2MM(400), px2MM(32), 'Expected Interest Rates on Liabilities', align='L')

    # ---------tables-------------------------
    # --card 1 table--
    # asset class table title row
    pdf.set_fill_color(*hex2RGB('#F3F6F9'))
    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.rect(px2MM(160), px2MM(244), px2MM(177), px2MM(45), 'DF')
    pdf.rect(px2MM(337), px2MM(244), px2MM(248), px2MM(45), 'DF')
    pdf.rect(px2MM(585), px2MM(244), px2MM(119), px2MM(45), 'DF')
    pdf.rect(px2MM(704), px2MM(244), px2MM(196), px2MM(45), 'DF')  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(18))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.set_xy(px2MM(180),px2MM(254))
    pdf.cell(px2MM(137), px2MM(25),'Asset Classes',border=0,align='L')
    pdf.set_xy(px2MM(357),px2MM(254))
    pdf.cell(px2MM(208), px2MM(25),'Examples',border=0,align='L')
    pdf.set_xy(px2MM(605),px2MM(254))
    pdf.cell(px2MM(79), px2MM(25),'Returns %',border=0,align='L')
    pdf.set_xy(px2MM(724),px2MM(254))
    pdf.cell(px2MM(156), px2MM(25),'Risk Level',border=0,align='C')

    instrument = []
    risk_images = []
    for i in range(len(df)):
        txt = df["Instrument 1"][i]+'\n'+df["Instrument 2"][i]+'\n'+df["Instrument 3"][i]
        instrument.append(txt)
        
        if df["Risk Level"][i]=="Moderate to High":
            risk_images.append('Riskmeter_m2h.png')
        elif df["Risk Level"][i]=="Low to High":
            risk_images.append('Riskmeter_l2h.png')
        elif df["Risk Level"][i]=="Very Low to Moderate":
            risk_images.append('Riskmeter_vl2m.png')
        elif df["Risk Level"][i]=="Low to Very High":
            risk_images.append('Riskmeter_l2vh.png')
              
    table1_col_vals = [list(df['Asset Classes']),instrument,list(df["Returns %"]),list(df['Risk Level'])]
    # risk_images = ['Riskmeter_m2h.png', 'Riskmeter_l2h.png', 'Riskmeter_l2h.png', 'Riskmeter_vl2m.png', 'Riskmeter_l2vh.png']

    pdf.set_fill_color(*hex2RGB('#F3F6F9'))
    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    col_x_pos = (160, 337, 585, 704)
    col_text_y_pos = (334, 309, 334, 359)
    col_widths = (177, 248, 119, 196)
    col_text_widths = (137, 218, 100, 166)
    for row in range(5):
        for column in range(4):
            # backgrounds
            pdf.set_draw_color(*hex2RGB('#E9EAEE'))
            if row%2 == 0:
                pdf.set_fill_color(*hex2RGB('#ffffff')) 
                pdf.rect(px2MM(col_x_pos[column]), px2MM(289+(row*115)), px2MM(col_widths[column]), px2MM(115), 'DF')
            else:
                pdf.set_fill_color(*hex2RGB('#F3F6F9'))
                pdf.rect(px2MM(col_x_pos[column]), px2MM(289+(row*115)), px2MM(col_widths[column]), px2MM(115), 'DF')
            
            # text weigth
            if column == 0 or column == 2:
                pdf.set_font('LeagueSpartan-Regular', size=px2pts(18))
            else:
                pdf.set_font('LeagueSpartan-Light', size=px2pts(18))

            # text color
            pdf.set_text_color(*hex2RGB('#000000'))

            # text positions
            pdf.set_xy(px2MM((col_x_pos[column])+20),px2MM(col_text_y_pos[column]+row*115))
            
            # text cells
            if column == 3:
                pdf.multi_cell(px2MM(col_text_widths[column]), px2MM(25),table1_col_vals[column][row],border=0,align='C')
            else:
                pdf.multi_cell(px2MM(col_text_widths[column]), px2MM(25),table1_col_vals[column][row],border=0,align='L')
            # Risk Images
            pdf.image(join(cwd,'assets','images','RiskMeters',risk_images[row]), px2MM(763), px2MM(309+row*115), px2MM(78), px2MM(40))
                
    # --card2 table--
    
    col_x_pos = (1020, 1175, 1299)
    col_widths = (155, 124, 156)
    col_text_widths = (115, 84, 116)
    col_align = ('L', 'C', 'R')
    # table2_col_vals = [
    #     ['Lifestage', 'Building', 'Growth', 'Sustainability', 'Pre-Retirement'],
    #     ['Age Range', '26-35', '36-45', '46-55', '55-60'],
    #     ['Income Growth', '15%', '20%', '5%', '-20%']
    # ]
    l1 = ['Lifestage']+list(df2["Expected Income Growth"])
    l2 = ['Age Range']+list(df2["Age Range"])
    l3 = ['Income Growth']+list(str(x*100)+'%' for x in df2["Percentage"])

    table2_col_vals = [
        l1,
        l2,
        l3
    ]
    
    for i in range(len(l1)-1):
        if i%2==0:
            pdf.set_fill_color(*hex2RGB('#F3F6F9'))
        else:
            pdf.set_fill_color(*hex2RGB('#FFFFFF'))
        if i==0:
            pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(18))
        else:
            pdf.set_font('LeagueSpartan-Regular', size=px2pts(18))
            
 
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.set_text_color(*hex2RGB('#000000'))
        
        pdf.rect(px2MM(1020),px2MM(276+(i*45)),px2MM(290),px2MM(45),'FD')
        pdf.set_xy(px2MM(1040),px2MM(286+(i*45)))
        pdf.cell(px2MM(250),px2MM(25),l1[i],align='L')
        
        pdf.rect(px2MM(1310),px2MM(276+(i*45)),px2MM(258),px2MM(45),'FD')
        pdf.set_xy(px2MM(1330),px2MM(286+(i*45)))
        pdf.cell(px2MM(218),px2MM(25),str(l2[i]),align='C')
        
        pdf.rect(px2MM(1568),px2MM(276+(i*45)),px2MM(192),px2MM(45),'FD')
        pdf.set_xy(px2MM(1588),px2MM(286+(i*45)))
        pdf.cell(px2MM(152),px2MM(25),str(l3[i]),align='R')
        
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(18))  
    pdf.set_xy(px2MM(1274),px2MM(511))
    pdf.cell(px2MM(250),px2MM(25),str(l1[-1])+': '+ str(int(l2[-1]*100))+'%',align='C')  
    
    pdf.set_font('LeagueSpartan-Light', size=px2pts(18))  
    pdf.set_xy(px2MM(1020),px2MM(556))
    pdf.multi_cell(px2MM(740),px2MM(25),'The timing of life stages varies based on profession, industry trends, career goals, and other factors, making it unique to each individual.',align='L') 
    

    # --card 3 table--
    col_x_pos = (1020, 1260)
    col_widths = (240, 220)
    col_text_widths = (200, 180)
    col_align = ('L', 'R')
    # table3_col_vals = [
    #     ['Liabilities', 'Home Loan', 'Auto Loan', 'Personal Loan', 'Education Loan'],
    #     ['Interest Rates', '8% - 12%', '8% - 12%', '11% - 15%', '8% - 12%']
    # ]
    l1=['Liabilities']+list(df3["Expected Interest Rate"])
    l2=['Interest Rates']+list(df3["percentage"])
    
    table3_col_vals = [
      l1,
      l2
    ]

    for row in range(5):
        for column in range(2):
            # cel backgrounds
            pdf.set_draw_color(*hex2RGB('#E9EAEE'))
            pdf.set_font('LeagueSpartan-Regular', size=px2pts(18))
            if row == 0:
                pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(18))
                pdf.set_fill_color(*hex2RGB('#F3F6F9'))
            elif row%2 == 0:
                pdf.set_fill_color(*hex2RGB('#F3F6F9'))
            else:
                pdf.set_fill_color(*hex2RGB('#ffffff'))
            pdf.rect(px2MM(col_x_pos[column]), px2MM(768+(row*45)), px2MM(col_widths[column]), px2MM(45), 'DF')
            # col text
            pdf.set_xy(px2MM(col_x_pos[column]+20), px2MM(778+row*45))
            pdf.set_text_color(*hex2RGB('#000000'))
            pdf.cell(px2MM(col_text_widths[column]), px2MM(25),table3_col_vals[column][row],border=0,align=col_align[column])


#//*----Your Financial wellness plan------*//

def fin_wellness_plan(pdf,json_data,c_MoneyS,money_signData):
    try:
        df_exp_lib_manage = pd.DataFrame.from_dict(json_data["Kt Expense_lib_manage"])
        df_asset = pd.DataFrame.from_dict(json_data["Kt Asset"])
        df_expense = pd.DataFrame.from_dict(json_data["Kt Expense"])
    except:
        return None
    
   
    #//*---Page setup----*//
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F') 

    # black background of page
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(0, 0, 1920, 1080, 'F')

    # white rectangular backgrount at bottom
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(px2MM(0), px2MM(736), px2MM(1920), px2MM(344), 'F')
    
    #//*--Purple vertical line
    # pdf.set_xy(px2MM(125),px2MM(78))
    pdf.set_fill_color(*hex2RGB('#ffffff'))
    pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84), 'F') 
    
    #//*---heading statement
    pdf.set_xy(px2MM(120),px2MM(80))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#ffffff'))
    pdf.cell(px2MM(724), px2MM(84),'Your Financial Wellness Plan')
        
    # subtitle
    pdf.set_xy(px2MM(120), px2MM(244))  
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(40))
    pdf.set_text_color(*hex2RGB('#ffffff'))
    pdf.cell(px2MM(234), px2MM(56), 'Key Takeaways')
    # Cards

    image_list = ['Expense.svg', 'Assets.svg', 'Shield.svg']
    card_title_list = ['Expense and Liability \nManagement', 'Asset Allocation', 'Emergency Planning']
   
    card_txt_pt_1_list =[df_exp_lib_manage["Expense and Liabiility Management"][0],
                         df_asset["Asset Allocation"][0],
                         df_expense["Emergency Planning"][0]]
    
    card_txt_pt_2_list =[df_exp_lib_manage["Expense and Liabiility Management"][1],
                         df_asset["Asset Allocation"][1],
                         df_expense["Emergency Planning"][1]]
    len_p = []
    for card_num in range(3):
        # Card Boxes
        pdf.set_fill_color(*hex2RGB('#FFFFFF'))
        pdf.set_draw_color(*hex2RGB('#D3D3D3'))
        pdf.rect(px2MM(120+(card_num*577)), px2MM(340), px2MM(527), px2MM(654), 'F')
        pdf.rect(px2MM(120+(card_num*577)), px2MM(340), px2MM(527), px2MM(654), 'D')
    
        # logo 
        logo = join(cwd,'assets','images','icons', image_list[card_num])
        pdf.image(logo, px2MM(160+(card_num*577)), px2MM(382), px2MM(80), px2MM(80))

        # Card titles  
        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
        pdf.set_text_color(*hex2RGB('#000000'))
        if card_num == 0:
            pdf.set_xy(px2MM(260), px2MM(380))
            pdf.multi_cell(px2MM(347), px2MM(42), card_title_list[card_num], align="L")
        else:
            pdf.set_xy(px2MM(260+(card_num*579)), px2MM(399))
            pdf.cell(px2MM(347), px2MM(42), card_title_list[card_num])  

        pdf.set_fill_color(*hex2RGB('#000000'))
        pdf.rect(px2MM(160+(card_num*576)), px2MM(522), px2MM(10), px2MM(10), 'F')

        pdf.set_xy(px2MM(195+(card_num*577)), px2MM(504))  
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
        pdf.set_text_color(*hex2RGB('#000000'))
        pdf.multi_cell(px2MM(417), px2MM(42), card_txt_pt_1_list[card_num], align='L', new_y='NEXT')
        
        len_p1 = len(card_txt_pt_1_list[card_num])/29
        if len_p1>int(len_p1):
            len_p1+=1
        len_p.append(len_p1)

    
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(px2MM(160), px2MM(522+(len_p[0]*40)), px2MM(10), px2MM(10), 'F')
    pdf.rect(px2MM(737), px2MM(522+(len_p[1]*40)), px2MM(10), px2MM(10), 'F')
    pdf.rect(px2MM(1313), px2MM(522+(len_p[2]*40)), px2MM(10), px2MM(10), 'F')

    pdf.set_xy(px2MM(190), px2MM(504+(len_p[0]*40)))  
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.multi_cell(px2MM(417), px2MM(42), card_txt_pt_2_list[0], align='L')

    pdf.set_xy(px2MM(767), px2MM(504+(len_p[1]*40)))  
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.multi_cell(px2MM(417), px2MM(42), card_txt_pt_2_list[1], align='L')

    pdf.set_xy(px2MM(1343), px2MM(504+(len_p[2]*40)))  
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.multi_cell(px2MM(417), px2MM(42), card_txt_pt_2_list[2], align='L')

#//*------cashflow_plan
def cashflow_plan(pdf,json_data,c_MoneyS,money_signData):
    try:
        df_cash_flow = pd.DataFrame.from_dict(json_data["Cash Flow Plan"])
        # df_cf_emerg_plan = pd.DataFrame.from_dict(json_data["Cf Emergency Planning"])
        # df_cf_asset_alloc = pd.DataFrame.from_dict(json_data["Cf Asset Allocation"])
        # df_cf_asset_lib_alloc = pd.DataFrame.from_dict(json_data["Cf Asset Lib Allooc"])
    except:
        return None
    
    lcol_val_list = ["Next 3M Cashflows"]+list(df_cash_flow["Next 3M Cashflows"])
    rcol_val_list = ["Amt"]+list('₹'+str(x) for x in df_cash_flow["Amount"])
    unit = ['']+list(x for x in df_cash_flow["Units"])

 

   
    #//*---Page setup----*//
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')
    pdf.image(join(cwd,'assets','images','backgrounds','doubleLine.png'),px2MM(1449),px2MM(0),px2MM(471),px2MM(1080))

    # purple rectangle
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84), 'F')

    # page tile 
    pdf.set_xy(px2MM(120), px2MM(80))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(792), px2MM(84), "Next 3 Months' Action Plan")

    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
    pdf.set_line_width(px2MM(0.05))
    # pdf.set_draw_color(*hex2RGB('#D3D3D3'))
    # pdf.rect(px2MM(120), px2MM(224), px2MM(516), px2MM(432), 'D')


    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#000000'))
    for row in range(len(lcol_val_list)):
        if row%2 == 0:
            pdf.set_fill_color(*hex2RGB('#F3F6F9'))
        else:
            pdf.set_fill_color(*hex2RGB('#ffffff'))
        if row == 0:
            pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
        else:
            pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))

        pdf.rect(px2MM(120), px2MM(224+(row*65)), px2MM(360), px2MM(65), 'FD')
        pdf.rect(px2MM(480), px2MM(224+(row*65)), px2MM(156), px2MM(65), 'FD')
        # col1 text
        pdf.set_xy(px2MM(140), px2MM(244+(row*65)))  
        pdf.cell(px2MM(320), px2MM(32), lcol_val_list[row], align='L')
        # col2 text
        pdf.set_xy(px2MM(500), px2MM(244+(row*65)))
        if rcol_val_list[row]=='₹0.0' or rcol_val_list[row]=='₹0':
            pdf.cell(px2MM(116), px2MM(32),'-', align='R')
        else:
            pdf.cell(px2MM(116), px2MM(32), rcol_val_list[row]+unit[row], align='R')

    # comment title
    pdf.set_xy(px2MM(716), px2MM(224))
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(40))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(170), px2MM(56), 'Comments', align='R')


    df_0 = pd.DataFrame.from_dict(json_data["Cf Emergency Planning"])
    df_1 = pd.DataFrame.from_dict(json_data["Cf Asset Allocation"])
    df_2 = pd.DataFrame.from_dict(json_data["Cf Asset Lib Allooc"])

    df_list = [df_0,df_1,df_2]

    y_hight = pdf.get_y()

    for i in range(3):
        # if i==0:
        #     top_mrgn = 80
        # else:
        #     top_mrgn = 30
        top_mrgn = 80   
        if not df_list[i].empty:
            
            pdf.set_xy(px2MM(716), px2MM(mm2PX(y_hight)+top_mrgn))
            pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(32))
            pdf.set_text_color(*hex2RGB('#000000'))
            pdf.cell(px2MM(300), px2MM(42),df_list[i].columns[0], align='L')
            
            cl_name = df_list[i].columns[0]
            h_end_high = pdf.get_y()
            pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
            pdf.set_text_color(*hex2RGB('#1A1A1D'))
            pdf.set_fill_color(*hex2RGB('#000000'))
            
            for k in range(len(df_list[i])):
                j = 0
                if k==0:
                    j = 50
                pdf.rect(px2MM(716), px2MM(mm2PX(h_end_high)+j+15), px2MM(10), px2MM(10), 'F')
                pdf.set_xy(px2MM(746), px2MM(mm2PX(h_end_high)+j))
                pdf.multi_cell(px2MM(1000), px2MM(42),df_list[i][cl_name][k] , align='L')
                
                h_end_high = pdf.get_y()
            top_mrgn = 30
        else:
            continue
                     
        y_hight = pdf.get_y()
        
        
    
    
    # #//**----Emergency Planning Comments-----*//
    # pdf.set_xy(px2MM(716), px2MM(300))
    # pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(32))
    # pdf.set_text_color(*hex2RGB('#000000'))
    # pdf.cell(px2MM(300), px2MM(42),"Emergency Planning", align='L')
    
    # pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
    # pdf.set_text_color(*hex2RGB('#1A1A1D'))
    # pdf.set_fill_color(*hex2RGB('#000000'))
    
    # emg_hed = pdf.get_y()
    
    # for i in range(len(df_cf_emerg_plan)):
    #     j = 0
    #     if i==0:
    #         j = 50
    #     pdf.rect(px2MM(716), px2MM(mm2PX(emg_hed)+j+15), px2MM(10), px2MM(10), 'F')
    #     pdf.set_xy(px2MM(746), px2MM(mm2PX(emg_hed)+j))
    #     pdf.multi_cell(px2MM(1000), px2MM(42),df_cf_emerg_plan["Emergency Planning"][i] , align='L')
        
    #     emg_hed = pdf.get_y()
        
    # emg_last_val = pdf.get_y()
     
    # #//**----Asset Allocation Comments-----*//
    # # pdf.set_xy(px2MM(716), px2MM(716))
    # pdf.set_xy(px2MM(716), px2MM(mm2PX(emg_last_val)+42))
    # pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(32))
    # pdf.set_text_color(*hex2RGB('#000000'))
    # pdf.cell(px2MM(300), px2MM(42),"Asset Allocation", align='L')
    
    # ass_hed =  pdf.get_y()
    
    
    
    # pdf.set_font('LeagueSpartan-Regular', size=px2pts(30))
    # pdf.set_text_color(*hex2RGB('#1A1A1D'))
    # pdf.set_fill_color(*hex2RGB('#000000'))
    
    # for i in range(len(df_cf_asset_alloc)):
    #     height = int(len(df_cf_asset_alloc["Asset Allocation"][i]))/60
    #     if height>int(height):
    #         height=height+1
    #     j = 0
    #     if i==0:
    #         j = 50
    #     pdf.rect(px2MM(716), px2MM(mm2PX(ass_hed)+j+15), px2MM(10), px2MM(10), 'F')
    #     pdf.set_xy(px2MM(746), px2MM(mm2PX(ass_hed)+j))
    #     pdf.multi_cell(px2MM(1000), px2MM(42),df_cf_asset_alloc["Asset Allocation"][i] , align='L')
        
    #     ass_hed =  pdf.get_y()

      
#//*-----Code written by Manjunath----*//        
#//*-----disclaimer----*//
def disclaimer(pdf,json_data,c_MoneyS,money_signData):
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(0,0,px2MM(1920),px2MM(1080),'F')

    pdf.set_xy(px2MM(140),px2MM(78))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(263), px2MM(84),"Disclaimer",border=0)
    
    # pdf.set_xy(px2MM(140),px2MM(170))  
    # pdf.set_font('Prata', size=px2pts(30))
    # pdf.set_text_color(*hex2RGB('#1A1A1D'))
    # pdf.cell(px2MM(286), px2MM(41),"As on 4th March 2022",border=0)
    
    pdf.set_xy(px2MM(142),px2MM(271))  
    pdf.set_font('LeagueSpartan-Medium', size=px2pts(36))
    pdf.set_text_color(*hex2RGB('#1A1A1D'))
    pdf.cell(px2MM(1143), px2MM(45),"The Disclaimer page should be read in conjunction with this report.",border=0)
    
    pdf.set_xy(px2MM(140),px2MM(356))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(1760), px2MM(40.32),"This report is based on the data and presumptions supplied by you (the client/ the user).",border=0)
    
    pdf.set_xy(px2MM(140),px2MM(416))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.multi_cell(px2MM(1444), px2MM(40.32),"This report is designed to assess your present financial condition and recommend planning ideas and concepts that may be beneficial. This report aims to demonstrate how well-established financial planning principles can enhance your existing financial situation. This report does not imply a recommendation of any specific method, but rather offers broad, general advice on the benefits of a few financial planning principles.",border=0)

    pdf.set_xy(px2MM(140),px2MM(556))  
    text1="""The reports give estimates based on multiple hypotheses; thus they are purely speculative and do not represent assurances of investment returns. Before making any financial decisions or adopting any transactions or plans, you should speak with your tax and/or legal counsel and solely decide on the execution and implementation. CERTIFIED FINANCIAL PLANNER    , 1 Finance Private Limited, or any of its representatives will not be liable or responsible for any losses or damages incurred by the client/user as a result of this report.\nPrices mentioned in this report may have come from sources we believe to be dependable, but they are not guaranteed. It's crucial to understand that past performance does not guarantee future outcomes and that actual results may vary from the forecasts in this report.\nUnless changes to your financial or personal situation necessitate a more frequent review, we advise that you evaluate your plan once a quarter.\nPlease be aware that some discrepancies could occur due to different calculation methods."""
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.multi_cell(px2MM(1444), px2MM(40.32),text1,border=0,align="L")
    
    pdf.set_xy(px2MM(788),px2MM(636))  
    text3="""CM"""
    pdf.set_font('LeagueSpartan-Light', size=px2pts(14))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(40), px2MM(29),text3,border=0,align="L")

    pdf.set_fill_color(*hex2RGB('#FFFFFF'))
    pdf.rect(px2MM(140),px2MM(916),px2MM(512),px2MM(42),'F')


    pdf.set_xy(px2MM(150),px2MM(926))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(18))
    pdf.set_text_color(*hex2RGB('#313236'))
    pdf.cell(px2MM(150), px2MM(22),"For any questions or queries, send an email to ",border=0,align="L")

    pdf.set_xy(px2MM(482),px2MM(926))  
    pdf.set_font('LeagueSpartan-Medium','U', size=px2pts(18))
    pdf.set_text_color(*hex2RGB('#313236'))
    pdf.multi_cell(px2MM(491), px2MM(22),"care@1finance.co.in",border=0,align="L")
    
    pdf.set_xy(px2MM(1595),px2MM(971))  
    pdf.set_font('LeagueSpartan-Regular','U', size=px2pts(20))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(185), px2MM(28),"https://1finance.co.in/",border=0)

#//*-----Def last Page
def lastpage(pdf,json_data,c_MoneyS,money_signData):
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(0,0,px2MM(1920),px2MM(1080),'F')

    pdf.image(logo,px2MM(904),px2MM(394),px2MM(104),px2MM(119.88))

    pdf.set_xy(px2MM(518),px2MM(579.27))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(40))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    text3="""Unit No. 1101 & 1102, 11th Floor, B - Wing, \nLotus Corporate Park, Goregaon (E), Mumbai-400063,"""
    pdf.multi_cell(px2MM(887),px2MM(56),text3,border=0,align="C")

    pdf.image(join(cwd,'assets','images','icons','gmail.png'),px2MM(114.22),px2MM(854.33),px2MM(28.15),px2MM(25.33))
    pdf.set_xy(px2MM(160),px2MM(849))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(25.33))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.cell(px2MM(241),px2MM(35.46),"Care@1finance.co.in",border=0,align="L")

    pdf.image(join(cwd,'assets','images','icons','globe.png'),px2MM(114.22),px2MM(905.81),px2MM(28.15),px2MM(28.15))
    pdf.set_xy(px2MM(160.66),px2MM(901.89))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(25.33))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.cell(px2MM(243),px2MM(35.46),"1finance.co.in",border=0,align="L")


    pdf.image(join(cwd,'assets','images','icons','call.png'),px2MM(114.22),px2MM(960.11),px2MM(25.33),px2MM(25.33))
    pdf.set_xy(px2MM(160.66),px2MM(954.78))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(25.33))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.cell(px2MM(183),px2MM(35.46),"022 - 6912 0000",border=0,align="L")


    pdf.set_xy(px2MM(1485),px2MM(857))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.cell(px2MM(325),px2MM(32),"Prepared by",border=0,align="R")

    pdf.set_xy(px2MM(1485),px2MM(899))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.cell(px2MM(325),px2MM(42),"Siddharth Tamboli",border=0,align="R")

    pdf.set_xy(px2MM(1485),px2MM(951))  
    pdf.set_font('LeagueSpartan-Light', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.cell(px2MM(325),px2MM(32),"Certified Financial Planner",border=0,align="R")
    
     #//*--To print superscritp TM  
    pdf.set_xy(px2MM(1804), px2MM((960)))
    pdf.set_font('LeagueSpartan-Light', size=9)
    pdf.set_text_color(*hex2RGB('#ffffff'))
    pdf.cell(px2MM(16), px2MM(8), 'CM')  

    pdf.line(110,791,1700,0)
    pdf.image(join(cwd,'assets','images','icons','Line 3.png'),px2MM(110),px2MM(791),px2MM(1700),px2MM(0.02))

#//*-----Def Your 1 view
def your_1_view_detail(pdf,json_data,c_MoneyS,money_signData,user_data):
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#000000'))
    pdf.rect(0,0,px2MM(1920),px2MM(1080),'F')
    df_asset = pd.DataFrame.from_dict(json_data['Asset'])

    # pdf.set_fill_color(*hex2RGB('#000000'))
    # pdf.rect(0,px2MM(730),px2MM(1920),px2MM(350),'F')

    pdf.set_fill_color(*hex2RGB('#FCF8ED'))
    pdf.rect(px2MM(0),px2MM(80),px2MM(15),px2MM(84),'F')

    pdf.set_xy(px2MM(120),px2MM(80))  
    pdf.set_font('LeagueSpartan-Bold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#ffffff'))
    pdf.cell(px2MM(293),px2MM(84),"Your 1 View",border=0,align="L")

    # card 1
    pdf.set_fill_color(*hex2RGB('#E6E0FF'))
    pdf.rect(px2MM(120),px2MM(204),px2MM(527),px2MM(520),'F')
    # pdf.image(join(cwd,'assets','images','1_view_table','table_bg1.png'),px2MM(120),px2MM(204),px2MM(527),px2MM(592))

    pdf.image(join(cwd,'assets','images','icons','Assets.png'),px2MM(160),px2MM(244),px2MM(60),px2MM(60))
    pdf.set_xy(px2MM(240),px2MM(246))  
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(40))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(105),px2MM(56),"Assets",border=0,align="L")

    
    tableL1=list(df_asset['Asset'])
    tableR1=list(df_asset['Amount'])
    tableU1=list(df_asset['Units'])
    
    pdf.set_xy(px2MM(500),px2MM(253))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(105),px2MM(42),'₹ '+str(tableR1[0])+tableU1[0],border=0,align="L")
    
    pdf.set_text_color(*hex2RGB('#000000'))
    for row in range(1,len(df_asset)):
        rows = row-1
        pdf.set_line_width(px2MM(0.1))
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.set_fill_color(*hex2RGB('#FFFFFF'))
        pdf.rect(px2MM(160), px2MM(324+(rows*72)), px2MM(290), px2MM(72), 'DF')
        pdf.rect(px2MM(450), px2MM(324+(rows*72)), px2MM(157), px2MM(72), 'DF')
        
        pdf.set_xy(px2MM(180), px2MM(344+(rows*72)))
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.cell(px2MM(250),px2MM(32),tableL1[row],border=0,align="L")

            #cal2 text
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.set_xy(px2MM(470), px2MM(344+(rows*72)))
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        if tableR1[row] =='-':
            pdf.cell(px2MM(117),px2MM(25),'-',border=0,align="R")
        else:
            pdf.cell(px2MM(117),px2MM(25),f"₹ {str(tableR1[row])} {tableU1[row]}",border=0,align="R")
  
     # card 2
    pdf.set_fill_color(*hex2RGB('#DEEDFF'))
    pdf.rect(px2MM(697),px2MM(558),px2MM(527),px2MM(304),'F')

    pdf.image(join(cwd,'assets','images','icons','Income.png'),px2MM(737),px2MM(598),px2MM(60),px2MM(60))
    pdf.set_xy(px2MM(817),px2MM(600))  
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(40))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(116),px2MM(56),"Income",border=0,align="L")

    # tableL2=("Regular Income","Other Income","Total")
    # tableR2=("50L","2L","52L")
    df_income = pd.DataFrame.from_dict(json_data['Income'])
    tableL2=list(df_income['Income'])
    tableR2=list(df_income['Amount'])
    tableU2=list(df_income['Units'])
    
    pdf.set_xy(px2MM(1050),px2MM(607))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(130),px2MM(42),'₹ '+str(tableR2[0])+tableU2[0],border=0,align="R")
    
    pdf.set_text_color(*hex2RGB('#000000'))
    for row in range(1,len(df_income)):
        rows = row-1
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.set_fill_color(*hex2RGB('#FFFFFF'))
        pdf.rect(px2MM(737), px2MM(678+(rows*72)), px2MM(447), px2MM(72), 'DF')
        pdf.rect(px2MM(1027), px2MM(678+(rows*72)), px2MM(157), px2MM(72), 'DF')

        # col1 text
        pdf.set_xy(px2MM(757), px2MM(698+(rows*72)))
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.cell(px2MM(250),px2MM(32),tableL2[row],border=0,align="L")

            #cal2 text
        pdf.set_xy(px2MM(1047), px2MM(698+(rows*72)))
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        if str(tableR2[row]) == '-':
            pdf.cell(px2MM(117),px2MM(32),'-',border=0,align="R")
        else:
            pdf.cell(px2MM(117),px2MM(32),f"₹ {str(tableR2[row])} {tableU2[row]}",border=0,align="R")

    pdf.set_fill_color(*hex2RGB('#FFDDDA'))
    pdf.rect(px2MM(1273),px2MM(558),px2MM(527),px2MM(304),'F')

    pdf.image(join(cwd,'assets','images','icons','Expense.png'),px2MM(1313),px2MM(598),px2MM(60),px2MM(60))
    pdf.set_xy(px2MM(1393),px2MM(600))  
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(40))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(154),px2MM(56),"Expenses",border=0,align="L")

    df_expense = pd.DataFrame.from_dict(json_data['Expense'])
    tableL3=list(df_expense['Expense'])
    tableR3=list(df_expense['Amount'])
    tableU3=list(df_expense['Units'])
    
    # tableL3=("Household + Lifestyle","Taxes","Total")
    # tableR3=("18L","13L","31L")
    
    pdf.set_xy(px2MM(1630),px2MM(607))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(130),px2MM(42),'₹ '+str(tableR3[0])+tableU3[0],border=0,align="R")
    
    pdf.set_text_color(*hex2RGB('#000000'))
    for row in range(1,len(df_expense)):
        rows = row-1
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.set_fill_color(*hex2RGB('#FFFFFF'))
        pdf.rect(px2MM(1313), px2MM(678+(rows*72)), px2MM(290), px2MM(72), 'DF')
        pdf.rect(px2MM(1603), px2MM(678+(rows*72)), px2MM(157), px2MM(72), 'DF')
        
        pdf.set_xy(px2MM(1333), px2MM(698+(rows*72)))
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.cell(px2MM(250),px2MM(32),tableL3[row],border=0,align="L")

            #cal2 text
        pdf.set_xy(px2MM(1623), px2MM(698+(rows*72)))
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        if tableR3[row] == '-':
            pdf.cell(px2MM(117),px2MM(32),"-",border=0,align="R")
        else:
            pdf.cell(px2MM(117),px2MM(32),f"₹ {str(tableR3[row])} {tableU3[row]}",border=0,align="R")
    
     # card 4
    df_insurance = pd.DataFrame.from_dict(json_data['Insurance'])
    tableL4=list(df_insurance['Insurance'])
    tableR4=list(df_insurance['Amount'])
    tableU4=list(df_insurance['Units'])
    
    pdf.set_fill_color(*hex2RGB('#FFE7CC'))
    pdf.rect(px2MM(1273),px2MM(204),px2MM(527),px2MM(304),'F')

    pdf.image(join(cwd,'assets','images','icons','Insurance.png'),px2MM(1313),px2MM(244),px2MM(60),px2MM(60))
    pdf.set_xy(px2MM(1393),px2MM(246))  
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(40))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(158),px2MM(56),"Insurance",border=0,align="L")
    
    
    pdf.set_text_color(*hex2RGB('#000000'))
    for row in range(2):
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.set_fill_color(*hex2RGB('#FFFFFF'))
        pdf.rect(px2MM(1313), px2MM(324+(row*72)), px2MM(290), px2MM(72), 'DF')
        pdf.rect(px2MM(1603), px2MM(324+(row*72)), px2MM(157), px2MM(72), 'DF')
        
        # col1 text
        pdf.set_xy(px2MM(1333), px2MM(344+(row*72)))
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.cell(px2MM(250),px2MM(32),tableL4[row],border=0,align="L")

            #cal2 text
        pdf.set_xy(px2MM(1623), px2MM(344+(row*72)))
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        if str(tableR4[row]) == '-':  
            pdf.cell(px2MM(117),px2MM(32),'-',border=0,align="R")
        else:            
            pdf.cell(px2MM(117),px2MM(32),f"₹ {str(tableR4[row])} {tableU4[row]}",border=0,align="R")

     # card 5
    df_lib_tot = pd.DataFrame.from_dict(json_data['Liability Snapshot'])
    try:
        val = str(df_lib_tot['Outstanding Amount'].iloc[-1])+str(df_lib_tot['Value'].iloc[-1])
    except:
        val = 'N/A'
    pdf.set_fill_color(*hex2RGB('#FFF3DB'))
    pdf.rect(px2MM(696),px2MM(204),px2MM(527),px2MM(304),'F')

    pdf.image(join(cwd,'assets','images','icons','Liabilities.png'),px2MM(736),px2MM(244),px2MM(60),px2MM(60))
    pdf.set_xy(px2MM(816),px2MM(246))  
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(40))
    pdf.set_text_color(*hex2RGB('#000000'))
    pdf.cell(px2MM(155),px2MM(56),"Liabilities",border=0,align="L")
    
    pdf.set_xy(px2MM(1050),px2MM(253))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(30))
    pdf.set_text_color(*hex2RGB('#000000'))
    if val == 'N/A':
        pdf.cell(px2MM(130),px2MM(42),val,border=0,align="R")
    else:
        pdf.cell(px2MM(130),px2MM(42),'₹ '+val,border=0,align="R")
    
    
    
    df_liabilities = pd.DataFrame.from_dict(json_data['Liabilities'])

    tableL5=list(df_liabilities['Liabilities'])
    tableR5=list(df_liabilities['Amount'])
    tableU5=list(df_liabilities['Units'])

    
    pdf.set_text_color(*hex2RGB('#000000'))
    for row in range(len(tableL5)):
        pdf.set_draw_color(*hex2RGB('#E9EAEE'))
        pdf.set_fill_color(*hex2RGB('#FFFFFF'))
        pdf.rect(px2MM(736), px2MM(324+(row*72)), px2MM(290), px2MM(72), 'DF')
        pdf.rect(px2MM(1026), px2MM(324+(row*72)), px2MM(157), px2MM(72), 'DF')

        #     # col1 text
        pdf.set_xy(px2MM(756), px2MM(344+(row*72)))
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        pdf.cell(px2MM(250),px2MM(32),tableL5[row],border=0,align="L")

            #cal2 text
        pdf.set_xy(px2MM(1046), px2MM(344+(row*72)))
        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
        if tableR5[row]=='-':
            pdf.cell(px2MM(117),px2MM(25),'-',border=0,align="R")
        else:
            pdf.cell(px2MM(117),px2MM(25),f'₹ {str(tableR5[row])} {tableU5[row]}',border=0,align="R")

    desc_text = '''Disclaimer: The accuracy and comprehensiveness of this information is dependent on the details provided to us. The more accurate the information, the better our financial suggestions will be.'''
    pdf.set_xy(px2MM(405), px2MM(976))
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#ffffff'))
    pdf.multi_cell(px2MM(1110),px2MM(32),desc_text,border=0,align="C")
#//*----Mutual Fundss
def mutual_funds(pdf,json_data,c_MoneyS,money_signData):
    pdf.add_page()
    pdf.set_fill_color(*hex2RGB('#222222'))
    pdf.rect(0,0,px2MM(1920),px2MM(1080),'F')
    #title
    pdf.set_xy(px2MM(120),px2MM(80))  
    pdf.set_font('LeagueSpartan-Bold', size=px2pts(60))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.cell(px2MM(346),px2MM(34),"Mutual Funds",border=0,align="L")
    #side react
    pdf.set_fill_color(*hex2RGB('#313236'))
    pdf.rect(px2MM(120),px2MM(204),px2MM(15),px2MM(258),'F')

    pdf.set_fill_color(*hex2RGB('#313236'))
    pdf.rect(px2MM(155),px2MM(204),px2MM(76),px2MM(42),'F')
    pdf.set_xy(px2MM(170),px2MM(209))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.cell(px2MM(46),px2MM(32),"Hold",border=0,align="L")


    pdf.set_fill_color(*hex2RGB('#FFFFFF'))
    pdf.rect(px2MM(155),px2MM(246),px2MM(1645),px2MM(216),'F')

    #card 1
    col_x_pos = (155,775,935,1155)
    col_widths = (620,160,220,645)
    col_text_widths = (580,120,180,605)
    col_align = ('L', 'C', 'R','L')
    table2_col_vals = [
        ['Fund Name', 'Axis Long Term Equity Fund - Direct Plan - Growth Option', 'Mirae Asset Tax Saver Fund-Direct Plan -Growth'],
        ['Ranking', '2/23', '1/23'],
        ['Current Amt', '₹ 1,19,388', '₹ 1,47,360'],
        ['Reason','Fund scored high by our algorithm.','Fund scored high by our algorithm.']
    ]
    for row in range(3):
        for column in range(4):
            # cel backgrounds
            pdf.set_draw_color(*hex2RGB('#E9EAEE'))
            if row%2 == 0:
                pdf.set_fill_color(*hex2RGB('#FFFFFF'))
            else:
                pdf.set_fill_color(*hex2RGB('#F3F6F9'))
            pdf.rect(px2MM(col_x_pos[column]), px2MM(246+(row*72)), px2MM(col_widths[column]), px2MM(72), 'DF')
            #col text
            pdf.set_xy(px2MM(col_x_pos[column]+20), px2MM(266+row*72))
            if row == 0:
                pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
            else:
                pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
            pdf.set_text_color(*hex2RGB('#1A1A1D'))
            pdf.cell(px2MM(col_text_widths[column]), px2MM(32),table2_col_vals[column][row],border=0,align=col_align[column])
    
    pdf.set_fill_color(*hex2RGB('#26A670'))
    pdf.rect(px2MM(120),px2MM(502),px2MM(15),px2MM(426),'F')

    pdf.set_fill_color(*hex2RGB('#26A670'))
    pdf.rect(px2MM(155),px2MM(502),px2MM(118),px2MM(42),'F')

    pdf.set_fill_color(*hex2RGB('#FFFFFF'))
    pdf.rect(px2MM(155),px2MM(544),px2MM(1645),px2MM(384),'F')

    pdf.set_xy(px2MM(170),px2MM(507))  
    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.cell(px2MM(88),px2MM(32),"Buy New",border=0,align="L")



    #card 2
    col_x_pos = (155,550,690,850,1010)
    col_widths = (395,140,160,160,790)
    col_text_widths = (360,110,130,130,760)
    col_align = ('L','C','C','C','L')
    table2_col_vals = [
        ['Fund Name','ICICI Prudential Nifty Index Fund - Direct Plan Cumulative Option','HDFC Index Fund-NIFTY 50 Plan - Direct Plan','Motilal Oswal Midcap 30 Fund (MOF30-Direct Plan-Growth Option'],
        ['Ranking','3/43', '2/43','1/19'],
        ['Buy Amt','₹ 1,55,000', '₹ 1,55,000','₹ 2,00,000'],
        ['SIP Amt','₹ 12,500', '₹ 12,500',' ₹ 37,500'],
        ['Reason','Large cap index funds are the most efficient and cost-effective way to gain exposure to that category.','Large cap index funds are the most efficient and cost-effective way to gain exposure to that category.','Exposure to the mid-cap category provides diversification benefits as well as the potential for strong growth. Our algorithm ranks this fund highly.']
    ]
    for row in range(4):
            for column in range(5):
                if row == 0:
                    # title row
                    pdf.set_xy(px2MM(col_x_pos[column]+20), px2MM(564))
                    pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
                    pdf.set_text_color(*hex2RGB('#1A1A1D'))
                    pdf.cell(px2MM(col_text_widths[column]), px2MM(32),table2_col_vals[column][row],border=0,align=col_align[column])
                else:
                    # content rows
                    # cel backgrounds
                    pdf.set_draw_color(*hex2RGB('#E9EAEE'))
                    if row%2 == 0:
                        pdf.set_fill_color(*hex2RGB('#FFFFFF'))
                    else:
                        pdf.set_fill_color(*hex2RGB('#F3F6F9'))
                    pdf.rect(px2MM(col_x_pos[column]), px2MM(616+((row-1)*104)), px2MM(col_widths[column]), px2MM(104), 'DF')
                    #col text
                    pdf.set_xy(px2MM(col_x_pos[column]+20), px2MM(636+(row-1)*104))
                    if row == 0:
                        pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(24))
                    else:
                        pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
                    pdf.set_text_color(*hex2RGB('#1A1A1D'))
                    pdf.multi_cell(px2MM(col_text_widths[column]), px2MM(32),table2_col_vals[column][row],border=0,align=col_align[column])


    pdf.set_xy(px2MM(120),px2MM(988))  
    pdf.set_font('LeagueSpartan-Regular', size=px2pts(24))
    pdf.set_text_color(*hex2RGB('#FFFFFF'))
    pdf.cell(px2MM(737),px2MM(32),"Disclaimer: These rankings are based on 1 Finance's proprietary algorithms.",border=0,align="L")
        
# Your 1 View
# def your_1_view(pdf,json_data,c_MoneyS,money_signData):
#     try:
#         df = pd.DataFrame.from_dict(json_data["Networth_2"])
#     except:
#         return None

#     #//*---Page setup----*//
#     pdf.add_page()
#     pdf.set_fill_color(*hex2RGB('#000000'))
#     pdf.rect(0, 0, px2MM(1920), px2MM(1080),'F')

#     # Purple rectangle
#     pdf.set_fill_color(*hex2RGB('#7C5FF2'))
#     pdf.rect(px2MM(0), px2MM(80), px2MM(15), px2MM(84), 'F')

#     # Page Title
#     pdf.set_xy(px2MM(120), px2MM(80))
#     pdf.set_font('LeagueSpartan-Medium', size=px2pts(60))
#     pdf.set_text_color(*hex2RGB('#FFFFFF'))
#     pdf.cell(px2MM(293), px2MM(84), 'Your 1 View', align='L')

#     # data lists/tuples
#     icons_list = ('Assets.svg', 'Liabilities.svg', 'Income.svg', 'Expense.svg', 'Life_insurance.svg', 'Health_insurance.svg')
#     card_topic_list = ("Assets", "Liabilities", "Income", "Expenses", "Life Insurance\nCover", "Health Insurance\nCover")
#     value_list = ('2.74Cr', '0Cr', '52L', '31L', '1Cr', '7.5L')

#     card_count = -1
#     # card background color
#     pdf.set_fill_color(*hex2RGB('#FFFFFF'))
#     for row in range(3):
#         for column in range(2):
#             card_count += 1 
#             # backgrounds
#             pdf.rect(px2MM(200+column*810), px2MM(264+row*240), px2MM(710), px2MM(180), 'F')
#             # card-logos
#             logo = join(cwd,'assets','images','icons',icons_list[card_count])
#             pdf.image(logo, px2MM(240+(column*810)), px2MM(304+(row*240)), px2MM(100), px2MM(100))
#             # cart topics
#             card_title_ypos = 790 if row == 2 else (326+(row*240))
#             pdf.set_xy(px2MM(372+(column*810)), px2MM(card_title_ypos))
#             pdf.set_font('LeagueSpartan-Regular', size=px2pts(40))
#             pdf.set_text_color(*hex2RGB('#000000')) 
#             pdf.multi_cell(px2MM(290), px2MM(42), card_topic_list[card_count], align='L')
#             # Money Amt
#             pdf.set_xy(px2MM(600+(column*810)), px2MM(326+(row*240)))
#             pdf.set_font('LeagueSpartan-SemiBold', size=px2pts(60))
#             pdf.set_text_color(*hex2RGB('#000000')) 
#             pdf.cell(px2MM(290), px2MM(42), f"₹ {value_list[card_count]}", align='R')            
        

money_sign_pdf(pdf,json_data)